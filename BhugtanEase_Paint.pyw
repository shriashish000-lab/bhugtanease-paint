# -*- coding: utf-8 -*-
"""
BHUGTANEASE CAR PAINT BILLING SOFTWARE
- Login protected with License System
- Vehicle Color Matching / Mixing System (OEM shade formula save + costing)
- Sale Bill (GST + Non-GST dono)
- Purchase Entry
- Sale History, Purchase History
- Stock Management (Liter/ML mein)
- Products, Parties, Ledger
- Color Formula Library (Vehicle OEM Shades)
- P&L Report, Sale Report
- Back date bill support
- Scrollable tables with mousewheel
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import datetime
import calendar
import os
import sys
import hashlib
import json
import io
import base64
import re
import shutil

try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False
# ─── REAL BHUGTANEASE LOGO ────────────────────────────────────────────────────
LOGO_B64 = """iVBORw0KGgoAAAANSUhEUgAAAHQAAAAcCAIAAAA/XwxHAAAMyUlEQVR42u1ZeVhUR7avunVvd0PfBmw2Zd8hajQmGkHR4AZuIy5PNIkYY8KLkzE6rgluY3ACo7jMS5gsLzHDOAHNqCAGUdwyRuISBAMhyQgaA4iILN109+3uu1TV/HERGcV8Wd/75j3OH3ynzi3Orf7VOadO/S6klII++WWE6YOgD9w+cPvkX4Xt1fp9CjEEAEDYh+B3QfSTDjRK+/D9YZGLMXY6HRBCFT0AAWIQAABjRZIkWZKwojCQceV5g5vbj8MXYwwAgBAyDPP/BVyMMULo0qXy5cuWuep5QggAgABisVhYxLi5u3sP8A8OD4+IiAwPCYnw8mLMZtegIHg/vpQCCBVZhgwD7zxSUwQh1P33f1MIAWrKqstTf8LPvdO9RK7D4ayvb+QNPCGEEKrhuNlzZg0dPCjcz98PMgbBom1oBIcP4evX5Pomy7oN7jNnAELuroxSAKGt08xxGq2r6/3+bTZrYeHBq1frkpKmjho1mhDyQ+OXEAIhhD+lIv2PZAzb23sZjVar0+nUoWDpiB33RLKHl23tKq2sEJMJyjICjEbr6grk9m0Z9oeiXKNjuvClFEAo3G4hDnuDJL/5pxxCoSBYnKLo7uaeOClpRnKyJEk1NTU5OW/evHk7Lm4UxphSihCilBJCVF1FUF0MhBAAijHpXqE6gVIKIaSUqgqEUP337oJzv0MIAIOQ7HB07M7lTe2UobIgMlCGDAdjRxumTQGEdMWHGtTqHlDaZVd3BcK7lu+Md+ZBB5Xd7mxvb+u0mLQuhpVLXqoUnWxQMGhv1RjcNf08oSdPXbTElTeaLGJmJlYwgBAQAiBUrJaOiosGvwA/f39PL+933nnHarElz5jpdIoLFqRuz95mNHquWLHK28uH4zgIIcdxKrIQQoQQy7IqUgghhNCd8OwaqnLjRmN7exuEMH3dmsJDB+9sAGAYRp2gInu/QwYhQAnSavVDBsH8XFR6DIWGsN5e+OA+8PHpu2AhBBC6ixqEXRaEupDttqjx9INaMQhBTEzUsGFDY2NjHx7yyEcHC1ZnvXYoeweT9jzrdEiEYR0WRs9TxAFDP315uWPvB3zqMxRjAEDL3/K9J0+jHKfnuMGDB3Eck5iUNHduyoSJE06dOnXkSPHqNWs7LRZJlrQ6jSg6Dxcd9vQ0jp8w0Ww2nT59uqXl9qxZMzmOO3aslGA8e84cvV5PKT18uKis7KyXl1d0dPTKlSsyMrYkJibl5e1zOuUB/f3CQsN8fH2rqj7/5JMz7e0dkyYlxsfHm83mv//94+bm5rlzU3Q63eHDReFhYSNj4xgIdIMHih7uNCzcLe15AICT4UD/AQAARXSSy587KyspxrqpU7QRkQAAxSY4PiqSmxo14ZG6hPFsPw/F6bQXFuLGBtdRYzXxcfBBpzrtIYqiUEo/OXMmIDAwNze356OXV65eu3ETKSrsDA+2bEzv3LRRGPaIEhVNh8dKEaEdLy3BikIpbX3/vbZDBymlsihSSvPz/mow8BlbMr799vr2HdkhoSFvvfUmIeTrr78KCQ2ZPn3a6lXLI8LDPI3Go0eP1Nd/OzP5VyzH5uS8UVd3JS4u1mDgPz59ilK6etWq8PCIPXv2pKam8jy/ZMkL586VTUqcEBYeHhMTM3jwQ3V1tVu3ZgUF+2dlZT7++IgB/X1rar5oaKifPXsmb+AXPP30S0tfDAwICAoMPPfpWUqpo+G6c9hQ59y5smBT2tpkm41Qiim9/UKaZXxC584dwsNRpomJcqdFcTpa56fYFj0rvPtui4/RduKYIopt81LMzy/qzMwwPzTQmpdHKaUY0/ukl7JAAQAQbH711RdeSGtqagIASJKUmf2HW1evlHAaw6wUpfQ4MyeF/fBDW/IUwWYS4sfymzIYhISPTyiVFcYZswAhkGEAAAQAnU539uyZjRs35OT8KTgoJDExUU1SSohGo01ftzEzK0uS5YqKiqCg4FWr1/AG3m63R0REzZs3HyGk0WpFUczP/2DsE6NTU1Ozs7Pd3T14Xh8XN/o3Ly4VbJbJk5POnCmLiIiklI5LGPfKK+nPPb+409pZW1cbGBi09uWXdTqNh4fHlt+/tmbN2vYOc21tLQCAYCK7uKD6JlPa4s6Z061vvwMphQCwnA78x3y3FSvB+Gma+uvEbsWtrbrzZzX9eG7hM3D7Tu3QYba9edrzZdrkOZrk2S7uemX3f2NF6bU+9F5zGQg5jt27d9+kieNOnjyu0WggZP6QvT1/b37Tkyk8w8LnFgOE3Lftorvfc92xg/Pylhsa7Bm/91i8uKtKqkUQQrtdmDJ5yvvv5+7fv7+1teXJp+abzWatVitJ0oABvkajp6+vL8sitezabDa1E1AURRBslAJRElmWHfbosIpLlTU1NSdPnjB1tIWHR1BKg4KCFEVxczMYjUaM8erVa9eufaWq6nJ9fYNWo2MgJIQINkF0isEhwe7uHj6+PgxDuzoTyEAZ0/6ehvUb9H983fWpJymEEAB+21Z2/Djpi2po7QBaDRUcbP8B8sJFQsFBJX6sK+8KvX3Ap2Wcl798/ITzjRz7oCHMs4sYBvXa7z/oQKN+fv7r129IXbhw8XPP7dq1E0I4wD/g2acXvFtyFKSnMy1N4vLlcns7PypeY/RUnKJjxTLdoKHaYcN7tmXqkaXn9SzLDn9s+MjHY7+s+erq1Tq9nmcYRn07IRRChuU49eiHALIsy7Ks2q7wegNCaPfuP3t7e2/dmlVefvGDvLwlS34NAVBkhWE4lmXV/mHbtq0p81IOFhy42dSEEMvzBoZhWJZBiKVqC0AwhECj0QAAGMQCqMju7rqBD2sfGyF+9WXHf72OFWxfly4vfVE4fYIztVFOC3kDw7J8RqZm3wEYHEiWLbd/VMT4GInDol23zuOttzzee89twSLIwF5rbi/gciwrCJaYmMi0tLT09A0lR47t3/+3Vat+CwCYmDTZW284zrLsnLnaykvCtky197Fv3649fx49MRpS2p0dlNJbLS2iKH9RXVNZWZmb++fS0mOPPvpYVFR0S8stu91561aLJEmSJAmC40bjDUkS9Xo9IfKpUyfz8vLOnfsUscyhQwU3bzb9dc9fWtta/PwCRowYGRYW2tbWCiCkgGKML1z8bOfO7d9cv1ZUVChLeGbyHN7Ai6KzpORIc3OzwyHabELzrWaMsVMSbYK9vr5eVmRq6mCsNu5mq7O01F5UhLa8yl75h2yz0sMHdB79NAkTJKBlre3OEyX2q1etqQuRQmD6KywGqKUFJc9RbFZp6W+k46XWvXubC/bf7d7+VdDmzZt7driUUh9f31GjRldXV+/YkV1bW5uUlLRs2fKSkiPFxcVJSUlDHxlaeLQ4OiHBpbyCqaqUo2Kka3Xs9ixW72KnUDd1GoSQUIoQ+qz84ocf7ouMjOrstJSVfXKl9h9Tp055LTOT5/nX39hlMBhkWfb19amurmYYZDK3BQYExMWNNpvNVdVVPM8vXPhMQ0MjpXjixEnl5ZcuX65qb+8oKDiYm5tbXFycMC4hLDT82je1zTdvDBkydEz8GG9vn/JL5VVVl381fYYgCBCCgQMfOnDggI+Pr81mDQoKvHjhgsHg3nq7JTTQX3v6DGSQaOznrPgM11RTrRZMStSPfNwJECn/zNnZyU6fJt9oUjw9UHA4OF8m7s+TL1ySx4x0SVuii4yWQkNIxQXxxAmJ0/KJiRpPL9ArjUUfLN9cu7YlY/OUKZN37tqBMc7Ly9u4aaOiKDXVn+cV7Je2ZklhweKYBGFsPB0y2BkVbX51M8aYEtKz97hfMMbdj+7RVUUQBFVxOp2UUrvdvvSlF0tKitWivGvXH/39A+rqaimlGCui6Oz2bLVaVG+iKN5Zg9ztvNs/uTPAlBJKCaUKpeSOE9ls7lJEqWs+pUpHh9Tc3DWHEEqpQohstdDvlN5ZMYxxN6tiMpny8z+4cuVKWtp/2u32umt1C55KPXn6VDBkQjK2sCYzlEQLr4PL1/BPz7+/41ObeVVRb7rdw560wz26ej1TjV99/eWY+PixY5+IGxXX1tp+/HhpSkrK+vUbuu/NKiVyz7D7qvagF/VOOXUbCQYMApSo3OpdCkLVu88VQgCED6KuvotyVOFQ745tba1Hjx6JiIw0Gj1lSY6JjKr59puwwkOanDfF4Y9oNmxyGT7i52Ige+6HClBhYcHbb79tMnf4ePumpqbOmze/55x7uKF7AP3xPGrPn9OT5XmQ5UfwuT0hbmxs4DhOr+c1HMdoOMfnVdKev7ht+p3GaAQYg1+Y68JYQYj9N6KRvy9Z3nXluIekoJRAqPZTvyjPpGb6PTv9f+1LRC/p9j2y4+eSn5Ts/36fefqk7+tvH7h94PbJD5F/Ak+xkty4jGcRAAAAAElFTkSuQmCC"""
LOGO_LOGIN_B64 = """iVBORw0KGgoAAAANSUhEUgAAASMAAABGCAYAAACHbvJwAAA4mUlEQVR42u2dd5wV1d3/3+fMzJ1b9m6jbAWkSJEiFkQEBKWIKNiiJj52Y4kaa2KNNYkpBmtixRZ7F0VFEUFEQUB678sudXu5dcr5/XELu7AIGvw9+uR+Xi9eF7gzc8+cOedzvuVzviOUUooMMsggg/9lyEwXZJBBBhkyyiCDDDLIkFEGGWSQIaMMMsgggwwZZZBBBhkyyiCDDDLIkFEGGWSQIaMMMsgggwwZZZBBBhkyyiCDDDLIkFEGGWSQIaMMMsgggwwZZZBBBhkyyiCDDDL4gdAPxEVc1/1fabwAEAIhROZJZpDBzxzi/0o9I9d1ERliyiCD/z7LSCmFEIL6+jqmTfsM5SqEFCiVNFkUCAEqcXAz+hOk+E8k/y4QKBSa1HCVQklAKISQCXJR4NgOynGwLRspJLZtYWg67fLzKGpXQNuC9gRzc/EGAokfziCDDP47LCPXdZFSsnLlCk4/bTyxuIWQMsk+pMkmRVxpEpMChEhYMoBAIKUEFFIIbNtGKYHt2LiugxWP4fF4yPJnkZeTR3Ywm149e9G1axdKO3agY0kJHTodRLviIsQupvxRCSl17z8WmvdXxtLLIGMZ7Sek1PB4vLiuhtT0hEmUso725D5c3IQVZGgJe0i5xGNRbNtCAH7Th6l7aJvXhq7dulLSoZQu3bpR2qED3Xr0oG1REaamtbiqU1dPeOkSwuXl6KWl5PU79EcjpIb6Orw+Hx6PeUDJJ2Vp7u5qpv4/gwwyZLRPK8FBuSCQSfdMgFDNJllzTlAJ1811iUajKOUQ8Ps4uGtnSktK6XxQZ3r1OoTuPXvRtrCI3Db5eLVmTYzHiZaV0VRVReO2CpyKctS69cgtW5EN9biNEUIFpZh/vAt/714HjpCS16navhWp6Xhyclu1YvZqfu6jDbsTUE1NNeFwiMLCYnRdz4zSDDJktJ+eHgpQuIjkCt/cLErNsfSkdRRZgSwGHjeAAUcNoPchvSnt0JHikpIWxpRTW0Ns40aqtmwhXr4Ja9NG2FyOb2cVWnU9IlyPKWxMW+GREqRE172EN6yn9pF/Ie+9A29BwX9OSCki2raFUH09HXv0ahHz2h+rRSmVdu1aO766uopFixYy55uv2bB+PbW1NdTV1XL88aO49tobCAazMxbSTw3/ad4n8ywPPBkJIRLcI5JR6734aEIIJAKUQig45bTTGXPCCQniiYSpXrQAa81qYuWb0Ssr0SoqUNt2ooVCmK6F6cTRUJh4MJUHIXVcoZCGATYIxwU3itfUaFw8l50TJ1L6+98nvv+hhJQ8r25rBTvXraHbwGP2IATXddNkA+AqNx03E0Kg6zpSSrSka7lH8F4IJj79BA8+OAGp6+iagZQ6muZh4sTnyM3N4+qrr82M1J8a9mcRSi/Xu/6doaAf1TJqEUGiRQR7d5cOF+lxqW+q4v77/0y3LgfRtVt3Vr8/Cfv5Fyiqb8AbakRIga4b6EpiKAFCgmYCCulKhOuibAehyeRvCnAVaKA5FgEDaqd8RE337rQ988wftooliSi0cydlM2fSdcRIPKaX2uoq3nztFSpraojEYkSjMeLxOLZt4boutmOlSUfXDYJZWRQWFNGjZ08GDhxE+/bt93DvjjzyKE4/40xqautYuWIlsVgcXTMJhUNs37Hj/7Bx8TML1CfHRMPq1TRMn4mpBFITKGWjlEIqhZIaSgika6O5oJSNo0lQAuFYSD1AVLhkjTqeYJduP3qy5b/MMqIZ/zcnI7fVtcIVDqaps2HdWiY88HceeOBROg09lurPv8A/bx4+j4EQEtdNXFsIwAGRupxyk+SiQGkgJUoIhEg8cOm4+IWNsMPUvfAMnl49yO7T7/s99OSx8YYG1r/zJiWDh5LVLkEijY2NvPrqy6xatx7d40WTejIrmHBVXVxsy05YSIChe5BSYmg63Q7uxplnnsU5/3MuXq83PRlHjBzNiJGjWb16Fb/+9SVs374DTSpc10HX9P8ThNMa2fzc3M6UJdtYVsbWN16nqK6JoGsh3Bi6ACkEMQUxBJoyUEKAFkHgoKQHqQS6lYXrRIkUF2TIaDdod999993/yYOprq7ijTfewHH2FdNIfC+Fjo6BaQZZtmoVWQEfxxw/ElVcRN3C+WjhRgylkK4NQuGS1Bup5LWFiytdkAqJhkIm6E+I5DNVaErhEQLZ0EB9ZRX+wUPQTHP/zOvk5FGOzbqXnsUsLKT02OPSg8bnD3D0oMEcP2I0G8vKqK9vxOf1o+k6psckLy+PkpISevbsxaH9+hONxojFLAzDQ1VlNTOmz6CycjuDBw/G4/G0yKRFYzHeffcd6usbkJqG4zgcemhfjjvu+J/d5N1bXC11r+FwmIcensDzz0/E6/XStWu3n3xcLNU2T34+2YMGEujWBX3DanzRMF7DgyGBgJ9YSQlk5WIHfGBqeARoQmLqEql50FDoI4/De3D3DBn9eG5aghASauhUHwuEUCgUVtzGcSSO5aJJDdObxXNP/5v+vQ9l0LFDafrVr2h87AnMWBRdUzi4uK5CSJkwfxUgFI50kK4C5SAQuCJJSAqk1EG5CFcQ1Axi38yl7v33aXfuufvvrgnB5jdeIVxbRb/zLmlxnsfjoXefvvTu05fXXn+F5cuWYugalm2Rn5vLnXfdyaBBx+Dz+TBNk4ULF3Dbbbexbt06fN4AHtfD66+9SreDu3HppVe0iCm4jo1y3QShymR2spVxmopP7SuArtIJhaQT3Yo2KvX99wnKN7/u7lqy1L9Tn1u3biEWi9G5c5cWZLR8+TKeeeZpqqoqKWhfwOhRY74X0bWWxdyX9mt/+yPRRhLWdivw5uXhzcvDLiyk/r3XoBqU0LAsB+egzuTccjuYXpzGRkSoET0cJrpgLk0fv4OXKBgSwx/Y96KoVMsFNEVc+0NeLc5vHsAS+7cg796/+/u7/2GQ58CthCrxqWkSKQWOYxONhohEQ9i2RVFRAUcceTinnD6eocceg+5YNNTUcv/f/0rlju0Unf4L3OHDCTvgkHB/PFIgkuprgUQpkEogBSAchLCRwkEIFyUFLhKFBkgkgqBt477xFpG165KS8O8gpKRvWPnpR9RNn8rBZ52H7vXtZlElgtW2bSfjQy5CJIjR8Gj06NGd/Px8fD4fQgiOOOJIrrr6SjweHcuJo6RA03U+/uhDotFoIpDdbOVNTeJEINxNklNLEpJSpjNzre0LTAXVhRDpY1ubeM230DS/5nfJFVLnND++tX/X1tZw//1/5bTTx/HAg3/HsqwWpDV16ifE4zY5OXnsrKxKJwFc18FxnJZC2SRB7O2+mt9fa21PnMce5+3+G83vTQi5V9LDTTwXJxzBEQ6OcHHRcIRG1OfH07kL3m7dCBx2GP4hx+IZPQZ5yumEDR+Go6GCOZCd3ToJuO4u0klmidNEkPr7d41hpRLx0xbny5bXauYB7G0OpI/f3/N+SpaREAIhwbZsLMvGsuKUlBbTuXNnBhw5gMMOO5zS0g60a9eOnJxcotEot//+RiZNmsSi1ct54KG/8+f7JpB32aXUbNiEXLUan2Yg7Hhi0imwlUKXBiTV2yAS21CEkxg8aOlnIUUibKVLia9iMw2vv4J5862J7NreHqKUNC6cz5Znn6L4/IsIHtS1VTM6PemSVh+otLwhHo+3GNhKKQYedTSdDurE6lVr8Po0hJTU19dTU1NDcXHxbh2ZsjBkeuKlJkVqwm3atJFYLEbHDh3x+f17tRBqa2uoqamhoaEB13Xp3/+wdFYvdZzrumzfvp3Gxgbq6uro3Lkz7dsX7H31kpJIJMzadWtZv24tlVWVKKXIycmlS5eu9OzRE6/Xx5133cGk995D03XWrl1PZVUlhQWFaJpGdXU1X8yciaGbxGJR6mrrEEK0aFtr7lHqs7q6ip07d7JlSwVbt27Ba/ro06cvh/Tu3SJLuXt/VFVVJvujnkBWkF49e6XJKvVMm5oaKS8vJxgMUlraocX3za0EIQRCk+i46I6DIRWuSgQ3levsIpbk8UZOHmS3I7RlI5GiItq0yd8zbJAkHQW4DQ1YlZVEa6qxqmrQXAejpBj/IYeg7U1wm7acwGpoILZ9G3Z9PW4kBJqGJycPT7v2GO3b72n9NiNAOxIhtmULdnU1TjSKFgziLSrCLCraRYY/gpWk/+cW0a4Bo5RLIOCnW7cunDBmDEcPPJru3btjmt49Viqv18ttd91DXTjMJ9Nn8Pb7kzmy/wDOOPtXxC68gMa//BUtVI8uNUBhuy5x18GRGh6hgTBwlUAoB+kmBoEAXJEQXSqSA0YJDGlhTf+U6MiR+I8+Zs/OTAWsN5ex9aEJ+Hv0ou3ok1pfBZrlZ2XSGgSBEgIlZAsVdWpi6LqOoRuARJM6lhKYppesrECrqeDUuNxdmb1w4QLeevNNvvxyJqFQiL59+3D2L3/FiSeelD7Ptm2mfjqFr77+mtWrVlNRXkFtfT2aBhdffAlXXXU1Xq+P+vo63nr7Db6dv4ANGzaxY/t2mhobOLR/P/7057/Qs2evVmM4n3wyhZdffomlSxdTV19PPBZD03Q8Hg8ej4fDDjuM9gUFTP30UwKBbBzHZuOGDWzbupXiomJqamqYOPEpNpVtRkoNwzCpqNjK3XffSVZWENe1sS2bMSeO5fDDj0iPl82by5g9+ysWLlrImjVr2LhxI6GmUMLCcSXBrADjTx3P1Vf/lrZt2zXbO1nPlCkfMWfOHDZu3ERFRTkNDfXk5uVy9133cOKJJyGlpLGxkcmTP+D9SZNYtWoV+fk5jBw5igsuvJji4pJW+0KIhCstFEgFmnIR2InAdcq6SD5MPT8f3wXnE92wGbPzQRi7Eb5SiljlTuzFS3C/XYS2ei3RLZuJREIYNmi2RcSrERo2jOBFFxPo2q2VcSmwQiGaJn+E+9lnxDavx25swHAVrq4TMXREQRFiwDG0OfccPAXtW7iCrnKp+XwazvsfYq9YAaFGlFSYmhc3K5/I8GF4zzkTb0nxj0JI+oGyihzHIT8/l1tuuYXx48e3WOV2jzGkVuQ2bdty1133Ul15HUsWLuaJh/9Fn7796DFqJFuWLk5YM7pEWDHcgiJUUSmhTRuJNdSh2S6G7sfQJDiAkwh4yyQpJNJvCW2TYUi8tdXUv/YGnv6Ho3u9uzoz+Wk1NLDjkQlYTXUcdPFlSF3fd4crEColL0hk83a5G27aHVi8ZDFlZWUYhomrFJZlc8QRR5KdnbNXlyhlCRkeDzU11Tz51OO8+fpr1NY2YOgmUkpmTP+COXNmU1VVxXnnXQBAOBzi4YcnMHfufIJZuZimH6lrxOMWzz//HEceeSTDhh1HdXU1Dz00gR3bq8nOzkfXNAzDy9y583hm4pPc/4+HWrhstmXxyKMP8/TTzxKNRtF1nW5de9CtWzd27NjBihXLcRzB7NnzUMrC4zHxeAyKiooIBgOEwyG+/HImEx64n5UrVoPQUQKkplFdXcMLL/wbpRIub0N9HSWlpRx++BFUVVXyz38+wscff0xNTTXRaAwhJDk5iXuzLBvTMAmFojw78Rl27NjO3/52P8Fgwg3atm0Lf/3Ln9hcvo3s7LzEwmB4qdxZw333/YWBAwexevUqHnxwAgsXLEK5YBgmZY1beOKJx1m0aBEPPvRIq4SklELYAulqoASaEmgKcJwWVhFKoZle2p/yi1bjk65tse2N17HefZusLeX465twPB7i2UECSAKxJoR0UDFo/GgyNdurkXfcjq9TxxZkEq2uYueECfinTicQDWP6PajiQhxboVfWY8RDRNcvo2FzOdEBRybIyHVB03DCYaqfepLYu2+RVd9AtqYT9Wk4bhyPFcHb2Ej0xZeoXrGE3FtvJtD9wAffD1g2zVVgORY1tTXk5uTQqVMnpJQ4jtPMDxctTG7XdcnNzaVXr57Mmf0VGzZtpHLHTkaNOQFfr0OoWbkcb1kFpjSIGhIx7Fj8l1yK1asnjbjEqitR0SY0TSKEjiNNUB6kAwgbJRVCCcIaxAQ4hx5GYPDQBNE0vxfXofLJx2j86CPaXXwpeUOG7b2jxa6Q87vvvUPZ5nIMw8Sx4wSDAc4882zy89uk73f27K/529//wtat29AMjaamBnr06MYf/nAn+fltWlg+DfX1vPPu2zQ0NKJpejIG5/LuO2/y3nvv4DV9gEJqOlJqmF4f0VicdetWM2bMiQSDQaTUKGhfQK9eh2B4dLZt34aUCctN0ySjR46iS9eumKaXg7sdTMcOHSmvqCAWjaIZGrZj0blLF8aOPbkFKb7+2mv84+//QNM9CCHo0+cQHnnkEc455xxOOOEEPB6Db79dgK6bmKaXWCzKyFEjefDBhzjzzLPIzcvjpptv5Kuvvsbvz0qOA5XUy4pk8kHg2BYjR47m8iuuJBgMsm3bNu7/x9/YubMKqWl06tSJe+75I1dffTVDhw5h0aIFVNfuxOPxoOsG69evpVevXvTo0ROlFH5/gHbt29OvX19C4SaqqyrRNA0pNISSbNy4gQcemMCWigr8fm9CTqF7kJqB6fWzbv1a8vJyGTjw6JYZTSGw62uJf/Ih3pp6JB4cVxEvaY9v3CkIw7Mr1tM8+Ns8KJ1aCMNhap6ZSM7ChQR1CTm5iEsvxXvV1RjDhhHdshFtRwWa1DGkH3fzdtz27fAd1j99DaUU2yc+jvb+u+RhgJmFuOzXeK+9HuOkk3GlBouWoxkSO9eL74SxGMUlacuz+vEn0f79CjnKxfZqOCePx//7W3B79KaprBy9qQm/R4OtWwlt2Yp3yGC05ov6T8lNU8rFcRRfzZrFt/PmM378OK644jd07do1HXjcPYiaspAO7X8ot915G7fedDPTPp/GS888zSVX/Zbci35Nw4bNyKoqVFMTjZPfJdCmDbnnXkjw5FMJz/6a6CefEFvwDYGqGoRXIfDgibmAxElaLxFXwLhTaX/tDYkOTNU3cV2Qkob3JxF5/U28Rw8m9+RTvlfWLWUiSSkIhUO88O/nKSwspLa2jvLycubO+4ba2npMr4/s7GyGDB7MNb+9hm7dDk73SWvWkVKg6QYrV65kwIAj+dc/n6BDh44sWbKYfz32GA0NYTQEHo9JbV0dFRXlFBUVo2kaY048iTEnnsSbb77O/PnfosSuByWTFqvX62Xs2HEMHTqMb+bOo7KyEt3QUUogNdmCiOpqa3n1lZdwlULTBaFQmJGjjqdbt264rkt+fj5XX30Ny1esYOonn+MP+AHB+nVrMAyDvLw8otEI48edQt8+h7JixUpWrlyZHA8Cj8fD4MGDyc3NwdA1Lrr4knQsLRAIkJeXR3VVLY5jc3D3gzn11NMA6N69B6tWreAfE+5HCRehJTK15eXl6fYHAgF++ctzAIjFoixatIBc00Rqgmg0zPz58zjv/PMZNmwYAnj+hef4fNoMdN2HEBJd97Bmzaq0pZsoedMskC0kSkuMJSl0jLoGol/PRukGTl0thELQ2Ei4qhrZvTvtTjsNsVtsTHo8mP4gPukDS9LQqRNtzjgbPTcHAKuhnvDKleTEQRMC040T3bF1V8RACKLbt8Hsr8lzbSQxIrl5+IYNQysuQQPiJ4yg6sMP8FRvw/H50HJzU5OQ+k8/xXrnTXJ1BzfmYB1xNDlX/RYzvw1mn34Y7Yuw7r4Tu74Gry6wvp1L48wvyR938k+LjFqm9QV+fwDHcXjzzTeZP38uV/zmSs468+y0JbQ3Qho16gTKr9rM3+77M88+/RRHHDWQ/gOPZvtZZxN5aiJZ0oWmME2PPU6svom2l19BcOQoAoOHEJn/DdEPJxGZ9xXZDQ0YQkdpHixNEhU28tTTyf3ttWhZwV2dlyKib2bT9NQTGL4AwbN/hRHM/t4dnJq0oVCIp556MqnETmwLyc7OIRDIIhaL06ljZ8479wIOPbT/fvFcLBbl5JNP4oEHHkDXE4H3I44cwNp163jppVfIzs4mHndxXUEoHE6fm7JGE+VYFKLZ/TRP5buuSzQa240MFcpVLazf8opyKrZUoOsathXF5zcpKS5JLzKO46BpGscNP47pn89AKYVhGJRXVLB06WKOO24EHo/JZZf9BoD77/8bixcvSo+VYDCLG2+8gUMO6d0iaC+EID8/n8LCIpYsXoaUUFJS0sIN7tevPz6fH9uy0TQDx3GJxWIt+tK27bSVnhhzAtt28HlNHn30UY448sj0sTk5uSxZvIy6+lBaZhGNxXAcZ4+xq5TCRRAXAqkLpC3Qy7fQeNfd2FKhuxb+mA2OQrddOG4katy4BBk1CxPoHhNPQSFR20JzddyCIoTPm3b1PL36EM0tgi3bkIaDkHEcK9xywNTU4G+ox7RdpOYi63bgTJmM9uvLkV4femkpjUcdhrZA4el1GFpRguytpgbq3n2DnHA9yhTEBGSdchpmfpv0pX3HHIM8+hjEpPcQARMzEiE2fwHuSScd0FI6BzSb1jw16vdnsXlzBXfecQeLFy3kiiuupGPHTq0qclOD7/wLL6KsbBPPPfkk/3rwQR566mnyzzyTqkVL8M2cQbbHQMOh4ZU3qPEFyT/vHKTPR2DocLwDByFnzyT66iuI+YtRrkW918Q8/ZfkXX4leiCQUG8Lmc6cRTauI/zoP9B3VuCcejbZAwb8IKZPtd/r9TLmhBPJzc1DKUU4FGbpsmWsX7cej+ll0cLFXHfd9Rx//DB+97vfpeMQLV3AXWTkug5t2+aj60Y63a1pGj169MDj0dMTRCkXK261aM8eafq96HI0TbYIlDdXe6SaY1lxXNdBSIXUwHEUlmXtET9p0yYfKcFNkpmuaWmxa/PfUKhE4BdwHAshVKt6Jdd18XhMBh41kDWr1+Dz+zj00P4tXH1/wJ9wQ9Xed36l+gMh0llYx7HxmEE6dOyAUgrbttE0jeKSYkpLS9m5czGBQBYIcBx7r9ayQuJquzRwuusirRiuplBuHOG6uJpGVIDmMRLjr5UsljngcJq+nYNwwDjq8F1ZXyGQphflDYCb3JaubITrtIy3ZGcT9/qxhcSQEsOOY7/ybxp2VuE76xw8fQ7h4D/cjVNdDX4felLnFF29Crl2NT7dRMRi6IEsxMqVNNXXE7PCONEw3rom/JvKkELHFiCkwq6uwYnHkAfQVTsg20FUMrYhNZEoIaIUjmvj9fpwXYeXX36Zb+fP56abb2XEiJEtVt3mk1nXda674UbWb1jPZ5/P4MVnnuGKa68lcMlFhNaswLt9Gz7ThxF3qX/meRpzA+ScfgY4LprHJHfYKJzDBtLwztvUffIh5rBjyb/48kQqVKldRCQE8fo66h59BHPFCmLFhWSfdjpCTwgmv/92xkQAPyc7m+uvvyEt8APYvHkzTzzxGG+99Ta67iESCfPGG6+zY+c2Hn7on7Rt2zYx0YXWat9alt3CnRNC4PP70HQtrWZLTDaxt6Y1m4R7qriVSqnjm2l0msVFAIqKisnPz6e8ogKv4SXUFGbduvXp5+a6Lrqus2XrVmzHwWsKolGLrKw8unY9eK/+fUpciEjUtdo9lZ9adc8551zGjBmL4fHQJr9NWgYghEDX9KQ4Ue6y7HYjjubjTMiUNitRXzRFqpqmIaXE9JjoupaoNCpVup2qdV8AkMkqp4kkiurUHt8lV6D7/BBqxI1Fcaw4RjiG6NgFNNly8ibvMWvQYHzdDkZIicjLTxQqTP2OJpGOldh5gIZwBFLJFhNQLynFHTyUus2bCOoKV0qEq+D9KcTmLKfpuKMInHk63m49WuiJIqtX469pwiN9iUB8g0vk6RcQjo3jUeA6CCWxpY40JFY8TENuFubxQ9FN86cYM1ItUtFSCKSQaUspGMxmw4YN3HDDtVx55dVceNHFmB6zhduWWgnz8vK59bY7uHTTJv751OP0PeIwBg85lm2//AU1Tz9BftxBxyYnGqbhXw/RGMwlOGpk2qTVsrPJu/AivGPH4snJbUZEu8RibjxOw2NP4pk1F5QH49jRBHr33UVY+3/3LSULyiUcDrVwIzp27Midd95NTW01n34yFb8vi5ycHGbO/IKXX3mRa6+5vhXyE7uytVLsIURMKYN3WT2tvxQhFd8QQuC4DqDtNRHRog2qpcVXWFjEuHHjefjhhzBNP6YZ4KOPPmbEiJEMHDgQKSXV1dV8+sknifylUljxGGPGnEinTgftmRZXKhFATsaMvsuqUUoRDGans2MqKThNuYaWZSXWj11028rcUOledd1dCYOUBbd718uUtegmNWR7CSGqpNZMVyDcBKFG8nPJGjESvz+rxbE5+8okmSZaM12Ta1ngOAgpkXW1+J0QjmYjNBMNA9yWVQCklOSffyFV1TsJTfsc0xIoQ8dneDG27aTprZeo+3oantPOJefsM9ECgYQ2rqYSPbnXM+6ClZuN6NYBJTW8rgIbIpqgViiEZeMoi6xxY2lz6mkHfOvOAbGMUtkQRHI1cxWanhgo8XgUgcIwNGLRKBMm3M+KlSu55eZbKCoqbkFIqfhR7969ufnmW7j+xut45MF/0Kd3H9qcdQ7lixfjmzmHgHDRdUV2TS31Dz6M1q49/v79WsjYfSkdR3PmTrpn9a++jPn2OxhKo7aoIzmnnp68h/9cXdpcDZyKU3i9Xn7xi7OY/vl0bCfhDhiGh2nTpnLRhZeQ3ZoaN73xuPUNpiL93XdbckqRrjOeiid936yplJKLL7mMjRs3Mun9DzBNP9u37+Smm37P6NGjyM3NZd78eXz77be4riIUauTU007hyiuv3sOVbekAyn20PUEctXW1vPzSS6xdswZNl8StGJFIBNdV1NbWpq2bXY9P7NOS/c7fTm3EFvu+iqYUUoASCespphv4HQfZPLXfXEtU34D0GOg+X2KcJMena1nUTZ5CaOFCfK6LcC2EFQUrht5Yj7euEqUnsqKOki0XluTY9bVrT+Gtd9HU43Aa33sHfds6siwbqXnwGV6cHTuIPf0vKrdupu2116MFgxh2PKHD0yTKcbFLuxK4525kfts0GfqT0ztRKUND9/t/lFIoB6SeUWogyKToL27FiUQiZGUFGDZsFAMGDGD655/xzTdz0A0fH7w/mfLNZdx775/o06dvq4Q0ftwpLF+2jKcfeIC3n3+Bi6+/npz/uYDalesxq6sTKmzDJLC1goZHH0K77z7MgsLd9EPNBkMyYF0/ayb2C8+SpSxCtoMcPhzfD96wKParf5RSlBSXEAxm01DXhBAJ8ePOHTuIRMJ7IaNdjsB3/XZCd6n2+zmlMmUttvHscaxsQa5KKXJzc/njn+6jb7/+vPHGG2zZsp2yss089NCDKBx0XSc3J59evQ7m1FNP5ZxzziUQCHzH5tcUGYj0NqLWNGn1DfXcfvutfPD+ZHRdx3EtLDtGMCtIfn5brLiVtMC172HHJqUEQmtl/1nzulxqnxeUSqGEi4OdXIi11KrcMlkiBLULvqXmsScQUZvAGadSeMZpCRc/Hqfm8SfhlVcJWlG8jiDmxGjMMqBNPqaj8FoOetICdkXipRWilcyuEQySd8E5eIYPpvGjd2ma/BHeihpkTMdvgAeb+snvUNejO23P/CVObhuUBL9yAIW1sxLqGtFKOux9z1gqlPFTs4xcV7UgE8uyKCwsYNCgYxg/fjwDBw7ENE1OO+10XnjhOV586VWshiaWLFnKNddcxb33/okhQ45tNbB99dXXsG7Jcp58/AkOHzqU/kcNpOzksdS8+Dz5aChHYXg1vAsWUvP0RNrfcssuseXuCmspaVq/gaZH/0XbmhqUrhEvKCAw+vgDZhWhEm5EakI1j6c0NTYRCcWQUk/vR/N6fa2XlRUqUf9m9/vYgwQFLXdBti49EIDtONgC4rFYOgPWPPay5xYpNx0w36UJy+Pkk09m+vTPWbduHUOHDuXIAUfi2BbZ2dl06dKV3r37UFBQuEdccHeKVem2J8q/pEgvvecu6WK+9dYbfDh5MtnZucm57XL6Gacxbtwp5Ofls3btGm666SbicSvtbqbub085idrtU+yjbLDa1+BPuKTYSENHRRSGKxPk03yPmVIoKYl99SXZC+eCK9EH9E8/wvp5c4m9/Trt3BCaVxJxgCEjyTvtdPTCQmRlJdxzF3LrNvAqlHRQIvEbSiU2kiMETtzCdR0Mr5dAp04EfnMd0REn0vja23imfEwg4qChk21HiM6aA2f+Es/BPYn6gqhoFKVrqOpqGj/9mLzehySysK7brLdEIuYlfpyXURwQN03XZPLhOxw7dDA33XQLPXv0TPu/juPQtm07brzxJo46ahD/mDCBZUuXUlZWzg03Xs/NN93CGWec2WIPluu6ZGdnc+Ptt3HRuefzyIMP8tjEibQ7+2y2fDuHwLLV+IQXEQ+TLRycye9R270rbc76VYtUdmpAxGprqH/4YXLWrsfweolFI+gHd01YRT9A8JkaxIm6RQkfVUoN0/S22GeVIptJ708iFA4RCARBSKy4Tb9+h7ZQYad34yeXvfRrnJr9nlKqWfBXpdcuISSuclsQQCooq5SDx9CIRMIsX7GcE088CY/Hk1Zs23YccJMK6ETqu/kkTj2Tr76axT333sma1as5ZvBQ7r3nXko7dEhujNb2Ml/3fJ9dYvtU6llrRKNxQqFwWo6QyKIl2jdnztdIKRBSEW6K0KtXT265+TYCgcAut1hLEXLiHgzD2G0HwK5xuEu0qFDJmlvN422JdL1KGk8J0kwQZGKhaW5pJzZmgxcN1xVoUmIoDekxd20FScktQo2Yy5fhExDxelB+367vFi/BH2lCl6BCNm5RKeZFv8bbp0/i+8Ji4noAQ5koBJpyMZRoEeQGqPp0Ck2ff0HR2WfjPfJwpKbh7d4D/ZbfUR9rID55Gj5bx3Qg2tSAa8Xx9elLU7cu2AsW4NUDmMqg7oOPqe/cneD4kxKLUfP7iMUIbdtKVknp3vd5/u9ZRi6hUCNKgcdrsn7dWp57biJjTzyJo44aiM/nb6F9GTp0KD179uSxxx7jzTdfZ+fOSu686w+Ew2HOO++CPQipd9++/P62W7n95lt449VXOP/iS8j9n/Opvu+vtGuK4XMBzSXLjlD97LM0de1O8IgjWlg6rutS/+wzZH0xE5+hoxwbTUncWKxFFmf/leeymWo45V7qxGI2K1auQDc8BAIBNE1j584dvP3O27z9zpv4A14QDpFIhHZt23HuuedjGEY6GJsouEZ6AkAi5ZyazKlUforgRJKElGsTt+JIIUHbNbnatmmTqESYDMTqus4bb7yOz+enS+eulJeXMfnDyaxdtw7T68F1bTRNsnbtaqZM+Qjbthg0aDBt2rTl82mfcfddd1CxrQKf38+mTZu4/IrLMQyDdm3bUlxaQmlJCZ06daJDx4506dyVrKysVi0j23FSe2kQQtLY0MSEByZw+GGHU15eRl19Lffc/Uc6d+6C6zoJ4nBdDEOnsTGxkbWgoD1r165hyidTiEaiaFJDKRdN01i8eDGfTZuKbVkMH35ceoEwdE9C3SFk+vdTxJvqM9M0MXQD27ETYlEpkrKJ3SZ/UlUthIvHlckXUQi0xiacefOIml7scBg3GkVFIogN6/CtXoUpPDQ5AtfrbRk0x8GWAqEb6E1xZPlW7JJiYlvKic+bh6euDlcTuNJBmBLP2tWEp3xMNOAneOhhaKaJnDKFnC++Irx4GdHjhuMdNxZvr57oHg++jp1xJbi6m9itUlyIEgJPTi7+M84itGoFetRB0zwEa+uwH3yQxgXz8QwZjGqTD7aNU76Fxm9mU1axju73/JW23Xse0GzaD94O0tyl2rlzB9U1NdTW1lFbW8vSpUv4fPrnLFq8CF3XKSkpwTRNhBBYlkUwGGT48OGUlpawZu0qduzYwTffzMHvD3Doof1bpLGVUvQ6pBfl5eW8++5bjBgxgqJ+h1G/eRPu+pV4NA+aEtiGjqhvIlpdg3fwMUjTl44TVX/4ATzxJLmAEongouYK6m0b7Zhj8CSrOO5PpwohqKgo56WXXmTGF1/g2IkJpWk68XicuXPn8Omnn/Dxxx/zwQeTeO211/hy5kxQCfFdLBahpLSYu+++m2HDhqeJd8fOHXz40WTmzJ6dfKecRDcMqqoqqa2tpmPHjuTk5BKLRZk6dSrffrsITepITSKEYv36tVRV7aRjh44Eg9kIITBNk08++Ziq6hpMjxdd04lEosz68ks++OB9Pv74YwraF3JY/8NYv24DhuFBCI3a2homvfc27733JkOGHEvnzl2YcP/9fD17DsFgNq6riERi1NTUUllZTXl5OUuWLGXWrC+ZOvVTpk37lKmfTWXx4kXEYjGKCoswTTP9TDdu3MC0z6ZimibgImSiT2d+OYPlK5YRCAQYNXI07dq1o6a6hunTZ6DrJrpu0NQU4euvv+bDDyfz7HNPs3z5MvLy2tLQ2JgoVazrbNq0kddfe5U538xm3MnjCQazWbBgAR988AE7dlRhGB40TcO2HTZsWE8gK0CH0g5omsa6dWuZNGkS9Q2NeAwTKTQaGxtYvmwJXo+HktJShILGlSuJvv0m2tLFmHEHDQ1Xl1ixMLE5s7E+nYIzdSpi6nTk9BmYK5Zg2Baa0AhJgThhFIGk92A5NqEvZyFjUZRXxxONYy1ZSuOsL2l4923Cc79By85CRJrQpINrKNy6KkJTp7Lzq6/xHTMYe/0mnGdfJN/QEVYTzoolRGd9RXzlauzaasS06WjVW7BFjNqsbIxLLyfQuQu4LmaXgwnH4oRWL8N1I/h0Bz0eQa5Yjvvll8RmzsSZOo341GmEKspQPQ8mMGAQWW3aJvfn/i+TUaoBgUCAE8acyKBBg+nc+SCi0ShNTSEaGxpYv249M2Z8weLFS9A0nY4dO6QHpatcevU8hMHHDKGquoqVK1Yxe/Y3OI7DwIEDW2hDpJT06deX9z+YxJbyMkaPHYferh3hObPxNIYwlAbKRRoCt6KCiD+A//DDQUrCy5bR9Jc/k9tQCwbYroMmBI4uiVkOqkMn/P32XZY2NZE+mzaViy4+j0mT3qW+vgHLsohGwsTiMWKxxATdtnUbm8s3U1a2iYaGOkyvF5/PS0lpCaeeehp/+MMdDBx4dNrS+Xr2V1xwwTm8887bNDQ2YMVjRKJRotEYtbVVvP/Bu3Ts0JEePXpw6WWX8OprrxKLxInFo0SjTURjEcrLN/HhR5PIb5PPkMHHJt3cHIqLi5k/fx7V1dU4roPt2AgEBQUFnH/+Bdxzz72MHXsS8+fPY9269en0edeuXfnNb65m2LDjCASyqKyqZt68+ZheE8dxiURiWJaddM8EXm9iT5qm6YRDESoqtrBs6XKmTPmQ0tJS+vbtl44tFhcXs3zFMtasWZUQSSobpVwOOaQXV1z+G26++Va6dOkKQLduB1NeXsGyZcuwrDi6rrFt2zYqKjbTq1cv7rvvr4wbdwozZ35BdVU1ylW4SjFgwACuu+5Gevc+hFtu/T1/uON2Nm7YiOMkLNNYLEYkEmbe/LmsXLmMX5xxJpM/mMQVv7mUss1l2LZDKBTCilvU19Xy7fxvWLLoW8acNJ4dX8xgw9334pm3AE88Tsy2iLkuTcoh6ti4TSGcWB0iHkZ3HRzpEBZWonhe2KZGSMxTxxM4qDO4LkZBIdFQBHvVWoiE0YSLbGgksmMH4eIi8i67nOyzf0nj2vWITduQSBp0SbRrN3J/+T9kHz2I0KZNRFatRtVXgwqj6S5mNIy2eh2xb77B3bkTF5va7Cy0X19O25PHp2OSQtPwHnY4Eb+f+KZN6HX1RHHA40FzFISjNLkudt9+BH/9a0ovupScktJWdWv/Ucjnh75RtnWtCoRCIb76ahafT5/Gl1/OZOvWbVhxi6ysIIMGDeL88y9g2LBhSCmxbRtd14lGozz73ESeePxJ6mrruen3N3L1Nde2iJNIKfl06ifcesuN/ONvD3LcyFFsf+IJjGeeJxcXiOMagljUJtS2iJz7/4Hs1IGdv7uB3IXzCUiBrRSOSBT0F7okZENs+Gja3/cn5D4EXKl7nDd/Lp9M+Riv15t4FXdKI5Pc6JkKLkspUSRiH23atKFTx4Po3r3HHgX5hRAsWrSQ999/D5/Pj8JFuSpRsRIQQhGNRjnxxJPo0qUrzz03kWgshq4ZyeqWqTiPJG5FOW74CAYNGtLi+osXL+KVV19mwYL5ZGfncMwxgxk96gT69u2Xvr8NGzYwYcL9rF+/jhPHjOX0M35Bhw4d0m2tra1l+fLlPP30U3z99VeMGnUCxcXFVFRUJOsKbSUUChGPxwCJ15sgpqamei6/4nJuveX2FvGjrVu38OxzE/nii+kYhsG4cadwyvhTKS4u3aPPGxrqeeWVV5gxYzo7duxE0wRjxpzIBRdcRLt27QD46qtZPPzIg8RiMc78xVmMG3cKOTk5OI7DxIlPUVlVidf0JhcAfZeo1LYoLCzkvHPPZ8aMacyaNYtAVlbSndul8XLsOIUFBfzil+dQs3QpjfO/JUdqaK6D0FJ1tSTYVjKGlygP6DoKxzRwDImMxvHYgia/j/yTxhIsLka5LkJKnFiMuo8+wp02FW3HVtw4qKMGknXeOfg6dkrMrTVraXzqBdTWCjwjhuAfMwZfkhRc1yW+dSuRWV+ivv0Gt2wjqq4RGVFYHgOVl43q3g3fySeTe8yQlpm4ZuM+snIl0RkzCK1ZjqyqwojEsLKzkSNHkT9mLJ78fH4sHBAyah4EbB74XLtuLZ9N/YTPp01j0aKFhEIR8tu0Zfjw47j44ovS9WpSg+7TTz/h73//KxvWb+Cee//Ieeedvwch3Xn37SxdMJ8XX34TIxpjx02/J7hoPgFTQzoKTZk4MYf40QMIFebDJ1PJFS6GZYEDruHBFi5ogpjlUtelG+0f/SdmYeH/l3rEuxdK+7HRfKHYsWM7Ho+HvLz8PdL6Qgji8Ti1tTXpbFjzZ+q6Ls88M5GHH36AsWNP4o9/vA/TNInH49TX11NdXcXKlSuZN28ey5Ytp2xTBbF4jHg8zNVXX8kNN/y+5WbTZJu2b9+GpulpUtk94N382KamRurq6tF1ncLCwj2Ob2iox7Zt8pP7qn7s15D/GHDCIey6OhACT/uCXfvYkuzpRGPYoRBmqjhbSsLSjF7ceByntga7qQlsB6Vp6FkBjLbtErsM9q5eTo9/OxrBaWwC20b6fRipF5furtL/KZLR7pOt+YCqqa5m9uyv+GDy+3z99Vy2b99OcUkh5593Pueeex7t2xfguolg4oYN67n3T39i5syZ3HnHHVx4wYXpmtqQKJZ1yUXncMLoMVz2m2vY8cEktL/eR3YsjKY0pKUjkFjSxpIuHmEiiKHhgmuAMFACok6Mhi6d0C+4kLyxJ+13ZqD5O9K+r0u7t/rS+3PNVMp7X6LF1l4UufukbG3haD7pdyepxsZGHnxwAhMnPkVJaQmvvPw6nTt3IRaLYRjGHhN+1apVXPPb69iwcQO6Lnj44Yc54YQxLdqxu5SjtazbdxF4a+envt/9WimZwnc9n9Qev1ZV2Xscl8hoir2ll/dnSu1eU3pvk7yVzPBej0+Jfr+LgJNx1H1sqWi9HT8SCR2QAPZ3ietS2hSlFP5AgO7dezBmzFj69+9PVpafDRvW8+nUKcz88gvy8/PpkQzm5eXlM3LEKBobG3nqqccpKCigT+8+6cHk8/koLe3AS/9+niHDhlPYuw+hxcvwbKxAejxJmaBC6Qo0ie4INOWAFMQMA0vpxC2FM/gYfLffRvbgIXuUdNjX/bVWf/m7/jTvkx96zeb7tfbnuNbIMNWHrR3XXCG9e92phx95gCeeeBy/PwvDMDjqqKM46KDO6Lre4jrRaJSFCxfwz3/+kxUrVxCPRTn00H5ceeVVeJPZo92J+bvatKewVu3xf621f/dr7at/d49P7s9xYvdaRT/kT2vk1Foh/NaIbn+ukSKn5tfcl6X4Xe34kb2GA05GextEKQFdx44dGTlyFEOHHkswO5vFixbz9ttvUlZWRq9eh5Cbm4vH42HYsGFkZ2Xxz8cf5aDOnenapWt6Re/U6SA2byxj4TdzGDpqNFZWkPDXs/HYMdBsZDIbpUSiWqNUCleDBqGoCwYRZ/2C4HXXYHbssKsA+X8B9vXmj729VuiTT6ewaPFSTNNPNBJhzpzZbN22harKSjZu2siqVSuZNu0zJk6cyMSJE1m1ahVW3CIYzOLOO++iV69eexdA7ucrwndf6H7I/f1MHtLeCWt/SWF/CfA/acdPOWa0vxLlhDu2i53XrVvHq6++wjvvvIXf5+PmW25j/PhT0t9P/ugDnnn+GW6/+XaOPGJAOgNVV1vLH2+9ict+ey09eh3C1lv/QPanU8gyJMJKvODR1QRIC+m6NNoWNV26kHXp5eSddHL6VduZd1Z9dxxQSsk338zh0kt/TTgcxev1EY/HiMWigErEjSwrUXZV09E0iW1Hyc/P5w9/uJNTTz39J/8+tAx+GvhRLaNWuK+FHw/Qpk0bhg0bxogRIwiHQ7z44vNsKtvEof0Pw+fz0f3gHnQo6cArr75Ely5dadeuHa7r4g8EyMnP4+svv+CoY4ag/D6avpqFP2whhIEtXJRu40qbiKuIHnE0wZtvJffYYQlXLkNE+2WJKKUoLi7B7zP5dsFc6usTeh6v15fUjkl0zUDT9GR2VDJ48EDuvffPjBgxKkNEGfxULaPWV9/m5vWqVSt5+uknaGhq5DdXXMXhhyUybosWLeTLWTP5n3POIz8/P71qv/ri8xx++JH06NmTslt+R95nM/BrBo4GUeUQ0U20cePIufgSPAUFGRL6DzBr1kwmTXqPuXPnUlVVhSKljPaQ36YNfXr3ZeTIkYwePRq/P5Ahogx+XmTUnJRSwUeAGTM+57PPp3Lc8BEcN/x4pJSUlW1i/YZ1HDNoCF7TTLxwcds21ixdzNEjR1MzbSrxP91Du6iF5TrUti/AvPgy2pwyHqlr+84kZLB3B7uZIn7r1i1s376Nuro6NE0jNy+P9u3bU9C+IPEiAX6eafUMMmS0V0upsbGRL2d9QU5ODocffiQ+r49t27YSjUQ4KFlNUQjBzi0VZOe3QVoWO66/Dv/c+cSP6E/wqt+SNWBAajZlLKIDFEP6PpZuBhn8bMmotYG/ZesWvKZJfrLkaCQSwTTNlnvYnEThp20vvkRsyVIKrrkaX4cO/1/0Ef9tFlLzz+bxpeafGWTwf4aMmg/47zPAraYm0DUMry9jDWWQQYaMfpx4xfc8KUNEGWTwM4L+c2jk9yKijFuWQQYZMvqJMFfmqWaQwc8QmdxrBhlkkCGjDDLIIIMMGWWQQQYZMsoggwwyyJBRBhlkkCGjDDLIIIMMGWWQQQYZMsoggwwyyJBRBhlkkCGjDDLIIIMMGWWQQQY/K/w/ByuJMlmQpEYAAAAASUVORK5CYII="""


def _get_logo_image(height=50):
    """height >= 50 = login/activate size, < 50 = header size"""
    import base64 as _b64, io as _io
    # Kaun sa logo use karna hai
    b64_data = LOGO_LOGIN_B64 if height >= 50 else LOGO_B64
    if PIL_OK:
        try:
            data = _b64.b64decode(b64_data)
            img = Image.open(_io.BytesIO(data)).convert("RGBA")
            return ImageTk.PhotoImage(img)
        except:
            pass
    # PIL nahi hai — PNG ko tk se load karo
    try:
        import tempfile, os as _os
        data = _b64.b64decode(b64_data)
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
        tmp.write(data); tmp.close()
        photo = tk.PhotoImage(file=tmp.name)
        _os.unlink(tmp.name)
        return photo
    except:
        return None


# ─── DATABASE PATH ────────────────────────────────────────────────────────────
if getattr(sys, 'frozen', False):
    _APP_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    _APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(_APP_DIR, "bhugtanease_paint.db")

# ─── LICENSE SYSTEM ───────────────────────────────────────────────────────────
_LIC_SECRET  = "BhugtanEase@2025#Paint#VNS"
_LIC_BASE_DT = datetime.date(2020, 1, 1)

def _make_serial_paint(expiry_date: datetime.date, customer_name: str = "") -> str:
    days = (expiry_date - _LIC_BASE_DT).days
    date_block = f"{days:04X}"
    cname = customer_name.strip().upper()
    raw = f"{_LIC_SECRET}|{expiry_date.isoformat()}|{cname}"
    h = hashlib.sha256(raw.encode()).hexdigest().upper()
    return f"BE-{date_block}-{h[0:4]}-{h[4:8]}-{h[8:12]}"

def _decode_serial(serial: str, customer_name: str = ""):
    serial = serial.strip().upper().replace(" ", "")
    parts = serial.split("-")
    if len(parts) != 5 or parts[0] != "BE" or len(parts[1]) != 4:
        return None
    try:
        days   = int(parts[1], 16)
        expiry = _LIC_BASE_DT + datetime.timedelta(days=days)
    except Exception:
        return None
    cname = customer_name.strip().upper()
    raw = f"{_LIC_SECRET}|{expiry.isoformat()}|{cname}"
    h   = hashlib.sha256(raw.encode()).hexdigest().upper()
    expected = f"BE-{days:04X}-{h[0:4]}-{h[4:8]}-{h[8:12]}"
    return expiry if expected == serial else None

def _get_license_info():
    try:
        conn = sqlite3.connect(DB_FILE)
        rows = {r[0]: r[1] for r in conn.execute(
            "SELECT key,value FROM settings WHERE key IN "
            "('install_date','license_expiry','license_customer','license_serial')"
        ).fetchall()}
        conn.close()
        install_date_str = rows.get('install_date', '').strip()
        expiry_str       = rows.get('license_expiry', '').strip()
        customer         = rows.get('license_customer', '').strip()
        if not install_date_str or not expiry_str:
            return ('new', 0, '', customer)
        expiry_date = datetime.date.fromisoformat(expiry_str)
        today       = datetime.date.today()
        days_left   = (expiry_date - today).days
        if today > expiry_date:
            return ('expired', days_left, install_date_str, customer)
        return ('ok', days_left, install_date_str, customer)
    except Exception:
        return ('new', 0, '', '')

def _activate_license(serial: str, customer_name: str = "") -> bool:
    expiry_date = _decode_serial(serial, customer_name)
    if expiry_date is None:
        return False
    if expiry_date <= datetime.date.today():
        return False
    try:
        today = datetime.date.today()
        conn  = sqlite3.connect(DB_FILE)
        for k, v in [
            ('install_date',     today.isoformat()),
            ('license_expiry',   expiry_date.isoformat()),
            ('license_customer', customer_name.strip()),
            ('license_serial',   serial.strip().upper()),
        ]:
            conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k, v))
        conn.commit(); conn.close()
        return True
    except Exception:
        return False

# ─── DATABASE ─────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db(); c = conn.cursor()
    c.executescript("""
        CREATE TABLE IF NOT EXISTS settings (
            key   TEXT PRIMARY KEY,
            value TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            username TEXT UNIQUE,
            password TEXT
        );
        CREATE TABLE IF NOT EXISTS products (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            name          TEXT UNIQUE NOT NULL,
            category      TEXT DEFAULT 'Paint',
            brand         TEXT DEFAULT '',
            hsn           TEXT DEFAULT '3208',
            unit          TEXT DEFAULT 'Ltr',
            sale_rate     REAL DEFAULT 0,
            purchase_rate REAL DEFAULT 0,
            mrp           REAL DEFAULT 0,
            gst_percent   REAL DEFAULT 18,
            opening_stock REAL DEFAULT 0,
            low_stock_alert REAL DEFAULT 5,
            is_base       INTEGER DEFAULT 0,
            is_tinter     INTEGER DEFAULT 0
        );
        CREATE TABLE IF NOT EXISTS parties (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            name     TEXT UNIQUE NOT NULL,
            ptype    TEXT DEFAULT 'Customer',
            mobile   TEXT DEFAULT '',
            gstin    TEXT DEFAULT '',
            address  TEXT DEFAULT '',
            state    TEXT DEFAULT 'Uttar Pradesh',
            email    TEXT DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS color_formulas (
            id           INTEGER PRIMARY KEY AUTOINCREMENT,
            shade_name   TEXT UNIQUE NOT NULL,
            shade_code   TEXT DEFAULT '',
            brand        TEXT DEFAULT '',
            base_product TEXT DEFAULT '',
            base_qty     REAL DEFAULT 0,
            base_unit    TEXT DEFAULT 'Ltr',
            components   TEXT DEFAULT '[]',
            total_cost   REAL DEFAULT 0,
            sale_price   REAL DEFAULT 0,
            notes        TEXT DEFAULT '',
            created_at   TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sales (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_no       TEXT UNIQUE NOT NULL,
            bill_date     TEXT NOT NULL,
            party         TEXT DEFAULT '',
            party_mobile  TEXT DEFAULT '',
            party_address TEXT DEFAULT '',
            party_gstin   TEXT DEFAULT '',
            gst_type      TEXT DEFAULT 'GST',
            grand_total   REAL DEFAULT 0,
            discount      REAL DEFAULT 0,
            pay_mode      TEXT DEFAULT 'Cash',
            due_date      TEXT DEFAULT '',
            notes         TEXT DEFAULT '',
            created_at    TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sale_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            sale_id     INTEGER NOT NULL,
            product     TEXT,
            shade_name  TEXT DEFAULT '',
            hsn         TEXT DEFAULT '3208',
            qty         REAL DEFAULT 0,
            unit        TEXT DEFAULT 'Ltr',
            rate        REAL DEFAULT 0,
            taxable     REAL DEFAULT 0,
            gst_percent REAL DEFAULT 18,
            gst_amt     REAL DEFAULT 0,
            grand       REAL DEFAULT 0,
            FOREIGN KEY(sale_id) REFERENCES sales(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS purchases (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_no     TEXT UNIQUE NOT NULL,
            bill_date   TEXT NOT NULL,
            party       TEXT DEFAULT '',
            grand_total REAL DEFAULT 0,
            pay_mode    TEXT DEFAULT 'Credit',
            due_date    TEXT DEFAULT '',
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS purchase_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            purchase_id INTEGER NOT NULL,
            product     TEXT,
            qty         REAL DEFAULT 0,
            unit        TEXT DEFAULT 'Ltr',
            rate        REAL DEFAULT 0,
            taxable     REAL DEFAULT 0,
            gst_percent REAL DEFAULT 18,
            gst_amt     REAL DEFAULT 0,
            total       REAL DEFAULT 0,
            batch_no    TEXT DEFAULT '',
            expiry_date TEXT DEFAULT '',
            FOREIGN KEY(purchase_id) REFERENCES purchases(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS bill_payments (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_type  TEXT NOT NULL,
            bill_no    TEXT NOT NULL,
            party      TEXT NOT NULL,
            pay_date   TEXT NOT NULL,
            amount     REAL NOT NULL,
            pay_mode   TEXT DEFAULT 'Cash',
            note       TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS stock_movements (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            move_date  TEXT NOT NULL,
            product    TEXT NOT NULL,
            move_type  TEXT NOT NULL,
            qty        REAL NOT NULL,
            ref_no     TEXT DEFAULT '',
            reason     TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS expenses (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            exp_date    TEXT NOT NULL,
            category    TEXT NOT NULL,
            description TEXT DEFAULT '',
            amount      REAL NOT NULL,
            pay_mode    TEXT DEFAULT 'Cash',
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sale_returns (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            return_no   TEXT UNIQUE NOT NULL,
            return_date TEXT NOT NULL,
            orig_bill   TEXT DEFAULT '',
            party       TEXT DEFAULT '',
            reason      TEXT DEFAULT '',
            grand_total REAL DEFAULT 0,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS sale_return_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            return_id   INTEGER NOT NULL,
            product     TEXT,
            qty         REAL DEFAULT 0,
            unit        TEXT DEFAULT 'Ltr',
            rate        REAL DEFAULT 0,
            total       REAL DEFAULT 0,
            FOREIGN KEY(return_id) REFERENCES sale_returns(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS purchase_returns (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            return_no   TEXT UNIQUE NOT NULL,
            return_date TEXT NOT NULL,
            orig_bill   TEXT DEFAULT '',
            party       TEXT DEFAULT '',
            reason      TEXT DEFAULT '',
            grand_total REAL DEFAULT 0,
            created_at  TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS purchase_return_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            return_id   INTEGER NOT NULL,
            product     TEXT,
            qty         REAL DEFAULT 0,
            unit        TEXT DEFAULT 'Ltr',
            rate        REAL DEFAULT 0,
            total       REAL DEFAULT 0,
            FOREIGN KEY(return_id) REFERENCES purchase_returns(id) ON DELETE CASCADE
        );
    """)
    # Migration: purane DB me batch/expiry column add karo (agar already nahi hai)
    for _col, _coldef in [("batch_no", "TEXT DEFAULT ''"), ("expiry_date", "TEXT DEFAULT ''")]:
        try:
            c.execute(f"ALTER TABLE purchase_items ADD COLUMN {_col} {_coldef}")
        except Exception:
            pass
    # Default settings
    for k, v in [
        ("install_date",""), ("license_expiry",""), ("license_serial",""), ("license_customer",""),
        ("shop_name","BhugtanEase Car Paint Shop"), ("shop_address",""), ("shop_city",""),
        ("shop_state","Uttar Pradesh"), ("shop_gstin",""), ("shop_mobile",""),
        ("shop_email",""), ("shop_bank",""), ("shop_ifsc",""), ("shop_account",""),
        ("shop_upi",""), ("shop_print_tnc","Maal wapas nahi hoga / No exchange policy"),
        ("bill_prefix","CP"), ("next_bill_no","1"),
    ]:
        c.execute("INSERT OR IGNORE INTO settings(key,value) VALUES(?,?)", (k,v))
    c.execute("INSERT OR IGNORE INTO users(username,password) VALUES(?,?)", ("admin","empo123"))
    # Sample products (Automotive / Car Paint)
    sample_products = [
        ("PPG Envirobase White Basecoat", "Paint", "PPG",          "3208", "Ltr", 850, 700, 950, 18, 20, 5,  1, 0),
        ("PPG Envirobase Black Basecoat", "Paint", "PPG",          "3208", "Ltr", 880, 720, 980, 18, 18, 5,  1, 0),
        ("Nippon 1K Silver Basecoat",     "Paint", "Nippon Paint", "3208", "Ltr", 790, 640, 880, 18, 15, 5,  1, 0),
        ("Tinter Red (R-1)",             "Tinter","Generic",      "3208", "ML",  2.5, 1.8, 3.0, 18, 500, 50, 0, 1),
        ("Tinter Yellow (Y-2)",          "Tinter","Generic",      "3208", "ML",  2.2, 1.6, 2.8, 18, 400, 50, 0, 1),
        ("Tinter Blue (B-3)",            "Tinter","Generic",      "3208", "ML",  2.8, 2.0, 3.2, 18, 300, 50, 0, 1),
        ("Tinter Green (G-4)",           "Tinter","Generic",      "3208", "ML",  2.6, 1.9, 3.0, 18, 350, 50, 0, 1),
        ("Tinter Black (K-5)",           "Tinter","Generic",      "3208", "ML",  2.0, 1.5, 2.5, 18, 600, 50, 0, 1),
        ("Tinter Orange (O-6)",          "Tinter","Generic",      "3208", "ML",  2.4, 1.7, 2.9, 18, 250, 50, 0, 1),
        ("Primer Surfacer Grey",         "Primer","PPG",          "3210", "Ltr", 320, 260, 360, 18, 25, 8,  0, 0),
        ("Body Filler Putty (1kg)",      "Putty", "P38",          "3214", "Kg",  260, 210, 300, 18, 40, 10, 0, 0),
        ("Thinner (Automotive Grade)",   "Thinner","Generic",     "3814", "Ltr", 150, 120, 180, 18, 50, 10, 0, 0),
        ("Touch-up Brush (Fine Tip)",    "Brush", "Generic",      "9603", "Pcs", 35,  22,  45,  12, 80,  15, 0, 0),
        ("Masking Tape (Automotive)",    "Other", "3M",           "3919", "Pcs", 95,  70,  120, 18, 60,  15, 0, 0),
    ]
    for p in sample_products:
        try:
            c.execute("""INSERT OR IGNORE INTO products
                (name,category,brand,hsn,unit,sale_rate,purchase_rate,mrp,gst_percent,
                 opening_stock,low_stock_alert,is_base,is_tinter)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""", p)
        except: pass
    # Sample color formulas (Vehicle OEM Shades)
    sample_formulas = [
        ("Maruti Fire Red", "085", "Maruti Suzuki",
         "PPG Envirobase White Basecoat", 1.0, "Ltr",
         json.dumps([{"product":"Tinter Red (R-1)","qty":90,"unit":"ML"},
                     {"product":"Tinter Orange (O-6)","qty":15,"unit":"ML"}]),
         410, 480, "Swift / Baleno / WagonR jaisa popular shade"),
        ("Hyundai Phantom Black", "X5", "Hyundai",
         "PPG Envirobase Black Basecoat", 1.0, "Ltr",
         json.dumps([{"product":"Tinter Black (K-5)","qty":50,"unit":"ML"}]),
         360, 430, "i20 / Creta / Verna ki premium black shade"),
        ("Honda Lunar Silver Metallic", "NH-700M", "Honda",
         "Nippon 1K Silver Basecoat", 1.0, "Ltr",
         json.dumps([{"product":"Tinter Blue (B-3)","qty":10,"unit":"ML"},
                     {"product":"Tinter Black (K-5)","qty":15,"unit":"ML"}]),
         330, 400, "City / Amaze ka common silver shade"),
    ]
    for f in sample_formulas:
        try:
            c.execute("""INSERT OR IGNORE INTO color_formulas
                (shade_name,shade_code,brand,base_product,base_qty,base_unit,
                 components,total_cost,sale_price,notes)
                VALUES(?,?,?,?,?,?,?,?,?,?)""", f)
        except: pass
    conn.commit(); conn.close()

def get_shop():
    try:
        conn = get_db()
        rows = {r[0]: r[1] for r in conn.execute(
            "SELECT key,value FROM settings WHERE key LIKE 'shop_%'").fetchall()}
        conn.close()
    except: rows = {}
    return {
        "name":    rows.get("shop_name",    "BhugtanEase Car Paint Shop"),
        "address": rows.get("shop_address", ""),
        "city":    rows.get("shop_city",    ""),
        "state":   rows.get("shop_state",   "Uttar Pradesh"),
        "gstin":   rows.get("shop_gstin",   ""),
        "mobile":  rows.get("shop_mobile",  ""),
        "email":   rows.get("shop_email",   ""),
        "bank":    rows.get("shop_bank",    ""),
        "ifsc":    rows.get("shop_ifsc",    ""),
        "account": rows.get("shop_account", ""),
        "upi":     rows.get("shop_upi",     ""),
        "tnc":     rows.get("shop_print_tnc",""),
    }

def _backup_dir():
    """Returns the backup folder (same folder as DB, 'backups' subfolder). Creates it if needed."""
    d = os.path.join(os.path.dirname(DB_FILE), "backups")
    os.makedirs(d, exist_ok=True)
    return d

def do_backup(silent=False):
    """
    Manual/auto backup — copies DB to backups/<timestamp>.db
    silent=True → no success popup (used for auto-backup on close).
    Keeps only the last 30 backup files.
    Returns the backup path on success, or None on failure.
    """
    try:
        ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(_backup_dir(), f"bhugtanease_paint_{ts}.db")
        shutil.copy2(DB_FILE, dest)
        # Prune: keep latest 30 backups
        files = sorted([
            os.path.join(_backup_dir(), f)
            for f in os.listdir(_backup_dir())
            if f.endswith(".db")
        ])
        for old in files[:-30]:
            try: os.remove(old)
            except: pass
        if not silent:
            messagebox.showinfo("✅ Backup Complete",
                f"Backup save ho gaya!\n\nFile:\n{os.path.basename(dest)}\n\nFolder:\n{_backup_dir()}")
        return dest
    except Exception as e:
        if not silent:
            messagebox.showerror("❌ Backup Failed", f"Backup nahi hua:\n{e}")
        return None

def get_drive_path():
    """Returns saved Google Drive sync folder path, or '' if not set."""
    try:
        conn = get_db()
        row  = conn.execute("SELECT value FROM settings WHERE key='gdrive_path'").fetchone()
        conn.close()
        return (row["value"] or "").strip() if row else ""
    except: return ""

def set_drive_path(path):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES('gdrive_path',?)", (path,))
    conn.commit(); conn.close()

def do_drive_backup(silent=False):
    """
    Copies the latest local backup (or current DB) to the user's Google Drive
    sync folder. Returns True on success, False on failure.
    silent=True → no success popup.
    """
    drive_path = get_drive_path()
    if not drive_path:
        if not silent:
            messagebox.showwarning("📁 Drive Folder Set Nahi",
                "Pehle Google Drive folder set karo\n(Settings → Backup & Restore → Drive section).")
        return False
    if not os.path.isdir(drive_path):
        if not silent:
            messagebox.showerror("❌ Folder Nahi Mila",
                f"Ye folder exist nahi karta:\n{drive_path}\n\nKripya sahi folder chunein.")
        return False
    try:
        ts   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(drive_path, f"bhugtanease_paint_{ts}.db")
        shutil.copy2(DB_FILE, dest)
        # Prune Drive folder: keep latest 10
        drive_files = sorted([
            os.path.join(drive_path, f)
            for f in os.listdir(drive_path)
            if f.startswith("bhugtanease_paint_") and f.endswith(".db")
        ])
        for old in drive_files[:-10]:
            try: os.remove(old)
            except: pass
        if not silent:
            messagebox.showinfo("☁️ Drive Backup Complete",
                f"Google Drive me backup ho gaya!\n\nFile: {os.path.basename(dest)}\nFolder: {drive_path}")
        return True
    except Exception as e:
        if not silent:
            messagebox.showerror("❌ Drive Backup Failed", f"Drive backup nahi hua:\n{e}")
        return False

def next_bill_no(commit=True):
    conn = get_db()
    prefix = conn.execute("SELECT value FROM settings WHERE key='bill_prefix'").fetchone()
    num    = conn.execute("SELECT value FROM settings WHERE key='next_bill_no'").fetchone()
    p = prefix["value"] if prefix else "PS"
    n = int(num["value"]) if num else 1
    if commit:
        conn.execute("UPDATE settings SET value=? WHERE key='next_bill_no'", (str(n+1),))
        conn.commit()
    conn.close()
    today = datetime.date.today()
    return f"{p}/{today.year % 100:02d}-{(today.year+1) % 100:02d}/{n:04d}"

def _party_balance(party_name, conn=None):
    """Net balance for a party: positive = party hamein dega (Dr), negative = hum denge (Cr)"""
    close_conn = conn is None
    if close_conn: conn = get_db()
    sale_total  = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM sales WHERE party=?", (party_name,)).fetchone()["t"]
    sale_recv   = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM bill_payments WHERE party=? AND bill_type='sale'", (party_name,)).fetchone()["t"]
    sale_ret    = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM sale_returns WHERE party=?", (party_name,)).fetchone()["t"]
    pur_total   = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM purchases WHERE party=?", (party_name,)).fetchone()["t"]
    pur_paid    = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM bill_payments WHERE party=? AND bill_type='purchase'", (party_name,)).fetchone()["t"]
    pur_ret     = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM purchase_returns WHERE party=?", (party_name,)).fetchone()["t"]
    if close_conn: conn.close()
    # sale_total - sale_recv - sale_ret = customer ka outstanding
    # pur_total - pur_paid - pur_ret = supplier ko dena hai
    return round((sale_total - sale_recv - sale_ret) - (pur_total - pur_paid - pur_ret), 2)

def get_stock(product_name):
    conn = get_db()
    prod = conn.execute("SELECT opening_stock FROM products WHERE name=?", (product_name,)).fetchone()
    opening = prod["opening_stock"] if prod else 0
    purchase_qty = conn.execute(
        "SELECT COALESCE(SUM(qty),0) as t FROM purchase_items WHERE product=?", (product_name,)
    ).fetchone()["t"]
    sale_qty = conn.execute(
        "SELECT COALESCE(SUM(qty),0) as t FROM sale_items WHERE product=?", (product_name,)
    ).fetchone()["t"]
    # Returns: sale return = stock wapas aaya, purchase return = stock gayi
    try:
        sale_ret_qty = conn.execute(
            "SELECT COALESCE(SUM(qty),0) as t FROM sale_return_items WHERE product=?", (product_name,)
        ).fetchone()["t"]
        pur_ret_qty = conn.execute(
            "SELECT COALESCE(SUM(qty),0) as t FROM purchase_return_items WHERE product=?", (product_name,)
        ).fetchone()["t"]
    except:
        sale_ret_qty = pur_ret_qty = 0
    conn.close()
    return opening + purchase_qty - sale_qty + sale_ret_qty - pur_ret_qty

MON = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

# ─── COLORS ───────────────────────────────────────────────────────────────────
C_ORANGE  = "#E65100"
C_DARK    = "#BF360C"
C_AMBER   = "#FF8F00"
C_YELLOW  = "#FFF8E1"
C_WHITE   = "#FFFFFF"
C_GRAY    = "#546E7A"
C_LIGHT   = "#F5F5F5"
C_GREEN   = "#2E7D32"
C_RED     = "#C62828"
C_BLUE    = "#1565C0"
C_TEAL    = "#00695C"
C_PURPLE  = "#6A1B9A"
C_BG      = "#FFF3E0"

def make_btn(parent, text, cmd, bg=C_ORANGE, fg="white", padx=10, pady=6, font=None):
    if font is None: font = ("Segoe UI", 9, "bold")
    b = tk.Button(parent, text=text, command=cmd, bg=bg, fg=fg,
                  font=font, relief="flat", bd=0, cursor="hand2",
                  padx=padx, pady=pady, activebackground=C_DARK,
                  activeforeground="white")
    return b

def bind_enter_nav(widgets):
    """
    Enter key se ek field se dusri field mein jaao.
    widgets: list of ttk.Entry / ttk.Combobox / tk.Button / DatePicker
    Last widget pe Enter = click (agar Button hai) ya kuch nahi.
    Disabled ya invisible widgets skip ho jaate hain.
    """
    def _make_handler(i):
        def _handler(ev):
            nxt_i = i + 1
            while nxt_i < len(widgets):
                nxt = widgets[nxt_i]
                # Button pe Enter = click
                if isinstance(nxt, tk.Button):
                    nxt.invoke()
                    return "break"
                # Skip disabled widgets
                try:
                    state = str(nxt.cget("state"))
                    if state == "disabled":
                        nxt_i += 1
                        continue
                except: pass
                # Focus pe ja
                nxt.focus_force()
                try: nxt.select_range(0, "end")
                except: pass
                return "break"
            return "break"
        return _handler

    for idx, w in enumerate(widgets):
        if isinstance(w, tk.Button):
            continue
        # DatePicker ke andar .entry attribute ho sakta hai
        target = getattr(w, 'entry', w)
        target.bind("<Return>",   _make_handler(idx))
        target.bind("<KP_Enter>", _make_handler(idx))

# (Combobox auto-open patch removed — native Tkinter behavior use hoga)

def rupee(amt):
    try: return f"₹{float(amt):,.2f}"
    except: return "₹0.00"

def mmyy_to_iso(mmyy_str):
    """MM/YY -> ISO date (us mahine ka aakhri din, jab tak product valid maana jaata hai)"""
    dt = datetime.datetime.strptime(mmyy_str.strip(), "%m/%y")
    last_day = calendar.monthrange(dt.year, dt.month)[1]
    return datetime.date(dt.year, dt.month, last_day).isoformat()

def iso_to_mmyy(iso_str):
    """ISO date -> MM/YY dikhane ke liye"""
    try:
        d = datetime.date.fromisoformat(iso_str)
        return d.strftime("%m/%y")
    except Exception:
        return iso_str or ""

def amt_words(n):
    """Number to words (Indian system) — invoice ke liye"""
    try: n = int(round(float(n)))
    except: return "Zero"
    if n == 0: return "Zero"
    ones = ["","One","Two","Three","Four","Five","Six","Seven","Eight","Nine",
            "Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen",
            "Seventeen","Eighteen","Nineteen"]
    tens = ["","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"]
    def two_digit(x):
        if x < 20: return ones[x]
        return tens[x//10] + (" " + ones[x%10] if x%10 else "")
    def three_digit(x):
        if x >= 100:
            return ones[x//100] + " Hundred" + (" " + two_digit(x%100) if x%100 else "")
        return two_digit(x)
    parts = []
    if n >= 10000000:
        parts.append(three_digit(n//10000000) + " Crore")
        n %= 10000000
    if n >= 100000:
        parts.append(three_digit(n//100000) + " Lakh")
        n %= 100000
    if n >= 1000:
        parts.append(three_digit(n//1000) + " Thousand")
        n %= 1000
    if n > 0:
        parts.append(three_digit(n))
    return " ".join(parts) + " Only"

# ─── DATE PICKER ──────────────────────────────────────────────────────────────
class DatePicker(tk.Frame):
    def __init__(self, parent, textvariable, width=13, bg="white", **kw):
        super().__init__(parent, bg=bg)
        self.var = textvariable
        self.entry = ttk.Entry(self, textvariable=self.var, width=width)
        self.entry.pack(side="left")
        today = datetime.date.today().strftime("%d-%m-%Y")
        if not self.var.get():
            self.var.set(today)
        # Auto-format: agar user DDMMYYYY type kare toh DD-MM-YYYY me convert karo
        def _auto_fmt(ev):
            val = self.var.get().replace("-","").replace("/","").strip()
            if len(val) == 8 and val.isdigit():
                try:
                    datetime.datetime.strptime(f"{val[:2]}-{val[2:4]}-{val[4:]}", "%d-%m-%Y")
                    self.var.set(f"{val[:2]}-{val[2:4]}-{val[4:]}")
                except: pass
        self.entry.bind("<FocusOut>", _auto_fmt, add="+")
        self.entry.bind("<Return>", _auto_fmt, add="+")

        # ── Calendar icon button — popup calendar se purani date bhi select ho sake ──
        self.cal_btn = tk.Button(self, text="📅", font=("Segoe UI",8), relief="flat",
                                  bg=bg, command=self._open_calendar, cursor="hand2",
                                  bd=0, padx=2)
        self.cal_btn.pack(side="left", padx=(2,0))

    def _open_calendar(self):
        # Current date parse karo (agar invalid hai toh aaj ki date)
        try:
            cur = datetime.datetime.strptime(self.var.get().strip(), "%d-%m-%Y").date()
        except Exception:
            cur = datetime.date.today()

        top = tk.Toplevel(self)
        top.title("Date Chunein")
        top.resizable(False, False)
        top.transient(self.winfo_toplevel())
        top.grab_set()
        top.configure(bg="white")

        state = {"year": cur.year, "month": cur.month}

        hdr = tk.Frame(top, bg=C_ORANGE); hdr.pack(fill="x")
        lbl_month = tk.Label(hdr, font=("Segoe UI",10,"bold"), bg=C_ORANGE, fg="white", width=16)
        lbl_month.pack(side="left", padx=8, pady=6)

        def _prev_month():
            state["month"] -= 1
            if state["month"] < 1:
                state["month"] = 12; state["year"] -= 1
            _draw()
        def _next_month():
            state["month"] += 1
            if state["month"] > 12:
                state["month"] = 1; state["year"] += 1
            _draw()

        tk.Button(hdr, text="◀", command=_prev_month, relief="flat", bg=C_ORANGE, fg="white",
                  bd=0, font=("Segoe UI",10,"bold")).pack(side="left", padx=4)
        tk.Button(hdr, text="▶", command=_next_month, relief="flat", bg=C_ORANGE, fg="white",
                  bd=0, font=("Segoe UI",10,"bold")).pack(side="right", padx=4)

        grid_frame = tk.Frame(top, bg="white", padx=6, pady=6)
        grid_frame.pack()

        def _pick(d):
            self.var.set(d.strftime("%d-%m-%Y"))
            top.destroy()

        def _draw():
            for w in grid_frame.winfo_children():
                w.destroy()
            y, m = state["year"], state["month"]
            lbl_month.config(text=datetime.date(y, m, 1).strftime("%B %Y"))
            days = ["Mo","Tu","We","Th","Fr","Sa","Su"]
            for i, d in enumerate(days):
                tk.Label(grid_frame, text=d, font=("Segoe UI",8,"bold"),
                         bg="white", fg=C_DARK, width=3).grid(row=0, column=i, padx=1, pady=1)
            import calendar as _calmod
            cal = _calmod.Calendar(firstweekday=0)
            row = 1
            for week in cal.monthdayscalendar(y, m):
                for col, day in enumerate(week):
                    if day == 0:
                        tk.Label(grid_frame, text="", bg="white", width=3).grid(row=row, column=col)
                        continue
                    this_date = datetime.date(y, m, day)
                    is_today = (this_date == datetime.date.today())
                    btn = tk.Button(grid_frame, text=str(day), width=3, relief="flat",
                                     bg=(C_ORANGE if is_today else "white"),
                                     fg=("white" if is_today else C_DARK),
                                     font=("Segoe UI",9, "bold" if is_today else "normal"),
                                     command=lambda d=this_date: _pick(d))
                    btn.grid(row=row, column=col, padx=1, pady=1)
                row += 1

        _draw()
        # Today shortcut
        bottom = tk.Frame(top, bg="white"); bottom.pack(fill="x", pady=(0,6))
        tk.Button(bottom, text="Aaj ki Date", bg=C_TEAL, fg="white", relief="flat",
                  command=lambda: _pick(datetime.date.today()),
                  font=("Segoe UI",9,"bold"), pady=4).pack(fill="x", padx=8)

        top.update_idletasks()
        x = self.winfo_rootx()
        y = self.winfo_rooty() + self.winfo_height()
        top.geometry(f"+{x}+{y}")

def make_date_entry(parent, textvariable, width=13, bg="white"):
    return DatePicker(parent, textvariable, width=width, bg=bg)

# ─── LICENSE SCREENS ──────────────────────────────────────────────────────────
def _show_activate_window():
    root = tk.Tk()
    root.title("BhugtanEase — Activate")
    root.resizable(False, False)
    root.configure(bg="#FFF3E0")
    W, H = 560, 460
    sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
    root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
    root.protocol("WM_DELETE_WINDOW", root.destroy)

    # Header
    hdr = tk.Frame(root, bg=C_ORANGE, height=100); hdr.pack(fill="x"); hdr.pack_propagate(False)
    _act_logo = _get_logo_image(52)
    if _act_logo:
        root._act_logo_ref = _act_logo
        tk.Label(hdr, image=_act_logo, bg=C_ORANGE).place(relx=0.5, rely=0.5, anchor="center")
    else:
        tk.Label(hdr, text="🚗  BhugtanEase", font=("Segoe UI", 22, "bold"),
                 bg=C_ORANGE, fg="white").place(relx=0.5, rely=0.4, anchor="center")
    sub = tk.Frame(root, bg=C_DARK, height=28); sub.pack(fill="x"); sub.pack_propagate(False)
    tk.Label(sub, text="Car Paint Billing & Color Matching Software",
             font=("Segoe UI", 10), bg=C_DARK, fg="#FFCCBC").place(relx=0.5, rely=0.5, anchor="center")

    tk.Frame(root, bg="#FFF3E0", height=16).pack()
    card = tk.Frame(root, bg="white", padx=30, pady=24, relief="groove", bd=1)
    card.pack(fill="x", padx=24)

    tk.Label(card, text="🔑 Software Activate Karein", font=("Segoe UI", 13, "bold"),
             bg="white", fg="#111").pack(anchor="w")
    tk.Label(card, text="Serial Format: BE-XXXX-XXXX-XXXX-XXXX",
             font=("Segoe UI", 9), bg="white", fg="#666").pack(anchor="w", pady=(3,12))

    tk.Label(card, text="👤 Shop/Owner ka Naam:", font=("Segoe UI", 9),
             bg="white", fg="#333").pack(anchor="w")
    v_name = tk.StringVar()
    e_name = tk.Entry(card, textvariable=v_name, font=("Segoe UI", 11),
                      relief="solid", bd=1)
    e_name.pack(fill="x", ipady=6, pady=(2,10))
    e_name.focus_set()

    tk.Label(card, text="🔢 Serial Number:", font=("Segoe UI", 9),
             bg="white", fg="#333").pack(anchor="w")
    v_serial = tk.StringVar()
    e_serial = tk.Entry(card, textvariable=v_serial, font=("Courier New", 12),
                        relief="solid", bd=1)
    e_serial.pack(fill="x", ipady=7, pady=(2,4))

    msg_var = tk.StringVar()
    tk.Label(card, textvariable=msg_var, font=("Segoe UI", 9, "bold"),
             bg="white", fg=C_RED, wraplength=460).pack(anchor="w", pady=(4,0))

    def activate(ev=None):
        nm = v_name.get().strip(); sr = v_serial.get().strip().upper()
        if not nm: msg_var.set("❌ Naam zaroori hai!"); return
        if not sr: msg_var.set("❌ Serial number daalo!"); return
        if _activate_license(sr, nm):
            root.destroy()
        else:
            msg_var.set("❌ Serial ya naam galat hai! Dealer se check karein.")

    e_name.bind("<Return>", lambda e: e_serial.focus_set())
    e_serial.bind("<Return>", activate)
    tk.Frame(card, bg="white", height=8).pack()
    make_btn(card, "✔  ACTIVATE KARO", activate, bg=C_ORANGE, pady=12,
             font=("Segoe UI", 12, "bold")).pack(fill="x")

    tk.Label(root, text="Serial lene ke liye apne dealer se sampark karein.",
             font=("Segoe UI", 8), bg="#FFF3E0", fg="#999").pack(pady=10)
    root.mainloop()

class LicenseExpiredWin:
    def __init__(self, days_left, customer):
        self._root = tk.Tk()
        self._root.title("BhugtanEase — License Expired")
        self._root.configure(bg="#1A1A2E")
        self._root.resizable(False, False)
        w, h = 560, 560
        sw, sh = self._root.winfo_screenwidth(), self._root.winfo_screenheight()
        self._root.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")
        self._root.protocol("WM_DELETE_WINDOW", self._root.destroy)

        tk.Frame(self._root, bg="#C53030", height=3).pack(fill="x")
        tk.Label(self._root, text="⏰  License Expire Ho Gayi!",
                 font=("Segoe UI", 15, "bold"), bg="#1A1A2E", fg="#FC8181").pack(pady=(18,2))
        tk.Label(self._root, text=f"Aapki license {abs(days_left)} din pehle expire hui.",
                 font=("Segoe UI", 10), bg="#1A1A2E", fg="#A0AEC0").pack()
        tk.Label(self._root, text="Software use karne ke liye naya serial number daalo.",
                 font=("Segoe UI", 10), bg="#1A1A2E", fg="#A0AEC0").pack(pady=(2,10))

        card = tk.Frame(self._root, bg="#2D3748", padx=24, pady=14)
        card.pack(fill="x", padx=28)
        tk.Label(card, text="👤 Aapka Naam:", font=("Segoe UI",9), bg="#2D3748", fg="#A0AEC0", anchor="w").pack(fill="x", pady=(8,2))
        self._customer_var = tk.StringVar()
        name_entry = tk.Entry(card, textvariable=self._customer_var, font=("Segoe UI",11),
                              bg="#1A202C", fg="#FBD38D", insertbackground="#FBD38D", relief="flat")
        name_entry.pack(fill="x", ipady=6); name_entry.focus_set()

        tk.Label(card, text="🔑 Naya Serial (BE-XXXX-XXXX-XXXX-XXXX):", font=("Segoe UI",9),
                 bg="#2D3748", fg="#A0AEC0", anchor="w").pack(fill="x", pady=(10,2))
        self._serial_var = tk.StringVar()
        serial_entry = tk.Entry(card, textvariable=self._serial_var, font=("Courier New",13,"bold"),
                                bg="#1A202C", fg="#68D391", insertbackground="#68D391", relief="flat")
        serial_entry.pack(fill="x", ipady=8)

        self._msg_var = tk.StringVar()
        tk.Label(self._root, textvariable=self._msg_var, font=("Segoe UI",9,"bold"),
                 bg="#1A1A2E", fg="#FC8181", wraplength=480).pack(pady=(10,4))

        name_entry.bind("<Return>", lambda e: serial_entry.focus_set())
        serial_entry.bind("<Return>", lambda e: self._do_activate())
        tk.Button(self._root, text="🔓  Activate Karo", font=("Segoe UI",11,"bold"),
                  bg=C_GREEN, fg="white", relief="flat", cursor="hand2",
                  padx=20, pady=9, bd=0, command=self._do_activate).pack(pady=(4,0))
        tk.Frame(self._root, bg="#C53030", height=3).pack(fill="x", side="bottom")
        self._root.mainloop()

    def _do_activate(self):
        customer = self._customer_var.get().strip()
        serial   = self._serial_var.get().strip().upper()
        if not customer: self._msg_var.set("❌ Naam daalo!"); return
        if not serial:   self._msg_var.set("❌ Serial daalo!"); return
        if _activate_license(serial, customer):
            self._root.destroy()
        else:
            self._msg_var.set("❌ Serial ya naam galat hai!")

# ─── LOGIN ────────────────────────────────────────────────────────────────────
class LoginWin:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("BhugtanEase — Login")
        self.root.resizable(False, False)
        self.root.configure(bg="#FFF3E0")
        W, H = 420, 500
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{W}x{H}+{(sw-W)//2}+{(sh-H)//2}")
        self.root.protocol("WM_DELETE_WINDOW", self.root.destroy)

        hdr = tk.Frame(self.root, bg=C_ORANGE, height=100); hdr.pack(fill="x"); hdr.pack_propagate(False)
        self._login_logo = _get_logo_image(52)
        if self._login_logo:
            tk.Label(hdr, image=self._login_logo, bg=C_ORANGE).place(relx=0.5, rely=0.5, anchor="center")
        else:
            tk.Label(hdr, text="🚗", font=("Segoe UI",28), bg=C_ORANGE, fg="white").place(relx=0.5, rely=0.35, anchor="center")
            tk.Label(hdr, text="BhugtanEase", font=("Segoe UI",20,"bold"), bg=C_ORANGE, fg="white").place(relx=0.5, rely=0.72, anchor="center")

        sub = tk.Frame(self.root, bg=C_DARK, height=26); sub.pack(fill="x"); sub.pack_propagate(False)
        tk.Label(sub, text="Billing Software",
                 font=("Segoe UI",9), bg=C_DARK, fg="#FFCCBC").place(relx=0.5, rely=0.5, anchor="center")

        tk.Frame(self.root, bg="#FFF3E0", height=20).pack()
        card = tk.Frame(self.root, bg="white", padx=32, pady=28, relief="groove", bd=1)
        card.pack(fill="x", padx=28)

        tk.Label(card, text="Username:", font=("Segoe UI",10), bg="white", fg="#333", anchor="w").pack(fill="x")
        self.v_user = tk.StringVar(value="admin")
        e_user = tk.Entry(card, textvariable=self.v_user, font=("Segoe UI",12), relief="solid", bd=1)
        e_user.pack(fill="x", ipady=6, pady=(2,12)); e_user.focus_set()

        tk.Label(card, text="Password:", font=("Segoe UI",10), bg="white", fg="#333", anchor="w").pack(fill="x")
        self.v_pass = tk.StringVar()
        e_pass = tk.Entry(card, textvariable=self.v_pass, font=("Segoe UI",12), show="*", relief="solid", bd=1)
        e_pass.pack(fill="x", ipady=6, pady=(2,4))

        self.msg = tk.StringVar()
        tk.Label(card, textvariable=self.msg, font=("Segoe UI",9,"bold"),
                 bg="white", fg=C_RED).pack(pady=(6,0))

        tk.Frame(card, bg="white", height=8).pack()
        make_btn(card, "LOGIN  →", self.do_login, bg=C_ORANGE, pady=12,
                 font=("Segoe UI",12,"bold")).pack(fill="x")

        tk.Label(card, text="Ashrisha Ecommerce Solution Pvt Ltd",
                 font=("Segoe UI", 9, "bold"), bg="white", fg="#555555").pack(pady=(10,4))

        e_user.bind("<Return>", lambda e: e_pass.focus_set())
        e_pass.bind("<Return>", lambda e: self.do_login())

        # License info bottom
        status, days_left, _, customer = _get_license_info()
        info_txt = f"License: {customer} | {days_left} din baaki" if status == 'ok' else ""
        tk.Label(self.root, text=info_txt, font=("Segoe UI",8),
                 bg="#FFF3E0", fg="#999").pack(pady=10)
        self.root.mainloop()

    def do_login(self):
        u = self.v_user.get().strip(); p = self.v_pass.get().strip()
        conn = get_db()
        row = conn.execute("SELECT * FROM users WHERE username=? AND password=?", (u, p)).fetchone()
        conn.close()
        if row:
            self.root.destroy()
            MainApp()
        else:
            self.msg.set("❌ Username ya Password galat hai!")

# ─── MAIN APPLICATION ──────────────────────────────────────────────────────────

def _show_notify_dialog(root_win, bill_no, party, mobile, email, grand):
    """Bill save ke baad WhatsApp/Email notification dialog"""
    dlg = tk.Toplevel(root_win)
    dlg.title("Notification Bhejein")
    dlg.resizable(False, False)
    dlg.transient(root_win)
    dlg.grab_set()
    dlg.configure(bg="white")
    dlg.geometry("420x320")

    shop = get_shop()
    msg_text = (
        f"Namaskar {party}!\n\n"
        f"Aapka bill {bill_no} generate ho gaya hai.\n"
        f"Amount: {grand}\n\n"
        f"Thank you for shopping with {shop.get('name','BhugtanEase Paint Shop')}!\n"
        f"Contact: {shop.get('mobile','')}\n"
        f"-- BhugtanEase"
    )

    tk.Label(dlg, text=f"Bill {bill_no} — Notification", font=("Segoe UI",11,"bold"),
             bg="white", fg=C_DARK).pack(pady=(14,4))
    tk.Label(dlg, text=f"Party: {party}  |  Mobile: {mobile or 'N/A'}  |  Email: {email or 'N/A'}",
             font=("Segoe UI",9), bg="white", fg="#666").pack()

    msg_frame = tk.LabelFrame(dlg, text="Message Preview", font=("Segoe UI",9),
                               bg="white", padx=8, pady=4)
    msg_frame.pack(fill="x", padx=16, pady=8)
    msg_txt = tk.Text(msg_frame, height=5, font=("Segoe UI",9), wrap="word",
                       relief="flat", bg="#f9f9f9")
    msg_txt.insert("1.0", msg_text)
    msg_txt.pack(fill="x")

    def get_msg():
        return msg_txt.get("1.0", "end").strip()

    btn_frame = tk.Frame(dlg, bg="white"); btn_frame.pack(pady=10)

    def send_whatsapp():
        mob = mobile.replace(" ","").replace("+","").replace("-","")
        if not mob:
            messagebox.showwarning("","Party ka mobile number nahi hai!", parent=dlg); return
        if not mob.startswith("91"):
            mob = "91" + mob
        import urllib.parse, webbrowser
        msg = urllib.parse.quote(get_msg())
        webbrowser.open(f"https://wa.me/{mob}?text={msg}")

    def send_email():
        if not email:
            messagebox.showwarning("","Party ka email nahi hai!", parent=dlg); return
        import urllib.parse, webbrowser
        subj = urllib.parse.quote(f"Bill {bill_no} - {shop.get('name','')}")
        body = urllib.parse.quote(get_msg())
        webbrowser.open(f"mailto:{email}?subject={subj}&body={body}")

    def copy_msg():
        dlg.clipboard_clear()
        dlg.clipboard_append(get_msg())
        messagebox.showinfo("Copied!","Message clipboard mein copy ho gaya!", parent=dlg)

    make_btn(btn_frame, "💬 WhatsApp", send_whatsapp, bg="#25D366", pady=6).pack(side="left", padx=6)
    make_btn(btn_frame, "📧 Email",    send_email,    bg=C_BLUE,    pady=6).pack(side="left", padx=6)
    make_btn(btn_frame, "📋 Copy Msg", copy_msg,      bg=C_GRAY,    pady=6).pack(side="left", padx=6)
    make_btn(btn_frame, "✖ Skip",      dlg.destroy,   bg=C_RED,     pady=6).pack(side="left", padx=6)

class MainApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("🚗 BhugtanEase — Car Paint Billing Software")
        self.root.configure(bg=C_BG)
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        self.root.geometry(f"{min(1400,sw)}x{min(800,sh)}+0+0")
        self.root.state("zoomed")

        self._build_header()
        self._build_sidebar()
        self._build_content()
        self._show_dashboard()

        def _on_close():
            do_backup(silent=True)          # local auto-backup
            do_drive_backup(silent=True)    # Drive auto-backup (only if path set)
            self.root.destroy()
        self.root.protocol("WM_DELETE_WINDOW", _on_close)

        self.root.mainloop()

    def _build_header(self):
        hdr = tk.Frame(self.root, bg=C_ORANGE, height=34); hdr.pack(fill="x"); hdr.pack_propagate(False)
        # Logo
        _hdr_logo = _get_logo_image(22)
        if _hdr_logo:
            self._hdr_logo_ref = _hdr_logo
            tk.Label(hdr, image=_hdr_logo, bg=C_ORANGE).pack(side="left", padx=(8,4), pady=3)
        else:
            tk.Label(hdr, text="🚗", font=("Segoe UI",11,"bold"), bg=C_ORANGE, fg="white").pack(side="left", padx=(8,4))
        shop = get_shop()
        tk.Label(hdr, text=f"🏪 {shop['name']}",
                 font=("Segoe UI",11), bg=C_ORANGE, fg="#FFCCBC").pack(side="left", padx=16)
        today_str = datetime.date.today().strftime("%d %b %Y")
        tk.Label(hdr, text=f"📅 {today_str}",
                 font=("Segoe UI",10), bg=C_ORANGE, fg="#FFF8E1").pack(side="right", padx=16)

    def _build_sidebar(self):
        sb = tk.Frame(self.root, bg=C_DARK, width=190); sb.pack(side="left", fill="y"); sb.pack_propagate(False)
        self.sidebar = sb

        menus = [
            ("🏠  Dashboard",        self._show_dashboard),
            ("🧾  Sale Bill",         self._show_sale_bill),
            ("🛒  Purchase Entry",    self._show_purchase),
            ("📋  Sale History",      self._show_sale_history),
            ("🔄  Sale Return",       self._show_sale_return),
            ("📦  Purchase History",  self._show_purchase_history),
            ("🔃  Purchase Return",   self._show_purchase_return),
            ("🚗  Color Formulas",    self._show_formulas),
            ("📊  Stock Register",    self._show_stock),
            ("📁  Products",          self._show_products),
            ("👥  Parties",           self._show_parties),
            ("📈  Reports",           self._show_reports),
            ("🔧  Settings",          self._show_settings),
            ("💸  Expenses",           self._show_expenses),
        ]
        for label, cmd in menus:
            b = tk.Button(sb, text=label, command=cmd, bg=C_DARK, fg="white",
                          font=("Segoe UI", 10), relief="flat", bd=0, cursor="hand2",
                          anchor="w", padx=16, pady=8,
                          activebackground=C_ORANGE, activeforeground="white")
            b.pack(fill="x")
            b.bind("<Enter>", lambda e, btn=b: btn.config(bg=C_ORANGE))
            b.bind("<Leave>", lambda e, btn=b: btn.config(bg=C_DARK))
        tk.Frame(sb, bg=C_ORANGE, height=1).pack(fill="x", side="bottom", pady=0)

    def _build_content(self):
        self.content = tk.Frame(self.root, bg=C_BG); self.content.pack(side="left", fill="both", expand=True)

    def _clear(self):
        for w in self.content.winfo_children(): w.destroy()

    def _section_header(self, title, subtitle=""):
        frm = tk.Frame(self.content, bg=C_ORANGE, height=46); frm.pack(fill="x"); frm.pack_propagate(False)
        tk.Label(frm, text=title, font=("Segoe UI",14,"bold"), bg=C_ORANGE, fg="white").pack(side="left", padx=16, pady=8)
        if subtitle:
            tk.Label(frm, text=subtitle, font=("Segoe UI",9), bg=C_ORANGE, fg="#FFCCBC").pack(side="left")

    # ── DASHBOARD ─────────────────────────────────────────────────────────────
    def _show_dashboard(self):
        self._clear()
        self._section_header("🏠 Dashboard", "Aaj ka overview")
        conn = get_db()
        today = datetime.date.today().isoformat()
        month_start = datetime.date.today().replace(day=1).isoformat()

        today_sale = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM sales WHERE bill_date=?", (today,)).fetchone()["t"]
        month_sale = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM sales WHERE bill_date>=?", (month_start,)).fetchone()["t"]
        total_bills = conn.execute("SELECT COUNT(*) as c FROM sales").fetchone()["c"]
        total_products = conn.execute("SELECT COUNT(*) as c FROM products").fetchone()["c"]
        total_formulas = conn.execute("SELECT COUNT(*) as c FROM color_formulas").fetchone()["c"]
        total_parties = conn.execute("SELECT COUNT(*) as c FROM parties").fetchone()["c"]

        # Low stock
        low_stock = []
        prods = conn.execute("SELECT name, unit, low_stock_alert FROM products").fetchall()
        for p in prods:
            stk = get_stock(p["name"])
            if stk <= p["low_stock_alert"]:
                low_stock.append((p["name"], stk, p["unit"]))

        # Expiring / expired stock (batch-wise, purchase entry me daali gayi expiry date se)
        horizon = (datetime.date.today() + datetime.timedelta(days=60)).isoformat()
        expiring_batches = conn.execute("""
            SELECT product, batch_no, qty, unit, expiry_date
            FROM purchase_items
            WHERE expiry_date != '' AND expiry_date <= ?
            ORDER BY expiry_date ASC
        """, (horizon,)).fetchall()
        conn.close()

        # Cards row
        cards_frame = tk.Frame(self.content, bg=C_BG); cards_frame.pack(fill="x", padx=16, pady=16)
        cards_data = [
            ("💰 Aaj Ki Bikri",  rupee(today_sale), C_ORANGE),
            ("📅 Mahine Ki Bikri",rupee(month_sale),  C_TEAL),
            ("🧾 Total Bills",    str(total_bills),   C_BLUE),
            ("📁  Products",       str(total_products),C_PURPLE),
            ("🚗 Color Formulas", str(total_formulas),C_DARK),
            ("👥 Parties",        str(total_parties), C_GREEN),
        ]
        for i, (lbl, val, bg) in enumerate(cards_data):
            card = tk.Frame(cards_frame, bg=bg, padx=18, pady=14, relief="flat")
            card.grid(row=0, column=i, padx=8, sticky="ew")
            cards_frame.grid_columnconfigure(i, weight=1)
            tk.Label(card, text=val, font=("Segoe UI",18,"bold"), bg=bg, fg="white").pack()
            tk.Label(card, text=lbl, font=("Segoe UI",9), bg=bg, fg="#FFE0B2").pack()

        # Quick actions
        qa_frame = tk.Frame(self.content, bg=C_BG); qa_frame.pack(fill="x", padx=16, pady=(0,12))
        tk.Label(qa_frame, text="⚡ Quick Actions:", font=("Segoe UI",11,"bold"),
                 bg=C_BG, fg=C_DARK).pack(side="left", padx=4)
        make_btn(qa_frame, "🧾 New Sale Bill", self._show_sale_bill, bg=C_ORANGE).pack(side="left", padx=4)
        make_btn(qa_frame, "🚗 Color Formulas",self._show_formulas, bg=C_TEAL).pack(side="left", padx=4)
        make_btn(qa_frame, "📦 Check Stock",   self._show_stock,   bg=C_BLUE).pack(side="left", padx=4)
        make_btn(qa_frame, "🛒 Purchase Entry",self._show_purchase, bg=C_GREEN).pack(side="left", padx=4)

        # Low stock alerts
        if low_stock:
            alert_frame = tk.LabelFrame(self.content, text="⚠️ Low Stock Alert",
                                        font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_RED,
                                        padx=10, pady=8)
            alert_frame.pack(fill="x", padx=16, pady=(0,10))
            for name, stk, unit in low_stock[:8]:
                tk.Label(alert_frame, text=f"🔴 {name}: {stk:.2f} {unit} baaki",
                         font=("Segoe UI",9), bg=C_BG, fg=C_RED).pack(anchor="w")

        # Expiry alerts (batch-wise)
        if expiring_batches:
            exp_frame = tk.LabelFrame(self.content, text="⏳ Expiry Alert (Batch-wise Stock)",
                                      font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_RED,
                                      padx=10, pady=8)
            exp_frame.pack(fill="x", padx=16, pady=(0,10))
            today_d = datetime.date.today()
            for r in expiring_batches[:10]:
                try:
                    ed = datetime.date.fromisoformat(r["expiry_date"])
                    dleft = (ed - today_d).days
                except Exception:
                    dleft = None
                batch_txt = f" (Batch: {r['batch_no']})" if r["batch_no"] else ""
                if dleft is not None and dleft < 0:
                    status, color = f"❌ {abs(dleft)} din pehle EXPIRE ho gaya", C_RED
                elif dleft is not None and dleft <= 15:
                    status, color = f"⚠️ {dleft} din me expire hoga", C_ORANGE
                elif dleft is not None:
                    status, color = f"🟡 {dleft} din baaki", C_DARK
                else:
                    status, color = "", C_DARK
                tk.Label(exp_frame,
                         text=f"{r['product']}{batch_txt}: {r['qty']:.2f} {r['unit']} — {status} ({iso_to_mmyy(r['expiry_date'])})",
                         font=("Segoe UI",9), bg=C_BG, fg=color).pack(anchor="w")
            if len(expiring_batches) > 10:
                tk.Label(exp_frame, text=f"... aur {len(expiring_batches)-10} batches",
                         font=("Segoe UI",8,"italic"), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(2,0))

        # Recent bills
        recent_frame = tk.LabelFrame(self.content, text="📋 Recent Bills",
                                     font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_DARK,
                                     padx=8, pady=6)
        recent_frame.pack(fill="both", expand=True, padx=16, pady=(0,12))
        cols = ("Bill No","Date","Party","Amount","Mode")
        tv = ttk.Treeview(recent_frame, columns=cols, show="headings", height=8)
        for c, w in zip(cols, [130,90,180,100,80]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb2 = ttk.Scrollbar(recent_frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb2.set)
        sb2.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)
        conn2 = get_db()
        rows = conn2.execute("SELECT bill_no,bill_date,party,grand_total,pay_mode FROM sales ORDER BY id DESC LIMIT 20").fetchall()
        conn2.close()
        for r in rows:
            tv.insert("", "end", values=(r["bill_no"], r["bill_date"], r["party"],
                                          rupee(r["grand_total"]), r["pay_mode"]))

    # ── SALE BILL ──────────────────────────────────────────────────────────────
    def _show_sale_bill(self, prefill_items=None):
        self._clear()
        self._section_header("🧾 New Sale Bill", "GST Bill / Non-GST Bill")

        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # ── TOP: Bill details ──
        top = tk.LabelFrame(main, text="Bill Details", font=("Segoe UI",9,"bold"),
                            bg=C_BG, fg=C_DARK, padx=8, pady=6)
        top.pack(fill="x", pady=(0,6))

        # Row 1
        r1 = tk.Frame(top, bg=C_BG); r1.pack(fill="x")
        tk.Label(r1, text="Bill No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_bill = tk.StringVar(value=next_bill_no(commit=False))
        ttk.Entry(r1, textvariable=v_bill, width=18).pack(side="left", padx=(4,16))
        tk.Label(r1, text="Date:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_date = tk.StringVar()
        make_date_entry(r1, v_date, width=12, bg=C_BG).pack(side="left", padx=(4,16))
        tk.Label(r1, text="Bill Type:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_gst_type = tk.StringVar(value="GST")
        bill_type_cb = ttk.Combobox(r1, textvariable=v_gst_type, width=13, state="readonly",
                     values=["GST","Non-GST"])
        bill_type_cb.pack(side="left", padx=(4,16))

        def _on_billtype_change(ev=None):
            pass
        bill_type_cb.bind("<<ComboboxSelected>>", _on_billtype_change)
        tk.Label(r1, text="Payment:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_pay = tk.StringVar(value="Cash")
        ttk.Combobox(r1, textvariable=v_pay, width=10, state="readonly",
                     values=["Cash","Online/UPI","Card","Credit","Cheque"]).pack(side="left", padx=4)

        # Row 2 — Party
        r2 = tk.Frame(top, bg=C_BG); r2.pack(fill="x", pady=(6,0))
        tk.Label(r2, text="Customer Name:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_party = tk.StringVar()
        conn = get_db()
        all_parties = [r["name"] for r in conn.execute("SELECT name FROM parties WHERE ptype='Customer' OR ptype='Both' ORDER BY name").fetchall()]
        conn.close()
        party_cb = ttk.Combobox(r2, textvariable=v_party, width=28, values=all_parties)
        party_cb.pack(side="left", padx=(4,16))
        tk.Label(r2, text="Mobile:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_mobile = tk.StringVar()
        ttk.Entry(r2, textvariable=v_mobile, width=14).pack(side="left", padx=(4,16))
        tk.Label(r2, text="GSTIN:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_gstin = tk.StringVar()
        ttk.Entry(r2, textvariable=v_gstin, width=18).pack(side="left", padx=4)

        def party_selected(ev=None):
            nm = v_party.get().strip()
            conn2 = get_db()
            row = conn2.execute("SELECT * FROM parties WHERE name=?", (nm,)).fetchone()
            conn2.close()
            if row:
                v_mobile.set(row["mobile"] or "")
                v_gstin.set(row["gstin"] or "")
        party_cb.bind("<<ComboboxSelected>>", party_selected)

        # Enter navigation: Bill No → Date → BillType → Payment → Customer → Mobile → GSTIN → Product
        # r1 widgets: BillNo Entry, Date Entry, BillType CB, Payment CB
        # r2 widgets: Customer CB, Mobile Entry, GSTIN Entry
        # Then jump to prod_cb (item row)
        # We bind after prod_cb is defined below

        # ── ITEMS TABLE ──
        items_frame = tk.LabelFrame(main, text="Items", font=("Segoe UI",9,"bold"),
                                    bg=C_BG, fg=C_DARK, padx=6, pady=6)
        items_frame.pack(fill="both", expand=True, pady=(0,6))

        # Item entry row
        entry_row = tk.Frame(items_frame, bg=C_BG); entry_row.pack(fill="x", pady=(0,6))

        tk.Label(entry_row, text="Product:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn2 = get_db()
        all_products = [r["name"] for r in conn2.execute("SELECT name FROM products ORDER BY name").fetchall()]
        conn2.close()
        v_prod = tk.StringVar()
        prod_cb = ttk.Combobox(entry_row, textvariable=v_prod, width=30, values=all_products)
        prod_cb.pack(side="left", padx=(4,8))

        tk.Label(entry_row, text="Shade:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_shade = tk.StringVar()
        shade_e = ttk.Entry(entry_row, textvariable=v_shade, width=14); shade_e.pack(side="left", padx=(4,8))

        tk.Label(entry_row, text="Qty:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_qty = tk.StringVar(value="1")
        qty_e = ttk.Entry(entry_row, textvariable=v_qty, width=7); qty_e.pack(side="left", padx=(4,4))
        v_unit = tk.StringVar(value="Ltr")
        ttk.Combobox(entry_row, textvariable=v_unit, width=5, state="readonly",
                     values=["Ltr","ML","Kg","Pcs"]).pack(side="left", padx=(0,8))

        tk.Label(entry_row, text="Rate:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_rate = tk.StringVar(value="0")
        rate_e = ttk.Entry(entry_row, textvariable=v_rate, width=9); rate_e.pack(side="left", padx=(4,8))

        tk.Label(entry_row, text="GST%:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_gst = tk.StringVar(value="18")
        gst_cb = ttk.Combobox(entry_row, textvariable=v_gst, width=5, state="readonly",
                     values=["0","5","12","18","28"]); gst_cb.pack(side="left", padx=(4,8))

        # Add + Remove + Formula buttons same row mein (rate/gst ke saath)
        _sale_add_btn = make_btn(entry_row, "➕ Add", None, bg=C_ORANGE, pady=3); _sale_add_btn.pack(side="left", padx=(8,4))
        _sale_rem_btn = make_btn(entry_row, "🗑️ Remove", None, bg=C_RED, pady=3); _sale_rem_btn.pack(side="left", padx=4)
        _sale_frm_btn = make_btn(entry_row, "🚗 Formula", None, bg=C_TEAL, pady=3); _sale_frm_btn.pack(side="left", padx=4)

        def product_selected(ev=None):
            nm = v_prod.get().strip()
            conn3 = get_db()
            row = conn3.execute("SELECT * FROM products WHERE name=?", (nm,)).fetchone()
            conn3.close()
            if row:
                v_rate.set(str(row["sale_rate"]))
                v_unit.set(row["unit"] or "Ltr")
                v_gst.set(str(int(row["gst_percent"])))
        prod_cb.bind("<<ComboboxSelected>>", product_selected)

        # ── Sale Bill Header Enter Navigation ──
        # Bill No Entry (r1 row 1st widget) → Date Entry → BillType CB → Payment CB
        # → Customer CB → Mobile Entry → GSTIN Entry → prod_cb (item section)
        r1_widgets = [w for w in r1.winfo_children() if isinstance(w, (ttk.Entry, ttk.Combobox))]
        r2_widgets = [w for w in r2.winfo_children() if isinstance(w, (ttk.Entry, ttk.Combobox))]
        # r1: BillNo entry, date entry (inside DatePicker Frame - skip frames), BillType CB, Payment CB
        # collect only direct Entry/CB, DatePicker is a Frame so handle separately
        _hdr_nav = []
        for w in r1.winfo_children():
            if isinstance(w, (ttk.Entry, ttk.Combobox)):
                _hdr_nav.append(w)
            elif isinstance(w, tk.Frame):   # DatePicker frame
                for c in w.winfo_children():
                    if isinstance(c, (ttk.Entry, ttk.Combobox)):
                        _hdr_nav.append(c)
        for w in r2.winfo_children():
            if isinstance(w, (ttk.Entry, ttk.Combobox)):
                _hdr_nav.append(w)
        # last header field → jump to product entry
        _hdr_nav.append(prod_cb)
        bind_enter_nav(_hdr_nav)

        # Enter navigation for item entry row:
        # Product → Shade → Qty → Rate → GST% → Add Item (trigger)
        # Enter navigation - Sale Bill items
        sale_nav_order = [prod_cb, shade_e, qty_e, rate_e, gst_cb]
        def _sale_enter(idx):
            def _fn(ev):
                if idx == len(sale_nav_order) - 1:
                    add_item()
                    self.root.after(10, lambda: prod_cb.focus_force())
                else:
                    nxt = sale_nav_order[idx + 1]
                    nxt.focus_force()
                    try: nxt.select_range(0, "end")
                    except: pass
                return "break"
            return _fn
        for _i, _w in enumerate(sale_nav_order):
            _w.bind("<Return>", _sale_enter(_i))
            _w.bind("<KP_Enter>", _sale_enter(_i))

        # Items treeview
        cols = ("Product","Shade","Qty","Unit","Rate","GST%","GST Amt","Total")
        tv = ttk.Treeview(items_frame, columns=cols, show="headings", height=6)
        for c, w in zip(cols, [220,100,60,50,80,50,80,90]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb3 = ttk.Scrollbar(items_frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb3.set)
        sb3.pack(side="right", fill="y"); tv.pack(fill="both", expand=False)

        self._bill_items = []

        def add_item():
            try:
                qty  = float(v_qty.get() or 0)
                rate = float(v_rate.get() or 0)
                gst_pct = float(v_gst.get() or 0) if v_gst_type.get() == "GST" else 0
                if not v_prod.get().strip(): messagebox.showwarning("Error","Product chunein!"); return
                if qty <= 0: messagebox.showwarning("Error","Qty daalen!"); return
                taxable = round(qty * rate, 2)
                gst_amt = round(taxable * gst_pct / 100, 2)
                total   = round(taxable + gst_amt, 2)
                item = {
                    "product": v_prod.get().strip(),
                    "shade":   v_shade.get().strip(),
                    "qty":     qty,
                    "unit":    v_unit.get(),
                    "rate":    rate,
                    "gst_pct": gst_pct,
                    "gst_amt": gst_amt,
                    "taxable": taxable,
                    "total":   total,
                }
                self._bill_items.append(item)
                tv.insert("", "end", values=(item["product"], item["shade"],
                          f"{qty:.2f}", item["unit"], f"{rate:.2f}",
                          f"{gst_pct:.0f}%", f"{gst_amt:.2f}", f"{total:.2f}"))
                v_prod.set(""); v_shade.set(""); v_qty.set("1"); v_rate.set("0")
                self._update_totals(tv_totals, disc_var)
            except ValueError:
                messagebox.showwarning("Error","Qty ya Rate sahi daalen!")

        def remove_item():
            sel = tv.selection()
            if not sel: return
            idx = tv.index(sel[0])
            tv.delete(sel[0])
            self._bill_items.pop(idx)
            self._update_totals(tv_totals, disc_var)

        def from_formula():
            self._pick_formula(v_prod, v_shade, v_qty, v_unit, v_rate)



        # Wire up inline buttons (defined after add_item/remove_item/from_formula)
        _sale_add_btn.config(command=add_item)
        _sale_rem_btn.config(command=remove_item)
        _sale_frm_btn.config(command=from_formula)

        # GST type change update
        def gst_type_changed(ev=None):
            pass  # Will apply on add
        v_gst_type.trace_add("write", lambda *a: gst_type_changed())

        # ── TOTALS ──
        tot_frame = tk.Frame(main, bg=C_BG); tot_frame.pack(fill="x")

        # Discount
        df = tk.Frame(tot_frame, bg=C_BG); df.pack(side="left", padx=8)
        tk.Label(df, text="Discount:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        disc_var = tk.StringVar(value="0")
        ttk.Entry(df, textvariable=disc_var, width=9).pack(side="left", padx=4)
        disc_type_var = tk.StringVar(value="₹")
        ttk.Combobox(df, textvariable=disc_type_var, values=["₹", "%"],
                     width=3, state="readonly").pack(side="left", padx=(0,4))
        disc_var.trace_add("write", lambda *a: self._update_totals(tv_totals, disc_var, disc_type_var))
        disc_type_var.trace_add("write", lambda *a: self._update_totals(tv_totals, disc_var, disc_type_var))

        # Totals display
        tv_totals = tk.Frame(tot_frame, bg=C_BG); tv_totals.pack(side="right", padx=16)

        self._tot_labels = {}
        for row_lbl, key in [("Taxable:", "taxable"), ("GST:", "gst"),
                               ("Discount:", "disc"), ("Grand Total:", "grand")]:
            rf = tk.Frame(tv_totals, bg=C_BG); rf.pack(anchor="e")
            tk.Label(rf, text=row_lbl, font=("Segoe UI",9,"bold"), bg=C_BG,
                     fg=C_DARK, width=12, anchor="e").pack(side="left")
            lv = tk.StringVar(value="₹0.00")
            fnt = ("Segoe UI",12,"bold") if key=="grand" else ("Segoe UI",10)
            clr = C_ORANGE if key=="grand" else C_DARK
            tk.Label(rf, textvariable=lv, font=fnt, bg=C_BG, fg=clr, width=12, anchor="e").pack(side="left")
            self._tot_labels[key] = lv

        # Notes
        nf = tk.Frame(main, bg=C_BG); nf.pack(fill="x", pady=(6,0))
        tk.Label(nf, text="Notes:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_notes = tk.StringVar()
        ttk.Entry(nf, textvariable=v_notes, width=50).pack(side="left", padx=4)

        # Action buttons
        af = tk.Frame(main, bg=C_BG); af.pack(fill="x", pady=8)

        def save_print(print_bill=True):
            if not self._bill_items:
                messagebox.showwarning("Error","Koi item add nahi kiya!"); return
            try: disc_val = float(disc_var.get() or 0)
            except: disc_val = 0
            taxable = sum(i["taxable"] for i in self._bill_items)
            gst_amt = sum(i["gst_amt"] for i in self._bill_items)
            if disc_type_var.get() == "%":
                disc = round((taxable + gst_amt) * disc_val / 100, 2)
            else:
                disc = disc_val
            grand   = round(taxable + gst_amt - disc, 2)

            # Ab bill no commit karo (tabhi increment hoga)
            next_bill_no(commit=True)
            conn3 = get_db()
            try:
                conn3.execute("""INSERT INTO sales
                    (bill_no,bill_date,party,party_mobile,party_gstin,gst_type,grand_total,discount,pay_mode,notes)
                    VALUES(?,?,?,?,?,?,?,?,?,?)""",
                    (v_bill.get(), v_date.get(), v_party.get(), v_mobile.get(),
                     v_gstin.get(), v_gst_type.get(), grand, disc, v_pay.get(), v_notes.get()))
                sale_id = conn3.execute("SELECT id FROM sales WHERE bill_no=?", (v_bill.get(),)).fetchone()["id"]
                for it in self._bill_items:
                    conn3.execute("""INSERT INTO sale_items
                        (sale_id,product,shade_name,qty,unit,rate,taxable,gst_percent,gst_amt,grand)
                        VALUES(?,?,?,?,?,?,?,?,?,?)""",
                        (sale_id, it["product"], it.get("shade",""), it["qty"], it["unit"],
                         it["rate"], it["taxable"], it["gst_pct"], it["gst_amt"], it["total"]))
                conn3.commit()
                if v_pay.get() != "Credit":
                    conn3.execute("""INSERT INTO bill_payments
                        (bill_type,bill_no,party,pay_date,amount,pay_mode,note)
                        VALUES(?,?,?,?,?,?,?)""",
                        ("sale", v_bill.get(), v_party.get(), v_date.get(),
                         grand, v_pay.get(), ""))
                    conn3.commit()
                conn3.close()
                # Notification dialog
                conn_p = get_db()
                p_row = conn_p.execute("SELECT mobile, email FROM parties WHERE name=?",
                                       (v_party.get(),)).fetchone()
                conn_p.close()
                p_mobile = (p_row["mobile"] if p_row else "") or v_mobile.get()
                p_email  = p_row["email"] if p_row else ""
                saved_bill = v_bill.get()
                saved_party = v_party.get()
                if print_bill:
                    self._print_bill(saved_bill)
                _show_notify_dialog(self.root, saved_bill, saved_party,
                                    p_mobile, p_email, rupee(grand))
                self._show_sale_bill()
            except Exception as ex:
                conn3.close()
                messagebox.showerror("Error", str(ex))

        make_btn(af, "💾 Save Bill", lambda: save_print(False), bg=C_BLUE, pady=10).pack(side="left", padx=4)
        make_btn(af, "🖨️ Save & Print", lambda: save_print(True), bg=C_GREEN, pady=10).pack(side="left", padx=4)
        make_btn(af, "🔄 Reset", self._show_sale_bill, bg=C_GRAY, pady=10).pack(side="left", padx=4)

        # Prefill items if coming from formula
        if prefill_items:
            for it in prefill_items:
                self._bill_items.append(it)
                gst_pct = it["gst_pct"] if v_gst_type.get()=="GST" else 0
                tv.insert("","end", values=(it["product"], it.get("shade",""),
                          f"{it['qty']:.2f}", it["unit"], f"{it['rate']:.2f}",
                          f"{gst_pct:.0f}%", f"{it['gst_amt']:.2f}", f"{it['total']:.2f}"))
            self._update_totals(tv_totals, disc_var)

    def _update_totals(self, frame, disc_var, disc_type_var=None):
        taxable = sum(i["taxable"] for i in self._bill_items)
        gst_amt = sum(i["gst_amt"] for i in self._bill_items)
        subtotal = taxable + gst_amt
        try: disc_val = float(disc_var.get() or 0)
        except: disc_val = 0
        if disc_type_var and disc_type_var.get() == "%":
            disc = round(subtotal * disc_val / 100, 2)
        else:
            disc = disc_val
        grand = subtotal - disc
        self._tot_labels["taxable"].set(rupee(taxable))
        self._tot_labels["gst"].set(rupee(gst_amt))
        self._tot_labels["disc"].set(rupee(disc))
        self._tot_labels["grand"].set(rupee(grand))

    def _pick_formula(self, v_prod, v_shade, v_qty, v_unit, v_rate):
        """Formula se product aur shade fill karo"""
        dlg = tk.Toplevel(self.root)
        dlg.title("🚗 Color Formula Chunein")
        dlg.configure(bg=C_BG)
        dlg.geometry("700x500")
        dlg.grab_set()

        tk.Label(dlg, text="🚗 Color Formula Library", font=("Segoe UI",12,"bold"),
                 bg=C_ORANGE, fg="white").pack(fill="x", ipady=10)

        conn = get_db()
        formulas = conn.execute("SELECT * FROM color_formulas ORDER BY shade_name").fetchall()
        conn.close()

        cols = ("Shade Name","Code","Brand","Base Paint","Sale Price","Notes")
        tv = ttk.Treeview(dlg, columns=cols, show="headings", height=14)
        for c, w in zip(cols, [140,80,100,150,90,150]):
            tv.heading(c, text=c); tv.column(c, width=w)
        sb4 = ttk.Scrollbar(dlg, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb4.set)
        sb4.pack(side="right", fill="y"); tv.pack(fill="both", expand=True, padx=8, pady=8)
        self._formulas_data = list(formulas)
        for f in formulas:
            tv.insert("","end", values=(f["shade_name"], f["shade_code"], f["brand"],
                                         f["base_product"], rupee(f["sale_price"]), f["notes"]))

        def use_selected():
            sel = tv.selection()
            if not sel: messagebox.showwarning("","Koi formula chunein!"); return
            idx = tv.index(sel[0])
            f = self._formulas_data[idx]
            v_prod.set(f["base_product"])
            v_shade.set(f["shade_name"])
            v_qty.set(str(f["base_qty"]))
            v_unit.set(f["base_unit"])
            v_rate.set(str(f["sale_price"]))
            dlg.destroy()

        make_btn(dlg, "✔ Is Formula Se Bill Banao", use_selected, bg=C_ORANGE, pady=10).pack(pady=8)

    def _print_bill(self, bill_no):
        conn = get_db()
        bill = conn.execute("SELECT * FROM sales WHERE bill_no=?", (bill_no,)).fetchone()
        if not bill: conn.close(); return
        items = conn.execute("SELECT * FROM sale_items WHERE sale_id=?", (bill["id"],)).fetchall()
        conn.close()
        shop = get_shop()

        dlg = tk.Toplevel(self.root)
        dlg.title(f"Bill — {bill_no}")
        dlg.configure(bg="white")
        dlg.geometry("720x700")

        txt = tk.Text(dlg, font=("Courier New",9), bg="white", wrap="none", padx=10, pady=10)
        sb5 = ttk.Scrollbar(dlg, orient="vertical", command=txt.yview)
        txt.configure(yscrollcommand=sb5.set)
        sb5.pack(side="right", fill="y"); txt.pack(fill="both", expand=True)

        W = 68
        def line(s=""): txt.insert("end", s.center(W) + "\n")
        def rule(c="─"): txt.insert("end", c*W + "\n")
        def lft(s): txt.insert("end", s + "\n")

        rule("═")
        line(shop["name"].upper())
        if shop["address"]: line(shop["address"])
        city_state = ", ".join(filter(None,[shop["city"],shop["state"]]))
        if city_state: line(city_state)
        if shop["mobile"]: line(f"Mob: {shop['mobile']}")
        if shop["gstin"]: line(f"GSTIN: {shop['gstin']}")
        rule("═")

        if bill["gst_type"] == "GST":
            gst_label = "PAKKA BILL  (TAX INVOICE)"
        else:
            gst_label = "KACHCHA BILL  (SALE BILL)"
        line(gst_label)
        rule("─")
        lft(f"Bill No: {bill['bill_no']:<30}Date: {bill['bill_date']}")
        if bill["party"]: lft(f"Customer: {bill['party']}")
        if bill["party_mobile"]: lft(f"Mobile: {bill['party_mobile']}")
        if bill["gst_type"]=="GST" and bill["party_gstin"]: lft(f"GSTIN: {bill['party_gstin']}")
        rule("─")

        # Header
        if bill["gst_type"]=="GST":
            lft(f"{'Product/Shade':<28}{'Qty':>6}{'Rate':>8}{'Taxable':>9}{'GST':>8}{'Total':>9}")
        else:
            lft(f"{'Product/Shade':<28}{'Qty':>6}{'Rate':>8}{'Total':>9}")
        rule("─")

        taxable_total = 0; gst_total = 0; grand_total = 0
        for it in items:
            name = it["product"]
            if it["shade_name"]: name += f" ({it['shade_name']})"
            if len(name) > 26: name = name[:25]+"…"
            qty_str = f"{it['qty']:.2f} {it['unit']}"
            if bill["gst_type"]=="GST":
                lft(f"{name:<28}{qty_str:>7}{it['rate']:>8.2f}{it['taxable']:>9.2f}{it['gst_amt']:>8.2f}{it['grand']:>9.2f}")
            else:
                lft(f"{name:<28}{qty_str:>7}{it['rate']:>8.2f}{it['grand']:>9.2f}")
            taxable_total += it["taxable"]; gst_total += it["gst_amt"]; grand_total += it["grand"]

        rule("─")
        if bill["gst_type"]=="GST":
            lft(f"{'Taxable Amount:':<40}{taxable_total:>9.2f}")
            lft(f"{'CGST (9%):':<40}{gst_total/2:>9.2f}")
            lft(f"{'SGST (9%):':<40}{gst_total/2:>9.2f}")
        if bill["discount"] and float(bill["discount"]) > 0:
            lft(f"{'Discount:':<40}{-float(bill['discount']):>9.2f}")
        rule("═")
        lft(f"{'GRAND TOTAL:':<40}{bill['grand_total']:>9.2f}")
        rule("═")
        lft(f"Amount in Words: {amt_words(bill['grand_total'])}")
        lft(f"Payment Mode: {bill['pay_mode']}")
        rule("─")
        if shop["upi"]: lft(f"UPI: {shop['upi']}")
        if shop["bank"]: lft(f"Bank: {shop['bank']} | A/c: {shop['account']} | IFSC: {shop['ifsc']}")
        if shop["tnc"]: rule("─"); lft(f"Terms: {shop['tnc']}")
        rule("─"); line("Thank you for your business! 🚗"); line("BhugtanEase Software"); rule("═")

        txt.config(state="disabled")

        bf = tk.Frame(dlg, bg="white"); bf.pack(pady=6)
        def do_print():
            try:
                dlg.withdraw()
                dlg.after(200, lambda: (dlg.deiconify(), dlg.lift()))
                import subprocess, tempfile, os as _os
                # Print via notepad (Windows)
                txt_content = txt.get("1.0", "end")
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.txt', mode='w', encoding='utf-8')
                tmp.write(txt_content); tmp.close()
                subprocess.Popen(["notepad", "/p", tmp.name])
            except Exception as e:
                messagebox.showerror("Print Error", f"Print nahi hua: {e}")
        make_btn(bf, "🖨️ Print Karo", do_print, bg=C_GREEN).pack(side="left", padx=6)
        make_btn(bf, "❌ Close", dlg.destroy, bg=C_GRAY).pack(side="left", padx=6)

    # ── PURCHASE ENTRY ─────────────────────────────────────────────────────────
    def _show_purchase(self):
        self._clear()
        self._section_header("🛒 Purchase Entry", "Maal kharidne ka record")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        top = tk.LabelFrame(main, text="Purchase Details", font=("Segoe UI",9,"bold"),
                            bg=C_BG, fg=C_DARK, padx=8, pady=6)
        top.pack(fill="x", pady=(0,6))
        r1 = tk.Frame(top, bg=C_BG); r1.pack(fill="x")
        tk.Label(r1, text="Bill No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_bill = tk.StringVar()
        ttk.Entry(r1, textvariable=v_bill, width=18).pack(side="left", padx=(4,16))
        tk.Label(r1, text="Date:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_date = tk.StringVar()
        make_date_entry(r1, v_date, width=12, bg=C_BG).pack(side="left", padx=(4,16))
        tk.Label(r1, text="Payment:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_pay = tk.StringVar(value="Credit")
        ttk.Combobox(r1, textvariable=v_pay, width=10, state="readonly",
                     values=["Cash","Credit","Cheque","Online/UPI"]).pack(side="left", padx=4)

        r2 = tk.Frame(top, bg=C_BG); r2.pack(fill="x", pady=(6,0))
        tk.Label(r2, text="Supplier:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn = get_db()
        suppliers = [r["name"] for r in conn.execute("SELECT name FROM parties WHERE ptype IN ('Supplier','Both') ORDER BY name").fetchall()]
        conn.close()
        v_party = tk.StringVar()
        ttk.Combobox(r2, textvariable=v_party, width=30, values=suppliers).pack(side="left", padx=(4,16))

        # Items
        items_frame = tk.LabelFrame(main, text="Items", font=("Segoe UI",9,"bold"),
                                    bg=C_BG, fg=C_DARK, padx=6, pady=6)
        items_frame.pack(fill="both", expand=True, pady=(0,6))

        entry_row = tk.Frame(items_frame, bg=C_BG); entry_row.pack(fill="x", pady=(0,6))
        tk.Label(entry_row, text="Product:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn2 = get_db()
        all_prods = [r["name"] for r in conn2.execute("SELECT name FROM products ORDER BY name").fetchall()]
        conn2.close()
        v_prod = tk.StringVar()
        prod_cb = ttk.Combobox(entry_row, textvariable=v_prod, width=30, values=all_prods)
        prod_cb.pack(side="left", padx=(4,8))
        tk.Label(entry_row, text="Qty:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_qty = tk.StringVar(value="1")
        p_qty_e = ttk.Entry(entry_row, textvariable=v_qty, width=7); p_qty_e.pack(side="left", padx=(4,4))
        v_unit = tk.StringVar(value="Ltr")
        ttk.Combobox(entry_row, textvariable=v_unit, width=5, state="readonly",
                     values=["Ltr","ML","Kg","Pcs"]).pack(side="left", padx=(0,8))
        tk.Label(entry_row, text="Rate:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_rate = tk.StringVar(value="0")
        p_rate_e = ttk.Entry(entry_row, textvariable=v_rate, width=9); p_rate_e.pack(side="left", padx=(4,8))
        tk.Label(entry_row, text="GST%:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_gst = tk.StringVar(value="18")
        p_gst_cb = ttk.Combobox(entry_row, textvariable=v_gst, width=5, state="readonly",
                     values=["0","5","12","18","28"]); p_gst_cb.pack(side="left", padx=4)

        # Add + Remove buttons same row mein
        _pur_add_btn = make_btn(entry_row, "➕ Add", None, bg=C_ORANGE, pady=3); _pur_add_btn.pack(side="left", padx=(8,4))
        _pur_rem_btn = make_btn(entry_row, "🗑️ Remove", None, bg=C_RED, pady=3); _pur_rem_btn.pack(side="left", padx=4)

        def prod_selected(ev=None):
            nm = v_prod.get().strip()
            conn3 = get_db()
            row = conn3.execute("SELECT * FROM products WHERE name=?", (nm,)).fetchone()
            conn3.close()
            if row:
                v_rate.set(str(row["purchase_rate"]))
                v_unit.set(row["unit"] or "Ltr")
                v_gst.set(str(int(row["gst_percent"])))
        prod_cb.bind("<<ComboboxSelected>>", prod_selected)

        # ── Purchase Header Enter Navigation ──
        # Bill No → Date → Payment CB → Supplier CB → Product (item section)
        _hdr_nav = []
        for w in r1.winfo_children():
            if isinstance(w, (ttk.Entry, ttk.Combobox)):
                _hdr_nav.append(w)
            elif isinstance(w, tk.Frame):   # DatePicker frame
                for c in w.winfo_children():
                    if isinstance(c, (ttk.Entry, ttk.Combobox)):
                        _hdr_nav.append(c)
        for w in r2.winfo_children():
            if isinstance(w, (ttk.Entry, ttk.Combobox)):
                _hdr_nav.append(w)
        _hdr_nav.append(prod_cb)
        bind_enter_nav(_hdr_nav)

        # Row 2 of item entry — Batch No + Expiry Date (optional, batch-wise tracking)
        entry_row2 = tk.Frame(items_frame, bg=C_BG); entry_row2.pack(fill="x", pady=(0,6))
        tk.Label(entry_row2, text="Batch No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_batch = tk.StringVar()
        p_batch_e = ttk.Entry(entry_row2, textvariable=v_batch, width=14); p_batch_e.pack(side="left", padx=(4,16))
        tk.Label(entry_row2, text="Expiry (MM/YY):", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_expiry = tk.StringVar()
        p_expiry_e = ttk.Entry(entry_row2, textvariable=v_expiry, width=8)
        p_expiry_e.pack(side="left", padx=(4,4))
        tk.Label(entry_row2, text="(jaise 08/27 — blank chhodein agar expiry na ho)", font=("Segoe UI",8),
                 bg=C_BG, fg=C_GRAY).pack(side="left", padx=(6,0))

        # Purchase Enter nav: Product → Qty → Rate → GST → Batch → Expiry → Add Item
        # Enter navigation - Purchase items
        pur_nav_order = [prod_cb, p_qty_e, p_rate_e, p_gst_cb, p_batch_e, p_expiry_e]
        def _pur_enter(idx):
            def _fn(ev):
                if idx == len(pur_nav_order) - 1:
                    add_item()
                    self.root.after(10, lambda: prod_cb.focus_force())
                else:
                    nxt = pur_nav_order[idx + 1]
                    nxt.focus_force()
                    try: nxt.select_range(0, "end")
                    except: pass
                return "break"
            return _fn
        for _i, _w in enumerate(pur_nav_order):
            _w.bind("<Return>", _pur_enter(_i))
            _w.bind("<KP_Enter>", _pur_enter(_i))

        cols = ("Product","Qty","Unit","Rate","Taxable","GST%","GST Amt","Total","Batch","Expiry (MM/YY)")
        tv = ttk.Treeview(items_frame, columns=cols, show="headings", height=8)
        for c, w in zip(cols, [190,55,50,70,80,50,70,80,90,90]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb6 = ttk.Scrollbar(items_frame, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb6.set)
        sb6.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)

        self._pur_items = []

        def add_item():
            try:
                qty = float(v_qty.get() or 0); rate = float(v_rate.get() or 0)
                gst_pct = float(v_gst.get() or 0)
                if not v_prod.get().strip(): messagebox.showwarning("","Product chunein!"); return
                # Expiry date: blank allowed, warna MM/YY -> ISO (us mahine ka aakhri din) me convert karo
                expiry_iso = ""
                exp_raw = v_expiry.get().strip()
                if exp_raw:
                    try:
                        expiry_iso = mmyy_to_iso(exp_raw)
                    except ValueError:
                        messagebox.showwarning("","Expiry sahi format me daalein: MM/YY (jaise 08/27)"); return
                taxable = round(qty * rate, 2)
                gst_amt = round(taxable * gst_pct / 100, 2)
                total   = round(taxable + gst_amt, 2)
                item = {"product":v_prod.get().strip(),"qty":qty,"unit":v_unit.get(),
                        "rate":rate,"gst_pct":gst_pct,"taxable":taxable,"gst_amt":gst_amt,"total":total,
                        "batch_no":v_batch.get().strip(),"expiry_date":expiry_iso}
                self._pur_items.append(item)
                tv.insert("","end", values=(item["product"], f"{qty:.2f}", item["unit"],
                          f"{rate:.2f}", f"{taxable:.2f}", f"{gst_pct:.0f}%",
                          f"{gst_amt:.2f}", f"{total:.2f}", item["batch_no"],
                          iso_to_mmyy(item["expiry_date"]) or "—"))
                grand_lbl.config(text=f"Grand Total: {rupee(sum(i['total'] for i in self._pur_items))}")
                v_batch.set(""); v_expiry.set("")
            except ValueError: messagebox.showwarning("","Sahi values daalen!")

        def remove_item():
            sel = tv.selection()
            if not sel: return
            idx = tv.index(sel[0]); tv.delete(sel[0]); self._pur_items.pop(idx)
            grand_lbl.config(text=f"Grand Total: {rupee(sum(i['total'] for i in self._pur_items))}")

        # Wire up inline buttons
        _pur_add_btn.config(command=add_item)
        _pur_rem_btn.config(command=remove_item)

        grand_lbl = tk.Label(items_frame, text="Grand Total: ₹0.00", font=("Segoe UI",12,"bold"),
                             bg=C_BG, fg=C_ORANGE); grand_lbl.pack(anchor="e", padx=12, pady=(2,0))

        af = tk.Frame(main, bg=C_BG); af.pack(fill="x", pady=6)

        def save_purchase():
            if not self._pur_items: messagebox.showwarning("","Items add karein!"); return
            if not v_bill.get().strip(): messagebox.showwarning("","Bill No daalen!"); return
            grand = sum(i["total"] for i in self._pur_items)
            conn3 = get_db()
            try:
                conn3.execute("""INSERT INTO purchases
                    (bill_no,bill_date,party,grand_total,pay_mode)
                    VALUES(?,?,?,?,?)""",
                    (v_bill.get(), v_date.get(), v_party.get(), round(grand,2), v_pay.get()))
                pur_id = conn3.execute("SELECT id FROM purchases WHERE bill_no=?", (v_bill.get(),)).fetchone()["id"]
                for it in self._pur_items:
                    conn3.execute("""INSERT INTO purchase_items
                        (purchase_id,product,qty,unit,rate,taxable,gst_percent,gst_amt,total,batch_no,expiry_date)
                        VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                        (pur_id,it["product"],it["qty"],it["unit"],it["rate"],
                         it["taxable"],it["gst_pct"],it["gst_amt"],it["total"],
                         it["batch_no"],it["expiry_date"]))
                conn3.commit(); conn3.close()
                messagebox.showinfo("Saved!",f"Purchase {v_bill.get()} save ho gaya!")
                self._show_purchase()
            except Exception as ex:
                conn3.close(); messagebox.showerror("Error",str(ex))

        make_btn(af, "💾 Save Purchase", save_purchase, bg=C_GREEN, pady=10).pack(side="left", padx=4)
        make_btn(af, "🔄 Reset", self._show_purchase, bg=C_GRAY, pady=10).pack(side="left", padx=4)


    # ── SALE HISTORY ──────────────────────────────────────────────────────────
    def _show_sale_history(self):
        self._clear()
        self._section_header("📋 Sale History", "Sabhi bills dekhein")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # Search
        sf = tk.Frame(main, bg=C_BG); sf.pack(fill="x", pady=(0,8))
        tk.Label(sf, text="🔍 Search:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_search = tk.StringVar()
        se = ttk.Entry(sf, textvariable=v_search, width=28); se.pack(side="left", padx=4)
        tk.Label(sf, text="From:", font=("Segoe UI",9), bg=C_BG).pack(side="left", padx=(12,0))
        v_from = tk.StringVar(); fde_sh = make_date_entry(sf, v_from, width=12, bg=C_BG); fde_sh.pack(side="left", padx=4)
        tk.Label(sf, text="To:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_to = tk.StringVar(); tde_sh = make_date_entry(sf, v_to, width=12, bg=C_BG); tde_sh.pack(side="left", padx=4)
        search_btn = make_btn(sf, "🔍 Search", lambda: load_bills(), bg=C_ORANGE); search_btn.pack(side="left", padx=8)
        make_btn(sf, "📋 All Bills", lambda: (v_search.set(""), load_bills()), bg=C_GRAY).pack(side="left")
        total_lbl = tk.Label(sf, text="", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_ORANGE)
        total_lbl.pack(side="right", padx=12)

        # Enter nav: Search → From → To → Search button
        bind_enter_nav([se, fde_sh.entry, tde_sh.entry, search_btn])

        cols = ("Bill No","Date","Customer","Bill Type","Grand Total","Payment","Action")
        tv = ttk.Treeview(main, columns=cols, show="headings")
        for c, w in zip(cols, [140,90,180,80,100,90,80]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb7 = ttk.Scrollbar(main, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb7.set)
        sb7.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)

        self._sale_rows = []

        def load_bills():
            tv.delete(*tv.get_children())
            q = "SELECT * FROM sales WHERE 1=1"
            params = []
            search = v_search.get().strip()
            if search:
                q += " AND (bill_no LIKE ? OR party LIKE ?)"; params += [f"%{search}%", f"%{search}%"]
            try:
                fd_str = v_from.var.get() if hasattr(v_from, 'var') else v_from.get()
                td_str = v_to.var.get() if hasattr(v_to, 'var') else v_to.get()
                fd = datetime.datetime.strptime(fd_str, "%d-%m-%Y").strftime("%Y-%m-%d")
                td = datetime.datetime.strptime(td_str, "%d-%m-%Y").strftime("%Y-%m-%d")
                q += " AND bill_date BETWEEN ? AND ?"; params += [fd, td]
            except: pass
            q += " ORDER BY id DESC"
            conn = get_db(); rows = conn.execute(q, params).fetchall(); conn.close()
            self._sale_rows = list(rows)
            total = sum(r["grand_total"] for r in rows)
            total_lbl.config(text=f"Total: {rupee(total)} ({len(rows)} bills)")
            for r in rows:
                tv.insert("","end", values=(r["bill_no"], r["bill_date"], r["party"],
                          r["gst_type"], rupee(r["grand_total"]), r["pay_mode"], "View"))

        load_bills()

        def on_double_click(ev):
            sel = tv.selection()
            if not sel: return
            idx = tv.index(sel[0])
            if idx < len(self._sale_rows):
                self._print_bill(self._sale_rows[idx]["bill_no"])

        tv.bind("<Double-1>", on_double_click)

        # Action buttons
        btn_row = tk.Frame(main, bg=C_BG); btn_row.pack(fill="x", pady=(4,0))

        def print_selected():
            sel = tv.selection()
            if not sel: messagebox.showwarning("","Pehle bill select karein!"); return
            idx = tv.index(sel[0])
            if idx < len(self._sale_rows):
                self._print_bill(self._sale_rows[idx]["bill_no"])

        def delete_selected():
            sel = tv.selection()
            if not sel: messagebox.showwarning("","Pehle bill select karein!"); return
            idx = tv.index(sel[0])
            if idx >= len(self._sale_rows): return
            bill = self._sale_rows[idx]["bill_no"]
            if not messagebox.askyesno("Delete", f"Bill {bill} delete karein? Stock bhi adjust ho jayega."): return
            conn = get_db()
            conn.execute("DELETE FROM sales WHERE bill_no=?", (bill,))
            conn.execute("DELETE FROM bill_payments WHERE bill_no=?", (bill,))
            conn.commit(); conn.close()
            messagebox.showinfo("Done!", f"Bill {bill} delete ho gaya!")
            load_bills()

        make_btn(btn_row, "🖨️ Print Bill", print_selected, bg=C_GREEN, pady=4).pack(side="left", padx=4)
        make_btn(btn_row, "🗑️ Delete Bill", delete_selected, bg=C_RED, pady=4).pack(side="left", padx=4)
        tk.Label(main, text="💡 Bill pe double-click ya Print button se print karein",
                 font=("Segoe UI",8), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=2)

    # ── PURCHASE HISTORY ──────────────────────────────────────────────────────
    def _show_purchase_history(self):
        self._clear()
        self._section_header("📦 Purchase History", "Sabhi purchases dekhein")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        sf = tk.Frame(main, bg=C_BG); sf.pack(fill="x", pady=(0,8))
        tk.Label(sf, text="🔍 Search:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_search = tk.StringVar()
        se_ph = ttk.Entry(sf, textvariable=v_search, width=24); se_ph.pack(side="left", padx=4)
        search_btn_ph = make_btn(sf, "🔍 Search", lambda: load(), bg=C_ORANGE); search_btn_ph.pack(side="left", padx=8)
        make_btn(sf, "All", lambda: (v_search.set(""), load()), bg=C_GRAY).pack(side="left")
        total_lbl = tk.Label(sf, text="", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_ORANGE)
        total_lbl.pack(side="right", padx=12)

        # Enter nav: Search → Search button
        bind_enter_nav([se_ph, search_btn_ph])

        cols = ("Bill No","Date","Supplier","Grand Total","Payment")
        tv = ttk.Treeview(main, columns=cols, show="headings")
        for c, w in zip(cols, [150,90,200,110,100]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb8 = ttk.Scrollbar(main, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb8.set)
        sb8.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)

        def load():
            tv.delete(*tv.get_children())
            s = v_search.get().strip()
            q = "SELECT * FROM purchases WHERE 1=1"
            params = []
            if s: q += " AND (bill_no LIKE ? OR party LIKE ?)"; params += [f"%{s}%", f"%{s}%"]
            q += " ORDER BY id DESC"
            conn = get_db(); rows = conn.execute(q, params).fetchall(); conn.close()
            total = sum(r["grand_total"] for r in rows)
            total_lbl.config(text=f"Total: {rupee(total)} ({len(rows)} entries)")
            for r in rows:
                tv.insert("","end", values=(r["bill_no"], r["bill_date"], r["party"],
                          rupee(r["grand_total"]), r["pay_mode"]))
        self._pur_rows = []

        def load_with_store():
            load()
            conn2 = get_db()
            s = v_search.get().strip()
            q = "SELECT * FROM purchases WHERE 1=1"
            params = []
            if s: q += " AND (bill_no LIKE ? OR party LIKE ?)"; params += [f"%{s}%", f"%{s}%"]
            q += " ORDER BY id DESC"
            self._pur_rows = conn2.execute(q, params).fetchall()
            conn2.close()

        def delete_purchase():
            sel = tv.selection()
            if not sel: messagebox.showwarning("","Pehle entry select karein!"); return
            load_with_store()
            idx = tv.index(sel[0])
            if idx >= len(self._pur_rows): return
            bill = self._pur_rows[idx]["bill_no"]
            if not messagebox.askyesno("Delete", f"Purchase {bill} delete karein? Stock bhi adjust ho jayega."): return
            conn2 = get_db()
            conn2.execute("DELETE FROM purchases WHERE bill_no=?", (bill,))
            conn2.execute("DELETE FROM bill_payments WHERE bill_no=? AND bill_type='purchase'", (bill,))
            conn2.commit(); conn2.close()
            messagebox.showinfo("Done!", f"Purchase {bill} delete ho gaya!")
            load()

        btn_ph = tk.Frame(main, bg=C_BG); btn_ph.pack(fill="x", pady=(4,0))
        make_btn(btn_ph, "🗑️ Delete Entry", delete_purchase, bg=C_RED, pady=4).pack(side="left", padx=4)
        load()

    # ── COLOR FORMULAS ────────────────────────────────────────────────────────
    def _show_formulas(self):
        self._clear()
        self._section_header("🚗 Color Formula Library", "Gaadi ka color match karne ke OEM shade formulas")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # Toolbar
        tb = tk.Frame(main, bg=C_BG); tb.pack(fill="x", pady=(0,8))
        make_btn(tb, "➕ Naya Formula", self._add_formula, bg=C_ORANGE).pack(side="left", padx=4)
        make_btn(tb, "✏️ Edit", self._edit_formula, bg=C_BLUE).pack(side="left", padx=4)
        make_btn(tb, "🗑️ Delete", self._delete_formula, bg=C_RED).pack(side="left", padx=4)
        v_search = tk.StringVar()
        ttk.Entry(tb, textvariable=v_search, width=22).pack(side="right", padx=4)
        tk.Label(tb, text="🔍", font=("Segoe UI",11), bg=C_BG).pack(side="right")

        # Split pane
        pane = tk.Frame(main, bg=C_BG); pane.pack(fill="both", expand=True)
        left = tk.Frame(pane, bg=C_BG, width=500); left.pack(side="left", fill="both")
        right = tk.LabelFrame(pane, text="Formula Details", font=("Segoe UI",9,"bold"),
                              bg=C_BG, fg=C_DARK, padx=8, pady=8)
        right.pack(side="left", fill="both", expand=True, padx=(8,0))

        cols = ("Shade Name","Code","Brand","Base Paint","Sale Price")
        self.formula_tv = ttk.Treeview(left, columns=cols, show="headings", height=20)
        for c, w in zip(cols, [150,70,90,160,90]):
            self.formula_tv.heading(c, text=c); self.formula_tv.column(c, width=w)
        sb9 = ttk.Scrollbar(left, orient="vertical", command=self.formula_tv.yview)
        self.formula_tv.configure(yscrollcommand=sb9.set)
        sb9.pack(side="right", fill="y"); self.formula_tv.pack(fill="both", expand=True)

        self._detail_text = tk.Text(right, font=("Segoe UI",10), bg=C_YELLOW, wrap="word",
                                    height=20, width=36, relief="flat")
        self._detail_text.pack(fill="both", expand=True)

        make_btn(right, "🧾 Is Formula Se Bill Banao",
                 self._bill_from_formula, bg=C_ORANGE, pady=8).pack(fill="x", pady=(8,0))

        self._formula_rows = []

        def load_formulas(*a):
            self.formula_tv.delete(*self.formula_tv.get_children())
            s = v_search.get().strip()
            q = "SELECT * FROM color_formulas WHERE 1=1"
            params = []
            if s: q += " AND (shade_name LIKE ? OR brand LIKE ? OR shade_code LIKE ?)"; params += [f"%{s}%"]*3
            q += " ORDER BY shade_name"
            conn = get_db(); rows = conn.execute(q, params).fetchall(); conn.close()
            self._formula_rows = list(rows)
            for f in rows:
                self.formula_tv.insert("","end", values=(f["shade_name"], f["shade_code"],
                                        f["brand"], f["base_product"], rupee(f["sale_price"])))

        def on_select(ev):
            sel = self.formula_tv.selection()
            if not sel: return
            idx = self.formula_tv.index(sel[0])
            if idx >= len(self._formula_rows): return
            f = self._formula_rows[idx]
            components = json.loads(f["components"] or "[]")
            self._detail_text.config(state="normal")
            self._detail_text.delete("1.0","end")
            self._detail_text.insert("end", f"🚗 {f['shade_name']}\n", "heading")
            self._detail_text.insert("end", f"Code: {f['shade_code']}  |  Brand: {f['brand']}\n\n")
            self._detail_text.insert("end", "BASE PAINT:\n", "bold")
            self._detail_text.insert("end", f"  {f['base_product']}: {f['base_qty']} {f['base_unit']}\n\n")
            self._detail_text.insert("end", "TINTERS:\n", "bold")
            for comp in components:
                self._detail_text.insert("end", f"  🔴 {comp['product']}: {comp['qty']} {comp['unit']}\n")
            self._detail_text.insert("end", f"\nCost: {rupee(f['total_cost'])}")
            self._detail_text.insert("end", f"\nSale Price: {rupee(f['sale_price'])}", "price")
            if f["notes"]: self._detail_text.insert("end", f"\n\nNotes: {f['notes']}")
            self._detail_text.config(state="disabled")
            self._detail_text.tag_config("heading", font=("Segoe UI",13,"bold"), foreground=C_ORANGE)
            self._detail_text.tag_config("bold", font=("Segoe UI",10,"bold"), foreground=C_DARK)
            self._detail_text.tag_config("price", font=("Segoe UI",12,"bold"), foreground=C_GREEN)

        self.formula_tv.bind("<<TreeviewSelect>>", on_select)
        v_search.trace_add("write", load_formulas)
        load_formulas()

    def _get_selected_formula(self):
        sel = self.formula_tv.selection() if hasattr(self,"formula_tv") else None
        if not sel: messagebox.showwarning("","Pehle ek formula chunein!"); return None
        idx = self.formula_tv.index(sel[0])
        return self._formula_rows[idx] if idx < len(self._formula_rows) else None

    def _add_formula(self):
        self._formula_dialog(None)

    def _edit_formula(self):
        f = self._get_selected_formula()
        if f: self._formula_dialog(f)

    def _delete_formula(self):
        f = self._get_selected_formula()
        if not f: return
        if messagebox.askyesno("Delete",f"'{f['shade_name']}' delete karna hai?"):
            conn = get_db()
            conn.execute("DELETE FROM color_formulas WHERE id=?", (f["id"],))
            conn.commit(); conn.close()
            self._show_formulas()

    def _bill_from_formula(self):
        f = self._get_selected_formula()
        if not f: return
        components = json.loads(f["components"] or "[]")
        items = [{
            "product": f["base_product"], "shade": f["shade_name"],
            "qty": f["base_qty"], "unit": f["base_unit"],
            "rate": f["sale_price"], "gst_pct": 18,
            "taxable": round(f["base_qty"] * f["sale_price"], 2),
            "gst_amt": round(f["base_qty"] * f["sale_price"] * 0.18, 2),
            "total": round(f["base_qty"] * f["sale_price"] * 1.18, 2),
        }]
        self._show_sale_bill(prefill_items=items)

    def _formula_dialog(self, formula=None):
        dlg = tk.Toplevel(self.root)
        dlg.title("Formula Add/Edit")
        dlg.configure(bg=C_BG)
        dlg.geometry("720x650")
        dlg.grab_set()

        is_edit = formula is not None
        tk.Label(dlg, text="🚗 Color Formula", font=("Segoe UI",13,"bold"),
                 bg=C_ORANGE, fg="white").pack(fill="x", ipady=10)

        form = tk.Frame(dlg, bg=C_BG, padx=16, pady=10); form.pack(fill="x")

        def row(parent, label, var, width=20):
            f = tk.Frame(parent, bg=C_BG); f.pack(fill="x", pady=3)
            tk.Label(f, text=label, font=("Segoe UI",9), bg=C_BG, width=16, anchor="w").pack(side="left")
            e = ttk.Entry(f, textvariable=var, width=width); e.pack(side="left")
            return e

        v_name   = tk.StringVar(value=formula["shade_name"] if is_edit else "")
        v_code   = tk.StringVar(value=formula["shade_code"] if is_edit else "")
        v_brand  = tk.StringVar(value=formula["brand"] if is_edit else "")
        v_notes  = tk.StringVar(value=formula["notes"] if is_edit else "")
        v_price  = tk.StringVar(value=str(formula["sale_price"]) if is_edit else "0")

        name_e  = row(form, "Shade Name *", v_name, 28)
        code_e  = row(form, "Shade Code", v_code, 14)
        brand_e = row(form, "Brand", v_brand, 20)
        price_e = row(form, "Sale Price ₹", v_price, 12)
        notes_e = row(form, "Notes", v_notes, 34)

        # Base paint
        base_frame = tk.LabelFrame(dlg, text="Base Paint", font=("Segoe UI",9,"bold"),
                                   bg=C_BG, fg=C_DARK, padx=8, pady=6)
        base_frame.pack(fill="x", padx=16, pady=(0,6))
        bf = tk.Frame(base_frame, bg=C_BG); bf.pack(fill="x")

        conn = get_db()
        base_products = [r["name"] for r in conn.execute("SELECT name FROM products WHERE is_base=1 ORDER BY name").fetchall()]
        if not base_products:
            base_products = [r["name"] for r in conn.execute("SELECT name FROM products WHERE category='Paint' ORDER BY name").fetchall()]
        conn.close()

        tk.Label(bf, text="Base Product:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_base = tk.StringVar(value=formula["base_product"] if is_edit else "")
        base_cb = ttk.Combobox(bf, textvariable=v_base, width=30, values=base_products)
        base_cb.pack(side="left", padx=4)
        tk.Label(bf, text="Qty:", font=("Segoe UI",9), bg=C_BG).pack(side="left", padx=(8,0))
        v_base_qty = tk.StringVar(value=str(formula["base_qty"]) if is_edit else "1")
        base_qty_e = ttk.Entry(bf, textvariable=v_base_qty, width=6); base_qty_e.pack(side="left", padx=4)
        v_base_unit = tk.StringVar(value=formula["base_unit"] if is_edit else "Ltr")
        base_unit_cb = ttk.Combobox(bf, textvariable=v_base_unit, width=5, state="readonly",
                     values=["Ltr","ML","Kg"])
        base_unit_cb.pack(side="left", padx=4)

        # ── Formula Header Enter Navigation ──
        # Name → Code → Brand → Price → Notes → Base Product → Qty → Unit → (Tinter, bound below)
        bind_enter_nav([name_e, code_e, brand_e, price_e, notes_e, base_cb])

        # Tinters / Components
        comp_frame = tk.LabelFrame(dlg, text="Tinters / Components", font=("Segoe UI",9,"bold"),
                                   bg=C_BG, fg=C_DARK, padx=8, pady=6)
        comp_frame.pack(fill="both", expand=True, padx=16, pady=(0,6))

        # Component entry
        ce_row = tk.Frame(comp_frame, bg=C_BG); ce_row.pack(fill="x", pady=(0,6))
        tk.Label(ce_row, text="Tinter:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn2 = get_db()
        tinters = [r["name"] for r in conn2.execute("SELECT name FROM products WHERE is_tinter=1 OR category='Tinter' ORDER BY name").fetchall()]
        conn2.close()
        v_tinter = tk.StringVar()
        tinter_cb = ttk.Combobox(ce_row, textvariable=v_tinter, width=28, values=tinters)
        tinter_cb.pack(side="left", padx=4)
        tk.Label(ce_row, text="Qty:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_tqty = tk.StringVar(value="10")
        tqty_e = ttk.Entry(ce_row, textvariable=v_tqty, width=6); tqty_e.pack(side="left", padx=4)
        v_tunit = tk.StringVar(value="ML")
        tunit_cb = ttk.Combobox(ce_row, textvariable=v_tunit, width=5, state="readonly",
                     values=["ML","Ltr","Kg"])
        tunit_cb.pack(side="left", padx=4)

        # Base row → Tinter row Enter nav
        bind_enter_nav([base_cb, base_qty_e, base_unit_cb, tinter_cb])

        cols = ("Tinter Product","Qty","Unit")
        comp_tv = ttk.Treeview(comp_frame, columns=cols, show="headings", height=6)
        for c, w in zip(cols, [280,80,60]):
            comp_tv.heading(c, text=c); comp_tv.column(c, width=w, anchor="center")
        comp_tv.pack(fill="both", expand=True)

        self._comp_items = []
        if is_edit:
            try:
                self._comp_items = json.loads(formula["components"] or "[]")
                for comp in self._comp_items:
                    comp_tv.insert("","end", values=(comp["product"], comp["qty"], comp["unit"]))
            except: pass

        def add_comp():
            if not v_tinter.get().strip(): messagebox.showwarning("","Tinter chunein!"); return
            try: qty = float(v_tqty.get())
            except: messagebox.showwarning("","Qty daalen!"); return
            comp = {"product": v_tinter.get().strip(), "qty": qty, "unit": v_tunit.get()}
            self._comp_items.append(comp)
            comp_tv.insert("","end", values=(comp["product"], comp["qty"], comp["unit"]))
            v_tinter.set(""); v_tqty.set("10")

        def remove_comp():
            sel = comp_tv.selection()
            if not sel: return
            idx = comp_tv.index(sel[0]); comp_tv.delete(sel[0]); self._comp_items.pop(idx)

        # Tinter row Enter nav: Tinter → Qty → Unit → Add Tinter (trigger, loops back)
        comp_nav_order = [tinter_cb, tqty_e, tunit_cb]
        def _comp_enter(idx):
            def _fn(ev):
                if idx == len(comp_nav_order) - 1:
                    add_comp()
                    dlg.after(10, lambda: tinter_cb.focus_force())
                else:
                    nxt = comp_nav_order[idx + 1]
                    nxt.focus_force()
                    try: nxt.select_range(0, "end")
                    except: pass
                return "break"
            return _fn
        for _i, _w in enumerate(comp_nav_order):
            _w.bind("<Return>", _comp_enter(_i))
            _w.bind("<KP_Enter>", _comp_enter(_i))

        cbrow = tk.Frame(comp_frame, bg=C_BG); cbrow.pack(fill="x", pady=4)
        make_btn(cbrow, "➕ Add Tinter", add_comp, bg=C_TEAL).pack(side="left", padx=4)
        make_btn(cbrow, "🗑️ Remove", remove_comp, bg=C_RED).pack(side="left", padx=4)

        def save_formula():
            if not v_name.get().strip(): messagebox.showwarning("","Shade name zaroori hai!"); return
            try:
                price = float(v_price.get() or 0)
                base_qty = float(v_base_qty.get() or 1)
            except: messagebox.showwarning("","Price ya Qty sahi daalen!"); return

            # Auto-calculate cost
            conn3 = get_db()
            cost = 0
            base_row = conn3.execute("SELECT purchase_rate,unit FROM products WHERE name=?", (v_base.get(),)).fetchone()
            if base_row:
                cost += base_row["purchase_rate"] * base_qty
            for comp in self._comp_items:
                cr = conn3.execute("SELECT purchase_rate,unit FROM products WHERE name=?", (comp["product"],)).fetchone()
                if cr:
                    qty = comp["qty"]
                    if comp["unit"] == "ML": qty = qty / 1000
                    cost += cr["purchase_rate"] * qty
            conn3.close()

            data = {
                "shade_name": v_name.get().strip(),
                "shade_code": v_code.get().strip(),
                "brand":      v_brand.get().strip(),
                "base_product": v_base.get().strip(),
                "base_qty":   base_qty,
                "base_unit":  v_base_unit.get(),
                "components": json.dumps(self._comp_items),
                "total_cost": round(cost, 2),
                "sale_price": price,
                "notes":      v_notes.get().strip(),
            }
            conn4 = get_db()
            if is_edit:
                conn4.execute("""UPDATE color_formulas SET
                    shade_name=?,shade_code=?,brand=?,base_product=?,base_qty=?,base_unit=?,
                    components=?,total_cost=?,sale_price=?,notes=? WHERE id=?""",
                    (*data.values(), formula["id"]))
            else:
                conn4.execute("""INSERT INTO color_formulas
                    (shade_name,shade_code,brand,base_product,base_qty,base_unit,
                     components,total_cost,sale_price,notes)
                    VALUES(?,?,?,?,?,?,?,?,?,?)""", tuple(data.values()))
            conn4.commit(); conn4.close()
            messagebox.showinfo("Saved!",f"Formula '{v_name.get()}' save ho gaya! Cost: {rupee(cost)}")
            dlg.destroy()
            self._show_formulas()

        make_btn(dlg, "💾 Save Formula", save_formula, bg=C_GREEN, pady=10).pack(fill="x", padx=16, pady=8)

    # ── STOCK REGISTER ─────────────────────────────────────────────────────────
    def _show_stock(self):
        self._clear()
        self._section_header("📦 Stock Register", "Current stock dekhein")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        sf = tk.Frame(main, bg=C_BG); sf.pack(fill="x", pady=(0,8))
        make_btn(sf, "🔄 Refresh", self._show_stock, bg=C_ORANGE).pack(side="left", padx=4)
        v_cat = tk.StringVar(value="All")
        tk.Label(sf, text="Category:", font=("Segoe UI",9), bg=C_BG).pack(side="left", padx=(12,0))
        cat_cb = ttk.Combobox(sf, textvariable=v_cat, width=12, state="readonly",
                               values=["All","Paint","Tinter","Primer","Putty","Thinner","Brush","Roller"])
        cat_cb.pack(side="left", padx=4)
        low_only = tk.BooleanVar()
        ttk.Checkbutton(sf, text="⚠️ Low Stock Only", variable=low_only).pack(side="left", padx=8)

        total_val_lbl = tk.Label(sf, text="", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_ORANGE)
        total_val_lbl.pack(side="right", padx=12)

        cols = ("Product","Category","Brand","Unit","Opening","Purchase","Sold","Current Stock","Value","Status")
        tv = ttk.Treeview(main, columns=cols, show="headings")
        for c, w in zip(cols, [200,80,100,50,70,80,70,100,90,80]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        tv.tag_configure("low", background="#FFEBEE", foreground=C_RED)
        tv.tag_configure("ok",  background="#E8F5E9", foreground=C_GREEN)
        sb10 = ttk.Scrollbar(main, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb10.set)
        sb10.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)

        def load_stock(*a):
            tv.delete(*tv.get_children())
            q = "SELECT * FROM products WHERE 1=1"
            params = []
            cat = v_cat.get()
            if cat != "All": q += " AND category=?"; params.append(cat)
            q += " ORDER BY category, name"
            conn = get_db(); prods = conn.execute(q, params).fetchall()
            total_val = 0
            for p in prods:
                opening = p["opening_stock"]
                pur_qty = conn.execute("SELECT COALESCE(SUM(qty),0) as t FROM purchase_items WHERE product=?", (p["name"],)).fetchone()["t"]
                sale_qty= conn.execute("SELECT COALESCE(SUM(qty),0) as t FROM sale_items WHERE product=?", (p["name"],)).fetchone()["t"]
                curr = opening + pur_qty - sale_qty
                if low_only.get() and curr > p["low_stock_alert"]:
                    continue
                val = curr * p["sale_rate"]
                total_val += val
                status = "⚠️ LOW" if curr <= p["low_stock_alert"] else "✅ OK"
                tag = "low" if curr <= p["low_stock_alert"] else "ok"
                tv.insert("","end", tag=tag, values=(
                    p["name"], p["category"], p["brand"], p["unit"],
                    f"{opening:.2f}", f"{pur_qty:.2f}", f"{sale_qty:.2f}",
                    f"{curr:.2f}", rupee(val), status))
            conn.close()
            total_val_lbl.config(text=f"Total Stock Value: {rupee(total_val)}")

        v_cat.trace_add("write", load_stock)
        low_only.trace_add("write", load_stock)
        load_stock()

    # ── PRODUCTS ──────────────────────────────────────────────────────────────

    # ── EXCEL UPLOAD — PRODUCTS ───────────────────────────────────────────────
    def _upload_products_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error","openpyxl install karo:\npip install openpyxl"); return

        path = filedialog.askopenfilename(
            title="Products Excel chunein",
            filetypes=[("Excel Files","*.xlsx *.xls"),("All Files","*.*")])
        if not path: return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

            # Expected columns
            col_map = {
                "name": None, "category": None, "brand": None, "unit": None,
                "sale_rate": None, "purchase_rate": None, "gst_percent": None,
                "is_base": None, "is_tinter": None
            }
            aliases = {
                "sale rate": "sale_rate", "purchase rate": "purchase_rate",
                "gst": "gst_percent", "gst%": "gst_percent",
                "base": "is_base", "tinter": "is_tinter",
                "sale": "sale_rate", "purchase": "purchase_rate"
            }
            for i, h in enumerate(headers):
                if h in col_map:
                    col_map[h] = i
                elif h in aliases:
                    col_map[aliases[h]] = i

            if col_map["name"] is None:
                messagebox.showerror("Error","'name' column nahi mila Excel mein!"); return

            conn = get_db()
            added = updated = skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[col_map["name"]]: continue
                def g(key, default=""):
                    idx = col_map.get(key)
                    return row[idx] if idx is not None and row[idx] is not None else default
                name = str(g("name","")).strip()
                if not name: continue
                cat  = str(g("category","General")).strip()
                brand= str(g("brand","")).strip()
                unit = str(g("unit","Ltr")).strip()
                try: sale_rate = float(g("sale_rate", 0))
                except: sale_rate = 0.0
                try: pur_rate = float(g("purchase_rate", 0))
                except: pur_rate = 0.0
                try: gst = float(g("gst_percent", 18))
                except: gst = 18.0
                is_base   = 1 if str(g("is_base","")).strip().lower() in ("1","yes","true","✓") else 0
                is_tinter = 1 if str(g("is_tinter","")).strip().lower() in ("1","yes","true","✓") else 0

                existing = conn.execute("SELECT id FROM products WHERE name=?", (name,)).fetchone()
                if existing:
                    conn.execute("""UPDATE products SET category=?,brand=?,unit=?,
                                   sale_rate=?,purchase_rate=?,gst_percent=?,is_base=?,is_tinter=?
                                   WHERE name=?""",
                                 (cat,brand,unit,sale_rate,pur_rate,gst,is_base,is_tinter,name))
                    updated += 1
                else:
                    conn.execute("""INSERT INTO products
                                   (name,category,brand,unit,sale_rate,purchase_rate,gst_percent,is_base,is_tinter)
                                   VALUES(?,?,?,?,?,?,?,?,?)""",
                                 (name,cat,brand,unit,sale_rate,pur_rate,gst,is_base,is_tinter))
                    added += 1
            conn.commit(); conn.close()
            messagebox.showinfo("Done!",f"Products upload ho gaye!\n✅ Added: {added}\n✏️ Updated: {updated}")
            self._show_products()
        except Exception as e:
            messagebox.showerror("Error", f"Upload failed:\n{e}")

    def _download_products_template(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error","openpyxl install karo:\npip install openpyxl"); return
        path = filedialog.asksaveasfilename(
            title="Template save karein",
            defaultextension=".xlsx",
            initialfile="products_template.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Products"
        headers = ["name","category","brand","unit","sale_rate","purchase_rate","gst_percent","is_base","is_tinter"]
        ws.append(headers)
        ws.append(["Asian Paint WB","Paint","Asian","Ltr",450,320,18,"",""])
        ws.append(["Berger Primer","Primer","Berger","Ltr",280,200,18,"1",""])
        wb.save(path)
        messagebox.showinfo("Done!",f"Template saved:\n{path}")

    # ── EXCEL UPLOAD — PARTIES ────────────────────────────────────────────────
    def _upload_parties_excel(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error","openpyxl install karo:\npip install openpyxl"); return

        path = filedialog.askopenfilename(
            title="Parties Excel chunein",
            filetypes=[("Excel Files","*.xlsx *.xls"),("All Files","*.*")])
        if not path: return

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_map = {"name":None,"ptype":None,"mobile":None,"gstin":None,"address":None,"state":None,"opening_balance":None}
            aliases = {"type":"ptype","party type":"ptype","phone":"mobile","gst":"gstin",
                       "opening balance":"opening_balance","balance":"opening_balance"}
            for i, h in enumerate(headers):
                if h in col_map: col_map[h] = i
                elif h in aliases: col_map[aliases[h]] = i

            if col_map["name"] is None:
                messagebox.showerror("Error","'name' column nahi mila!"); return

            conn = get_db()
            added = updated = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[col_map["name"]]: continue
                def g(key, default=""):
                    idx = col_map.get(key)
                    return row[idx] if idx is not None and row[idx] is not None else default
                name = str(g("name","")).strip()
                if not name: continue
                ptype = str(g("ptype","Customer")).strip()
                mobile= str(g("mobile","")).strip()
                gstin = str(g("gstin","")).strip()
                addr  = str(g("address","")).strip()
                state = str(g("state","Uttar Pradesh")).strip()
                try: ob = float(g("opening_balance",0))
                except: ob = 0.0

                existing = conn.execute("SELECT id FROM parties WHERE name=?", (name,)).fetchone()
                if existing:
                    conn.execute("""UPDATE parties SET ptype=?,mobile=?,gstin=?,address=?,state=?
                                   WHERE name=?""", (ptype,mobile,gstin,addr,state,name))
                    updated += 1
                else:
                    conn.execute("""INSERT INTO parties (name,ptype,mobile,gstin,address,state,opening_balance)
                                   VALUES(?,?,?,?,?,?,?)""",
                                 (name,ptype,mobile,gstin,addr,state,ob))
                    added += 1
            conn.commit(); conn.close()
            messagebox.showinfo("Done!",f"Parties upload ho gayin!\n✅ Added: {added}\n✏️ Updated: {updated}")
            self._show_parties()
        except Exception as e:
            messagebox.showerror("Error", f"Upload failed:\n{e}")

    def _download_parties_template(self):
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("Error","openpyxl install karo:\npip install openpyxl"); return
        path = filedialog.asksaveasfilename(
            title="Template save karein",
            defaultextension=".xlsx",
            initialfile="parties_template.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not path: return
        wb = openpyxl.Workbook()
        ws = wb.active; ws.title = "Parties"
        ws.append(["name","ptype","mobile","gstin","address","state","opening_balance"])
        ws.append(["Ashish Verma","Customer","9876543210","","12 Civil Lines","Uttar Pradesh",0])
        ws.append(["Paint Depot","Supplier","9988776655","09ABCDE1234F1Z5","MG Road","Uttar Pradesh",5000])
        wb.save(path)
        messagebox.showinfo("Done!",f"Template saved:\n{path}")

    def _show_products(self):
        self._clear()
        self._section_header("📁  Products", "Products manage karein")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        tb = tk.Frame(main, bg=C_BG); tb.pack(fill="x", pady=(0,8))
        make_btn(tb, "➕ Add Product", self._add_product, bg=C_ORANGE).pack(side="left", padx=4)
        make_btn(tb, "✏️ Edit", lambda: self._edit_product(tv, prod_rows), bg=C_BLUE).pack(side="left", padx=4)
        make_btn(tb, "📤 Excel Upload", self._upload_products_excel, bg=C_GREEN).pack(side="left", padx=4)
        make_btn(tb, "📥 Template", self._download_products_template, bg=C_TEAL).pack(side="left", padx=4)
        v_search = tk.StringVar()
        ttk.Entry(tb, textvariable=v_search, width=24).pack(side="right", padx=4)
        tk.Label(tb, text="🔍", font=("Segoe UI",11), bg=C_BG).pack(side="right")

        cols = ("Name","Category","Brand","Unit","Sale Rate","Purchase Rate","GST%","Stock","Base","Tinter")
        tv = ttk.Treeview(main, columns=cols, show="headings")
        for c, w in zip(cols, [200,80,100,50,90,95,50,70,50,50]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb11 = ttk.Scrollbar(main, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb11.set)
        sb11.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)
        prod_rows = []

        def load(*a):
            tv.delete(*tv.get_children()); prod_rows.clear()
            s = v_search.get().strip()
            q = "SELECT * FROM products WHERE 1=1"
            params = []
            if s: q += " AND (name LIKE ? OR category LIKE ? OR brand LIKE ?)"; params += [f"%{s}%"]*3
            q += " ORDER BY category, name"
            conn = get_db(); rows = conn.execute(q, params).fetchall(); conn.close()
            for r in rows:
                prod_rows.append(dict(r))
                stock = get_stock(r["name"])
                tv.insert("","end", values=(r["name"], r["category"], r["brand"], r["unit"],
                          rupee(r["sale_rate"]), rupee(r["purchase_rate"]), f"{r['gst_percent']:.0f}%",
                          f"{stock:.2f}", "✓" if r["is_base"] else "", "✓" if r["is_tinter"] else ""))

        v_search.trace_add("write", load); load()

    def _add_product(self):
        self._product_dialog(None)

    def _edit_product(self, tv, prod_rows):
        sel = tv.selection()
        if not sel: messagebox.showwarning("","Product chunein!"); return
        idx = tv.index(sel[0])
        if idx < len(prod_rows): self._product_dialog(prod_rows[idx])

    def _product_dialog(self, product=None):
        dlg = tk.Toplevel(self.root)
        dlg.title("Product Add/Edit")
        dlg.configure(bg=C_BG); dlg.geometry("560x580"); dlg.grab_set()
        is_edit = product is not None

        tk.Label(dlg, text="🛍️ Product Details", font=("Segoe UI",12,"bold"),
                 bg=C_ORANGE, fg="white").pack(fill="x", ipady=10)

        form = tk.Frame(dlg, bg=C_BG, padx=16, pady=12); form.pack(fill="x")

        fields = {}
        def frow(label, key, default="", width=24):
            f = tk.Frame(form, bg=C_BG); f.pack(fill="x", pady=4)
            tk.Label(f, text=label, font=("Segoe UI",9), bg=C_BG, width=16, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(product[key]) if is_edit and key in product else default)
            ttk.Entry(f, textvariable=var, width=width).pack(side="left")
            fields[key] = var

        def cbrow(label, key, values, default=""):
            f = tk.Frame(form, bg=C_BG); f.pack(fill="x", pady=4)
            tk.Label(f, text=label, font=("Segoe UI",9), bg=C_BG, width=16, anchor="w").pack(side="left")
            var = tk.StringVar(value=str(product[key]) if is_edit and key in product else default)
            ttk.Combobox(f, textvariable=var, width=18, state="readonly", values=values).pack(side="left")
            fields[key] = var

        frow("Name *", "name", width=34)
        cbrow("Category", "category", ["Paint","Tinter","Primer","Putty","Thinner","Brush","Roller","Other"], "Paint")
        frow("Brand", "brand", "Generic")
        frow("HSN Code", "hsn", "3208")
        cbrow("Unit", "unit", ["Ltr","ML","Kg","Pcs","Box"], "Ltr")
        frow("Sale Rate ₹", "sale_rate", "0")
        frow("Purchase Rate ₹", "purchase_rate", "0")
        frow("MRP ₹", "mrp", "0")
        cbrow("GST %", "gst_percent", ["0","5","12","18","28"], "18")
        frow("Opening Stock", "opening_stock", "0")
        frow("Low Stock Alert", "low_stock_alert", "5")

        ckf = tk.Frame(form, bg=C_BG); ckf.pack(fill="x", pady=6)
        v_base   = tk.BooleanVar(value=bool(product["is_base"]) if is_edit else False)
        v_tinter = tk.BooleanVar(value=bool(product["is_tinter"]) if is_edit else False)
        ttk.Checkbutton(ckf, text="Base Paint hai?", variable=v_base).pack(side="left", padx=8)
        ttk.Checkbutton(ckf, text="Tinter hai?", variable=v_tinter).pack(side="left", padx=8)

        def save():
            if not fields["name"].get().strip(): messagebox.showwarning("","Name zaroori!"); return
            try:
                data = {
                    "name": fields["name"].get().strip(),
                    "category": fields["category"].get(),
                    "brand": fields["brand"].get().strip(),
                    "hsn": fields["hsn"].get().strip(),
                    "unit": fields["unit"].get(),
                    "sale_rate": float(fields["sale_rate"].get() or 0),
                    "purchase_rate": float(fields["purchase_rate"].get() or 0),
                    "mrp": float(fields["mrp"].get() or 0),
                    "gst_percent": float(fields["gst_percent"].get() or 18),
                    "opening_stock": float(fields["opening_stock"].get() or 0),
                    "low_stock_alert": float(fields["low_stock_alert"].get() or 5),
                    "is_base": 1 if v_base.get() else 0,
                    "is_tinter": 1 if v_tinter.get() else 0,
                }
            except ValueError: messagebox.showwarning("","Numeric fields sahi daalen!"); return
            conn = get_db()
            if is_edit:
                conn.execute("""UPDATE products SET name=?,category=?,brand=?,hsn=?,unit=?,sale_rate=?,
                    purchase_rate=?,mrp=?,gst_percent=?,opening_stock=?,low_stock_alert=?,is_base=?,is_tinter=?
                    WHERE id=?""", (*data.values(), product["id"]))
            else:
                conn.execute("""INSERT INTO products
                    (name,category,brand,hsn,unit,sale_rate,purchase_rate,mrp,gst_percent,
                     opening_stock,low_stock_alert,is_base,is_tinter) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    tuple(data.values()))
            conn.commit(); conn.close()
            messagebox.showinfo("Saved!","Product save ho gaya!")
            dlg.destroy(); self._show_products()

        save_btn = make_btn(dlg, "💾 Save Product", save, bg=C_GREEN, pady=10)
        save_btn.pack(fill="x", padx=16, pady=8)

        # Enter nav: all Entry/Combobox/Checkbutton widgets in form → Save button
        all_entries = [w for f in form.winfo_children()
                       for w in f.winfo_children()
                       if isinstance(w, (ttk.Entry, ttk.Combobox, ttk.Checkbutton))]
        bind_enter_nav(all_entries + [save_btn])

    # ── PARTIES ───────────────────────────────────────────────────────────────
    def _show_parties(self):
        self._clear()
        self._section_header("👥 Parties", "Customers & Suppliers — Ledger & Reminders")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        nb = ttk.Notebook(main); nb.pack(fill="both", expand=True)

        # ═══════════════════════════════════════════════
        # TAB 1 — Parties List
        # ═══════════════════════════════════════════════
        tab_list = tk.Frame(nb, bg=C_BG); nb.add(tab_list, text="👥 Parties List")

        tb = tk.Frame(tab_list, bg=C_BG); tb.pack(fill="x", pady=(6,6))
        make_btn(tb, "➕ Add Party", self._add_party, bg=C_ORANGE).pack(side="left", padx=4)
        make_btn(tb, "📤 Excel Upload", self._upload_parties_excel, bg=C_GREEN).pack(side="left", padx=4)
        make_btn(tb, "📥 Template", self._download_parties_template, bg=C_TEAL).pack(side="left", padx=4)

        def send_party_reminder():
            sel = tv_party.selection()
            if not sel:
                messagebox.showwarning("", "Pehle ek party select karein!"); return
            vals = tv_party.item(sel[0])["values"]
            name   = vals[0]
            mobile = vals[2] or ""
            bal_str= vals[6]  # "Dr ₹500.00"
            conn2  = get_db()
            p_row  = conn2.execute("SELECT email FROM parties WHERE name=?", (name,)).fetchone()
            conn2.close()
            email = p_row["email"] if p_row else ""
            _show_notify_dialog(self.root, "Outstanding Payment", name, mobile, email, bal_str)

        make_btn(tb, "🔔 Reminder", send_party_reminder, bg="#25D366").pack(side="left", padx=4)
        v_psearch = tk.StringVar()
        ttk.Entry(tb, textvariable=v_psearch, width=22).pack(side="right", padx=4)
        tk.Label(tb, text="🔍", font=("Segoe UI",11), bg=C_BG).pack(side="right")

        cols = ("Name","Type","Mobile","GSTIN","Address","State","Balance")
        tv_party = ttk.Treeview(tab_list, columns=cols, show="headings")
        for c, w in zip(cols, [180,80,110,150,180,110,100]):
            tv_party.heading(c, text=c); tv_party.column(c, width=w, anchor="center")
        tv_party.tag_configure("debit",  foreground=C_RED)
        tv_party.tag_configure("credit", foreground=C_GREEN)
        sb12 = ttk.Scrollbar(tab_list, orient="vertical", command=tv_party.yview)
        tv_party.configure(yscrollcommand=sb12.set)
        sb12.pack(side="right", fill="y"); tv_party.pack(fill="both", expand=True)

        def load_parties(*a):
            tv_party.delete(*tv_party.get_children())
            s = v_psearch.get().strip()
            conn = get_db()
            q = "SELECT * FROM parties WHERE 1=1"
            params = []
            if s: q += " AND (name LIKE ? OR mobile LIKE ?)"; params += [f"%{s}%", f"%{s}%"]
            q += " ORDER BY name"
            rows = conn.execute(q, params).fetchall()
            for r in rows:
                bal = _party_balance(r["name"], conn)
                bal_str = f"{'Dr ' if bal>0 else 'Cr '}{rupee(abs(bal))}"
                tag = "debit" if bal > 0 else "credit"
                tv_party.insert("","end", tag=tag,
                    values=(r["name"], r["ptype"], r["mobile"], r["gstin"],
                            r["address"], r["state"], bal_str))
            conn.close()

        v_psearch.trace_add("write", load_parties)
        load_parties()

        # Double-click → open ledger for that party
        def on_party_dblclick(ev):
            sel = tv_party.selection()
            if not sel: return
            name = tv_party.item(sel[0])["values"][0]
            nb.select(tab_ledger)
            v_led_party.set(name)
            load_ledger()

        tv_party.bind("<Double-1>", on_party_dblclick)
        tk.Label(tab_list, text="💡 Party pe double-click karein Ledger dekhne ke liye",
                 font=("Segoe UI",8), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=2)

        # ═══════════════════════════════════════════════
        # TAB 2 — Ledger
        # ═══════════════════════════════════════════════
        tab_ledger = tk.Frame(nb, bg=C_BG); nb.add(tab_ledger, text="📒 Ledger")

        lf_top = tk.Frame(tab_ledger, bg=C_BG); lf_top.pack(fill="x", pady=(6,4))
        tk.Label(lf_top, text="Party:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn0 = get_db()
        all_party_names = [r["name"] for r in conn0.execute("SELECT name FROM parties ORDER BY name").fetchall()]
        conn0.close()
        v_led_party = tk.StringVar()
        led_party_cb = ttk.Combobox(lf_top, textvariable=v_led_party, width=28, values=all_party_names)
        led_party_cb.pack(side="left", padx=(4,12))
        tk.Label(lf_top, text="From:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_led_from = tk.StringVar()
        led_fde = make_date_entry(lf_top, v_led_from, width=12, bg=C_BG)
        led_fde.pack(side="left", padx=4)
        v_led_from.set(datetime.date.today().replace(day=1).strftime("%d-%m-%Y"))
        tk.Label(lf_top, text="To:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_led_to = tk.StringVar()
        led_tde = make_date_entry(lf_top, v_led_to, width=12, bg=C_BG)
        led_tde.pack(side="left", padx=4)
        led_load_btn = make_btn(lf_top, "📒 Load", lambda: None, bg=C_BLUE)
        led_load_btn.pack(side="left", padx=8)

        # Enter nav: Party → From → To → Load button
        bind_enter_nav([led_party_cb, led_fde.entry, led_tde.entry, led_load_btn])

        # Summary bar
        led_sum_frame = tk.Frame(tab_ledger, bg=C_DARK, height=32); led_sum_frame.pack(fill="x"); led_sum_frame.pack_propagate(False)
        led_sum_var = tk.StringVar(value="")
        tk.Label(led_sum_frame, textvariable=led_sum_var, font=("Segoe UI",9,"bold"),
                 bg=C_DARK, fg="white").pack(side="left", padx=12, pady=6)

        led_cols = ("Date","Type","Bill No","Debit (Dr)","Credit (Cr)","Balance","Remarks")
        led_tv = ttk.Treeview(tab_ledger, columns=led_cols, show="headings")
        for c, w in zip(led_cols, [90,80,130,100,100,110,200]):
            led_tv.heading(c, text=c); led_tv.column(c, width=w, anchor="center")
        led_tv.tag_configure("dr",      foreground=C_RED)
        led_tv.tag_configure("cr",      foreground=C_GREEN)
        led_tv.tag_configure("bal_row", background="#FFF3E0", font=("Segoe UI",9,"bold"))
        sb_led = ttk.Scrollbar(tab_ledger, orient="vertical", command=led_tv.yview)
        led_tv.configure(yscrollcommand=sb_led.set)
        sb_led.pack(side="right", fill="y"); led_tv.pack(fill="both", expand=True)

        def load_ledger():
            nm = v_led_party.get().strip()
            if not nm: messagebox.showwarning("","Party chunein!"); return
            try:
                fd = datetime.datetime.strptime(led_fde.var.get(), "%d-%m-%Y").strftime("%Y-%m-%d")
                td = datetime.datetime.strptime(led_tde.var.get(), "%d-%m-%Y").strftime("%Y-%m-%d")
            except:
                fd = datetime.date.today().replace(day=1).isoformat()
                td = datetime.date.today().isoformat()

            led_tv.delete(*led_tv.get_children())
            conn = get_db()

            # Collect all transactions for this party
            txns = []

            # Sales (debit to party — they owe us)
            sales = conn.execute(
                "SELECT bill_no, bill_date, grand_total, pay_mode, notes FROM sales "
                "WHERE party=? AND bill_date BETWEEN ? AND ? ORDER BY bill_date, id",
                (nm, fd, td)).fetchall()
            for s in sales:
                txns.append({
                    "date": s["bill_date"], "type": "Sale",
                    "ref": s["bill_no"], "dr": s["grand_total"], "cr": 0,
                    "remarks": f"Pay: {s['pay_mode']}" + (f" | {s['notes']}" if s["notes"] else "")
                })

            # Sale payments received (credit — party paid us)
            sale_pmts = conn.execute(
                "SELECT pay_date, bill_no, amount, pay_mode, note FROM bill_payments "
                "WHERE party=? AND bill_type='sale' AND pay_date BETWEEN ? AND ? ORDER BY pay_date, id",
                (nm, fd, td)).fetchall()
            for p in sale_pmts:
                txns.append({
                    "date": p["pay_date"], "type": "Receipt",
                    "ref": p["bill_no"], "dr": 0, "cr": p["amount"],
                    "remarks": f"{p['pay_mode']}" + (f" | {p['note']}" if p["note"] else "")
                })

            # Purchases (credit to party — we owe them)
            purchases = conn.execute(
                "SELECT bill_no, bill_date, grand_total, pay_mode FROM purchases "
                "WHERE party=? AND bill_date BETWEEN ? AND ? ORDER BY bill_date, id",
                (nm, fd, td)).fetchall()
            for p in purchases:
                txns.append({
                    "date": p["bill_date"], "type": "Purchase",
                    "ref": p["bill_no"], "dr": 0, "cr": p["grand_total"],
                    "remarks": f"Pay: {p['pay_mode']}"
                })

            # Purchase payments made (debit — we paid them)
            pur_pmts = conn.execute(
                "SELECT pay_date, bill_no, amount, pay_mode, note FROM bill_payments "
                "WHERE party=? AND bill_type='purchase' AND pay_date BETWEEN ? AND ? ORDER BY pay_date, id",
                (nm, fd, td)).fetchall()
            for p in pur_pmts:
                txns.append({
                    "date": p["pay_date"], "type": "Payment",
                    "ref": p["bill_no"], "dr": p["amount"], "cr": 0,
                    "remarks": f"{p['pay_mode']}" + (f" | {p['note']}" if p["note"] else "")
                })

            conn.close()

            # Sort by date
            txns.sort(key=lambda x: x["date"])

            total_dr = 0; total_cr = 0; running_bal = 0
            for t in txns:
                total_dr += t["dr"]; total_cr += t["cr"]
                running_bal += t["dr"] - t["cr"]
                bal_str = f"{'Dr ' if running_bal>0 else 'Cr '}{rupee(abs(running_bal))}"
                tag = "dr" if t["dr"] > 0 else "cr"
                led_tv.insert("","end", tag=tag, values=(
                    t["date"], t["type"], t["ref"],
                    rupee(t["dr"]) if t["dr"] else "—",
                    rupee(t["cr"]) if t["cr"] else "—",
                    bal_str, t["remarks"]
                ))

            # Closing balance row
            closing = total_dr - total_cr
            cbal_str = f"{'Dr (Payable)' if closing>0 else 'Cr (Receivable)'} {rupee(abs(closing))}"
            led_tv.insert("","end", tag="bal_row", values=(
                "", "━━ CLOSING", "━━━━━━━━━━",
                rupee(total_dr), rupee(total_cr), cbal_str, ""))

            led_sum_var.set(
                f"  Party: {nm}   |   Total Debit: {rupee(total_dr)}   |   "
                f"Total Credit: {rupee(total_cr)}   |   "
                f"Closing Balance: {cbal_str}"
            )

        # Bind load button and combobox after load_ledger is defined
        led_load_btn.config(command=load_ledger)
        led_party_cb.bind("<<ComboboxSelected>>", lambda e: load_ledger())

        # ═══════════════════════════════════════════════
        # TAB 3 — Payment Reminders
        # ═══════════════════════════════════════════════
        tab_rem = tk.Frame(nb, bg=C_BG); nb.add(tab_rem, text="🔔 Payment Reminders")

        rf_top = tk.Frame(tab_rem, bg=C_BG); rf_top.pack(fill="x", pady=(6,4))
        tk.Label(rf_top, text="Filter:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_rem_type = tk.StringVar(value="All")
        rem_filter_cb = ttk.Combobox(rf_top, textvariable=v_rem_type, width=14, state="readonly",
                     values=["All","Customer (Dr)","Supplier (Cr)"])
        rem_filter_cb.pack(side="left", padx=4)
        tk.Label(rf_top, text="Min Amount ₹:", font=("Segoe UI",9), bg=C_BG).pack(side="left", padx=(12,0))
        v_min_amt = tk.StringVar(value="0")
        rem_min_e = ttk.Entry(rf_top, textvariable=v_min_amt, width=10)
        rem_min_e.pack(side="left", padx=4)
        rem_refresh_btn = make_btn(rf_top, "🔄 Refresh", lambda: None, bg=C_ORANGE)
        rem_refresh_btn.pack(side="left", padx=8)

        # Enter nav: Filter → Min Amount → Refresh button
        bind_enter_nav([rem_filter_cb, rem_min_e, rem_refresh_btn])

        # Summary cards
        rem_card_frame = tk.Frame(tab_rem, bg=C_BG); rem_card_frame.pack(fill="x", padx=4, pady=(0,6))
        rem_total_dr_var  = tk.StringVar(value="₹0.00")
        rem_total_cr_var  = tk.StringVar(value="₹0.00")
        rem_count_dr_var  = tk.StringVar(value="0")
        rem_count_cr_var  = tk.StringVar(value="0")
        for col_i, (lbl, var, cnt, bg) in enumerate([
            ("💰 Total Receivable\n(Customers ke)",  rem_total_dr_var, rem_count_dr_var, C_RED),
            ("💸 Total Payable\n(Suppliers ko)",     rem_total_cr_var, rem_count_cr_var, C_GREEN),
        ]):
            card = tk.Frame(rem_card_frame, bg=bg, padx=20, pady=10); card.grid(row=0, column=col_i, padx=6, pady=4, sticky="ew")
            rem_card_frame.grid_columnconfigure(col_i, weight=1)
            tk.Label(card, textvariable=var, font=("Segoe UI",16,"bold"), bg=bg, fg="white").pack()
            tk.Label(card, text=lbl, font=("Segoe UI",8), bg=bg, fg="#FFE0B2").pack()
            tk.Label(card, textvariable=cnt, font=("Segoe UI",8), bg=bg, fg="#FFE0B2").pack()

        rem_cols = ("Party","Type","Mobile","Total Debit","Total Credit","Balance","Status")
        rem_tv = ttk.Treeview(tab_rem, columns=rem_cols, show="headings")
        for c, w in zip(rem_cols, [180,80,110,110,110,130,120]):
            rem_tv.heading(c, text=c); rem_tv.column(c, width=w, anchor="center")
        rem_tv.tag_configure("overdue",  background="#FFEBEE", foreground=C_RED)
        rem_tv.tag_configure("payable",  background="#E8F5E9", foreground=C_GREEN)
        rem_tv.tag_configure("settled",  foreground=C_GRAY)
        sb_rem = ttk.Scrollbar(tab_rem, orient="vertical", command=rem_tv.yview)
        rem_tv.configure(yscrollcommand=sb_rem.set)
        sb_rem.pack(side="right", fill="y"); rem_tv.pack(fill="both", expand=True)

        tk.Label(tab_rem,
                 text="🔴 Laal = Customer ka paisa baaki (Receivable)   🟢 Hara = Supplier ko dena hai (Payable)",
                 font=("Segoe UI",8), bg=C_BG, fg=C_DARK).pack(anchor="w", pady=2)

        def load_reminders():
            rem_tv.delete(*rem_tv.get_children())
            try: min_amt = float(v_min_amt.get() or 0)
            except: min_amt = 0
            ftype = v_rem_type.get()

            conn = get_db()
            parties = conn.execute("SELECT name, ptype, mobile FROM parties ORDER BY name").fetchall()

            total_receivable = 0; total_payable = 0
            count_dr = 0; count_cr = 0

            for p in parties:
                nm = p["name"]
                # Total sales to party
                sale_total = conn.execute(
                    "SELECT COALESCE(SUM(grand_total),0) as t FROM sales WHERE party=?", (nm,)).fetchone()["t"]
                # Payments received from party
                sale_recv = conn.execute(
                    "SELECT COALESCE(SUM(amount),0) as t FROM bill_payments WHERE party=? AND bill_type='sale'", (nm,)).fetchone()["t"]
                # Total purchases from party
                pur_total = conn.execute(
                    "SELECT COALESCE(SUM(grand_total),0) as t FROM purchases WHERE party=?", (nm,)).fetchone()["t"]
                # Payments made to party
                pur_paid = conn.execute(
                    "SELECT COALESCE(SUM(amount),0) as t FROM bill_payments WHERE party=? AND bill_type='purchase'", (nm,)).fetchone()["t"]

                net_dr = round(sale_total - sale_recv, 2)   # Party hamein dega (receivable)
                net_cr = round(pur_total - pur_paid,  2)    # Hum party ko denge (payable)
                balance = round(net_dr - net_cr, 2)

                if abs(balance) < 0.01 and abs(net_dr) < 0.01 and abs(net_cr) < 0.01:
                    if ftype != "All": continue
                if abs(balance) < min_amt and min_amt > 0: continue

                if ftype == "Customer (Dr)" and balance <= 0: continue
                if ftype == "Supplier (Cr)" and balance >= 0: continue

                if balance > 0:
                    status = f"⚠️ Lena Hai"
                    tag = "overdue"
                    total_receivable += balance; count_dr += 1
                elif balance < 0:
                    status = f"💸 Dena Hai"
                    tag = "payable"
                    total_payable += abs(balance); count_cr += 1
                else:
                    status = "✅ Settled"
                    tag = "settled"

                bal_str = f"{'Dr ' if balance>0 else 'Cr '}{rupee(abs(balance))}" if balance != 0 else "✅ 0"
                rem_tv.insert("","end", tag=tag, values=(
                    nm, p["ptype"], p["mobile"],
                    rupee(net_dr), rupee(net_cr), bal_str, status))

            conn.close()
            rem_total_dr_var.set(rupee(total_receivable))
            rem_total_cr_var.set(rupee(total_payable))
            rem_count_dr_var.set(f"{count_dr} parties")
            rem_count_cr_var.set(f"{count_cr} parties")

        # Bind refresh button after load_reminders is defined
        rem_refresh_btn.config(command=load_reminders)
        v_rem_type.trace_add("write", lambda *a: load_reminders())
        load_reminders()

        # Double-click on reminder → open ledger
        def on_rem_dblclick(ev):
            sel = rem_tv.selection()
            if not sel: return
            name = rem_tv.item(sel[0])["values"][0]
            nb.select(tab_ledger)
            v_led_party.set(name)
            load_ledger()

        rem_tv.bind("<Double-1>", on_rem_dblclick)
        tk.Label(tab_rem, text="💡 Party pe double-click karein Ledger dekhne ke liye",
                 font=("Segoe UI",8), bg=C_BG, fg=C_GRAY).pack(anchor="w")

        # ── Action buttons: WhatsApp + Email reminder ──
        act_row = tk.Frame(tab_rem, bg=C_BG); act_row.pack(fill="x", pady=(4,6))

        def send_wa_reminder():
            sel = rem_tv.selection()
            if not sel:
                messagebox.showwarning("","Pehle ek party select karein!"); return
            vals = rem_tv.item(sel[0])["values"]
            name   = vals[0]
            mobile = str(vals[2]).strip()
            bal    = str(vals[5]).strip()
            if not mobile or mobile == "-":
                messagebox.showwarning("","Is party ka mobile number nahi hai!"); return
            mob = mobile.replace(" ","").replace("+","").replace("-","")
            if not mob.startswith("91"): mob = "91" + mob
            shop = get_shop()
            msg = (
                f"Namaskar {name}!\n\n"
                f"Aapka {bal} outstanding baki hai.\n"
                f"Kripya jald se jald payment karein.\n\n"
                f"Thank you,\n{shop.get('name','')}\n{shop.get('mobile','')}"
            )
            import urllib.parse, webbrowser
            webbrowser.open(f"https://wa.me/{mob}?text={urllib.parse.quote(msg)}")

        def send_email_reminder():
            sel = rem_tv.selection()
            if not sel:
                messagebox.showwarning("","Pehle ek party select karein!"); return
            vals = rem_tv.item(sel[0])["values"]
            name = vals[0]
            bal  = str(vals[5]).strip()
            conn2 = get_db()
            p = conn2.execute("SELECT email FROM parties WHERE name=?", (name,)).fetchone()
            conn2.close()
            email = (p["email"] if p else "") or ""
            if not email:
                messagebox.showwarning("","Is party ka email nahi hai!"); return
            shop = get_shop()
            import urllib.parse, webbrowser
            subj = urllib.parse.quote(f"Payment Reminder - {shop.get('name','')}")
            body = urllib.parse.quote(
                f"Dear {name},\n\nAapka {bal} outstanding baki hai.\n"
                f"Kripya jald payment karein.\n\nRegards,\n{shop.get('name','')}\n{shop.get('mobile','')}"
            )
            webbrowser.open(f"mailto:{email}?subject={subj}&body={body}")

        make_btn(act_row, "💬 WhatsApp Reminder", send_wa_reminder, bg="#25D366", pady=4).pack(side="left", padx=4)
        make_btn(act_row, "📧 Email Reminder",    send_email_reminder, bg=C_BLUE, pady=4).pack(side="left", padx=4)

    def _add_party(self):
        dlg = tk.Toplevel(self.root)
        dlg.title("Party Add"); dlg.configure(bg=C_BG); dlg.geometry("480x440"); dlg.grab_set()
        tk.Label(dlg, text="👥 Party Details", font=("Segoe UI",12,"bold"),
                 bg=C_ORANGE, fg="white").pack(fill="x", ipady=10)
        form = tk.Frame(dlg, bg=C_BG, padx=16, pady=12); form.pack(fill="x")
        fields = {}
        for label, key, default in [
            ("Name *","name",""), ("Type","ptype","Customer"), ("Mobile","mobile",""),
            ("GSTIN","gstin",""), ("Address","address",""),
            ("State","state","Uttar Pradesh"), ("Email","email","")
        ]:
            f = tk.Frame(form, bg=C_BG); f.pack(fill="x", pady=4)
            tk.Label(f, text=label, font=("Segoe UI",9), bg=C_BG, width=14, anchor="w").pack(side="left")
            if key == "ptype":
                var = tk.StringVar(value=default)
                ttk.Combobox(f, textvariable=var, width=16, state="readonly",
                             values=["Customer","Supplier","Both"]).pack(side="left")
            else:
                var = tk.StringVar(value=default)
                ttk.Entry(f, textvariable=var, width=28).pack(side="left")
            fields[key] = var
        def save():
            if not fields["name"].get().strip(): messagebox.showwarning("","Name zaroori!"); return
            conn = get_db()
            try:
                conn.execute("INSERT INTO parties (name,ptype,mobile,gstin,address,state,email) VALUES(?,?,?,?,?,?,?)",
                    (fields["name"].get().strip(), fields["ptype"].get(), fields["mobile"].get(),
                     fields["gstin"].get(), fields["address"].get(), fields["state"].get(), fields["email"].get()))
                conn.commit(); conn.close()
                messagebox.showinfo("Saved!","Party save ho gayi!"); dlg.destroy(); self._show_parties()
            except Exception as ex:
                conn.close(); messagebox.showerror("Error",str(ex))
        party_save_btn = make_btn(dlg, "💾 Save Party", save, bg=C_GREEN, pady=10)
        party_save_btn.pack(fill="x", padx=16, pady=8)

        # Enter nav for party form
        party_entries = [w for f in form.winfo_children()
                         for w in f.winfo_children()
                         if isinstance(w, (ttk.Entry, ttk.Combobox))]
        bind_enter_nav(party_entries + [party_save_btn])

    # ── REPORTS ───────────────────────────────────────────────────────────────
    def _show_reports(self):
        self._clear()
        self._section_header("📊 Reports", "Business analysis, GST & Balance Sheet")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # Date range
        df = tk.Frame(main, bg=C_BG); df.pack(fill="x", pady=(0,10))
        tk.Label(df, text="From:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_from = tk.StringVar()
        fde = make_date_entry(df, v_from, width=12, bg=C_BG); fde.pack(side="left", padx=4)
        v_from.set(datetime.date.today().replace(day=1).strftime("%d-%m-%Y"))
        tk.Label(df, text="To:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_to = tk.StringVar()
        tde = make_date_entry(df, v_to, width=12, bg=C_BG); tde.pack(side="left", padx=4)

        def get_dates():
            try:
                fd = datetime.datetime.strptime(fde.var.get(), "%d-%m-%Y").strftime("%Y-%m-%d")
            except:
                fd = datetime.date.today().replace(day=1).isoformat()
            try:
                td = datetime.datetime.strptime(tde.var.get(), "%d-%m-%Y").strftime("%Y-%m-%d")
            except:
                td = datetime.date.today().isoformat()
            return fd, td

        nb = ttk.Notebook(main); nb.pack(fill="both", expand=True)

        # ── Tab 1: Sale Summary ──
        tab1 = tk.Frame(nb, bg=C_BG); nb.add(tab1, text="💰 Sale Report")
        cols1 = ("Date","Bill No","Customer","Bill Type","Taxable","GST","Grand Total","Payment")
        tv1 = ttk.Treeview(tab1, columns=cols1, show="headings")
        for c, w in zip(cols1, [90,130,160,80,90,80,100,90]):
            tv1.heading(c, text=c); tv1.column(c, width=w, anchor="center")
        sb13 = ttk.Scrollbar(tab1, orient="vertical", command=tv1.yview)
        tv1.configure(yscrollcommand=sb13.set)
        sb13.pack(side="right", fill="y"); tv1.pack(fill="both", expand=True)
        sum_lbl1 = tk.Label(tab1, text="", font=("Segoe UI",11,"bold"), bg=C_BG, fg=C_ORANGE)
        sum_lbl1.pack(pady=4)

        # ── Tab 2: Product-wise ──
        tab2 = tk.Frame(nb, bg=C_BG); nb.add(tab2, text="🛍️ Product-wise Sale")
        cols2 = ("Product","Total Qty","Unit","Total Amount")
        tv2 = ttk.Treeview(tab2, columns=cols2, show="headings")
        for c, w in zip(cols2, [250,100,60,120]):
            tv2.heading(c, text=c); tv2.column(c, width=w, anchor="center")
        sb14 = ttk.Scrollbar(tab2, orient="vertical", command=tv2.yview)
        tv2.configure(yscrollcommand=sb14.set)
        sb14.pack(side="right", fill="y"); tv2.pack(fill="both", expand=True)

        # ── Tab 3: P&L ──
        tab3 = tk.Frame(nb, bg=C_BG); nb.add(tab3, text="📈 P&L Summary")
        pl_text = tk.Text(tab3, font=("Segoe UI",11), bg=C_YELLOW, padx=20, pady=16, wrap="word")
        pl_text.pack(fill="both", expand=True)

        # ── Tab 4: GSTR-1 ──
        tab4 = tk.Frame(nb, bg=C_BG); nb.add(tab4, text="📋 GSTR-1")
        tk.Label(tab4, text="GSTR-1 — Outward Supplies (B2C/B2B)",
                 font=("Segoe UI",11,"bold"), bg=C_ORANGE, fg="white").pack(fill="x", ipady=6)

        gstr1_nb = ttk.Notebook(tab4); gstr1_nb.pack(fill="both", expand=True, pady=4)

        # GSTR-1: B2B (GST bills with GSTIN)
        g1_b2b_tab = tk.Frame(gstr1_nb, bg=C_BG); gstr1_nb.add(g1_b2b_tab, text="B2B (GSTIN wale)")
        g1b2b_cols = ("Bill No","Date","Customer","GSTIN","Taxable","CGST","SGST","IGST","Total")
        g1_b2b_tv = ttk.Treeview(g1_b2b_tab, columns=g1b2b_cols, show="headings")
        for c, w in zip(g1b2b_cols, [120,90,150,160,90,70,70,70,90]):
            g1_b2b_tv.heading(c, text=c); g1_b2b_tv.column(c, width=w, anchor="center")
        sb_g1b2b = ttk.Scrollbar(g1_b2b_tab, orient="vertical", command=g1_b2b_tv.yview)
        g1_b2b_tv.configure(yscrollcommand=sb_g1b2b.set)
        sb_g1b2b.pack(side="right", fill="y"); g1_b2b_tv.pack(fill="both", expand=True)
        g1b2b_sum = tk.Label(g1_b2b_tab, text="", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_DARK)
        g1b2b_sum.pack(pady=3)

        # GSTR-1: B2C (GST bills without GSTIN)
        g1_b2c_tab = tk.Frame(gstr1_nb, bg=C_BG); gstr1_nb.add(g1_b2c_tab, text="B2C (Retail / Bina GSTIN)")
        g1b2c_cols = ("Bill No","Date","Customer","Taxable","CGST","SGST","Total")
        g1_b2c_tv = ttk.Treeview(g1_b2c_tab, columns=g1b2c_cols, show="headings")
        for c, w in zip(g1b2c_cols, [130,90,180,100,80,80,100]):
            g1_b2c_tv.heading(c, text=c); g1_b2c_tv.column(c, width=w, anchor="center")
        sb_g1b2c = ttk.Scrollbar(g1_b2c_tab, orient="vertical", command=g1_b2c_tv.yview)
        g1_b2c_tv.configure(yscrollcommand=sb_g1b2c.set)
        sb_g1b2c.pack(side="right", fill="y"); g1_b2c_tv.pack(fill="both", expand=True)
        g1b2c_sum = tk.Label(g1_b2c_tab, text="", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_DARK)
        g1b2c_sum.pack(pady=3)

        # GSTR-1: HSN Summary
        g1_hsn_tab = tk.Frame(gstr1_nb, bg=C_BG); gstr1_nb.add(g1_hsn_tab, text="HSN Summary")
        g1hsn_cols = ("HSN","Description","UQC","Total Qty","Taxable","CGST","SGST","Total GST")
        g1_hsn_tv = ttk.Treeview(g1_hsn_tab, columns=g1hsn_cols, show="headings")
        for c, w in zip(g1hsn_cols, [80,180,55,80,100,80,80,90]):
            g1_hsn_tv.heading(c, text=c); g1_hsn_tv.column(c, width=w, anchor="center")
        sb_g1hsn = ttk.Scrollbar(g1_hsn_tab, orient="vertical", command=g1_hsn_tv.yview)
        g1_hsn_tv.configure(yscrollcommand=sb_g1hsn.set)
        sb_g1hsn.pack(side="right", fill="y"); g1_hsn_tv.pack(fill="both", expand=True)

        # ── Tab 5: GSTR-3B ──
        tab5 = tk.Frame(nb, bg=C_BG); nb.add(tab5, text="📝 GSTR-3B")
        tk.Label(tab5, text="GSTR-3B — Monthly Summary Return",
                 font=("Segoe UI",11,"bold"), bg=C_ORANGE, fg="white").pack(fill="x", ipady=6)
        gstr3b_text = tk.Text(tab5, font=("Courier New",10), bg=C_YELLOW, padx=20, pady=14, wrap="word")
        gstr3b_text.pack(fill="both", expand=True)

        # ── Tab 6: Balance Sheet ──
        tab6 = tk.Frame(nb, bg=C_BG); nb.add(tab6, text="⚖️ Balance Sheet")
        tk.Label(tab6, text="Balance Sheet — Assets & Liabilities",
                 font=("Segoe UI",11,"bold"), bg=C_ORANGE, fg="white").pack(fill="x", ipady=6)
        bs_text = tk.Text(tab6, font=("Courier New",10), bg=C_YELLOW, padx=20, pady=14, wrap="word")
        bs_text.pack(fill="both", expand=True)

        # ═══════════════ LOAD ALL REPORTS ═══════════════
        def load_reports():
            fd, td = get_dates()
            conn = get_db()

            # ── Sale Report ──
            tv1.delete(*tv1.get_children())
            sales = conn.execute("""
                SELECT s.*, COALESCE(SUM(si.gst_amt),0) as total_gst,
                       COALESCE(SUM(si.taxable),0) as total_taxable
                FROM sales s LEFT JOIN sale_items si ON si.sale_id=s.id
                WHERE s.bill_date BETWEEN ? AND ? GROUP BY s.id ORDER BY s.bill_date
            """, (fd, td)).fetchall()
            total_sale = 0; total_gst_sale = 0
            for r in sales:
                tv1.insert("","end", values=(
                    r["bill_date"], r["bill_no"], r["party"], r["gst_type"],
                    rupee(r["total_taxable"]), rupee(r["total_gst"]),
                    rupee(r["grand_total"]), r["pay_mode"]))
                total_sale += r["grand_total"]; total_gst_sale += r["total_gst"]
            sum_lbl1.config(text=f"Total Sale: {rupee(total_sale)} | GST Collected: {rupee(total_gst_sale)} | Bills: {len(sales)}")

            # ── Product-wise ──
            tv2.delete(*tv2.get_children())
            pwise = conn.execute("""
                SELECT si.product, si.unit, SUM(si.qty) as tqty, SUM(si.grand) as tamt
                FROM sale_items si JOIN sales s ON s.id=si.sale_id
                WHERE s.bill_date BETWEEN ? AND ? GROUP BY si.product ORDER BY tamt DESC
            """, (fd, td)).fetchall()
            for r in pwise:
                tv2.insert("","end", values=(r["product"], f"{r['tqty']:.2f}", r["unit"], rupee(r["tamt"])))

            # ── P&L ──
            total_purchase = conn.execute(
                "SELECT COALESCE(SUM(grand_total),0) as t FROM purchases WHERE bill_date BETWEEN ? AND ?",
                (fd, td)).fetchone()["t"]
            total_expense = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as t FROM expenses WHERE exp_date BETWEEN ? AND ?",
                (fd, td)).fetchone()["t"]
            gross_profit = total_sale - total_purchase
            net_profit   = gross_profit - total_expense

            pl_text.config(state="normal"); pl_text.delete("1.0","end")
            pl_text.insert("end", f"📅 Period: {fd}  to  {td}\n\n", "heading")
            pl_text.insert("end", f"{'INCOME':─^44}\n")
            pl_text.insert("end", f"  Total Sale Revenue:      {rupee(total_sale):>16}\n")
            pl_text.insert("end", f"\n{'EXPENSES':─^44}\n")
            pl_text.insert("end", f"  Total Purchase Cost:     {rupee(total_purchase):>16}\n")
            pl_text.insert("end", f"  Other Expenses:          {rupee(total_expense):>16}\n")
            pl_text.insert("end", f"\n{'━'*44}\n")
            pl_text.insert("end", f"  Gross Profit:            {rupee(gross_profit):>16}\n",
                           "green" if gross_profit >= 0 else "red")
            pl_text.insert("end", f"  Net Profit:              {rupee(net_profit):>16}\n",
                           "big_green" if net_profit >= 0 else "big_red")
            pl_text.insert("end", f"{'━'*44}\n")
            pl_text.insert("end", f"\n  GST Collected (Sales):   {rupee(total_gst_sale):>16}\n")
            for tag, cfg in [
                ("heading",   {"font":("Segoe UI",12,"bold"),  "foreground":C_ORANGE}),
                ("green",     {"font":("Segoe UI",11,"bold"),  "foreground":C_GREEN}),
                ("red",       {"font":("Segoe UI",11,"bold"),  "foreground":C_RED}),
                ("big_green", {"font":("Segoe UI",14,"bold"),  "foreground":C_GREEN}),
                ("big_red",   {"font":("Segoe UI",14,"bold"),  "foreground":C_RED}),
            ]:
                pl_text.tag_config(tag, **cfg)
            pl_text.config(state="disabled")

            # ════════════════════════════════════════════
            # ── GSTR-1 ──
            # Only GST-type bills
            gst_sales = conn.execute("""
                SELECT s.*, COALESCE(SUM(si.gst_amt),0) as t_gst,
                       COALESCE(SUM(si.taxable),0) as t_taxable
                FROM sales s LEFT JOIN sale_items si ON si.sale_id=s.id
                WHERE s.bill_date BETWEEN ? AND ? AND s.gst_type='GST'
                GROUP BY s.id ORDER BY s.bill_date
            """, (fd, td)).fetchall()

            # B2B (has GSTIN)
            g1_b2b_tv.delete(*g1_b2b_tv.get_children())
            b2b_taxable=0; b2b_cgst=0; b2b_sgst=0; b2b_total=0
            for r in gst_sales:
                if r["party_gstin"] and r["party_gstin"].strip():
                    cgst = round(r["t_gst"]/2, 2); sgst = round(r["t_gst"]/2, 2)
                    g1_b2b_tv.insert("","end", values=(
                        r["bill_no"], r["bill_date"], r["party"], r["party_gstin"],
                        rupee(r["t_taxable"]), rupee(cgst), rupee(sgst), "₹0.00",
                        rupee(r["grand_total"])))
                    b2b_taxable += r["t_taxable"]; b2b_cgst += cgst
                    b2b_sgst += sgst; b2b_total += r["grand_total"]
            g1b2b_sum.config(text=f"B2B Total — Taxable: {rupee(b2b_taxable)} | CGST: {rupee(b2b_cgst)} | SGST: {rupee(b2b_sgst)} | Grand: {rupee(b2b_total)}")

            # B2C (no GSTIN)
            g1_b2c_tv.delete(*g1_b2c_tv.get_children())
            b2c_taxable=0; b2c_cgst=0; b2c_sgst=0; b2c_total=0
            for r in gst_sales:
                if not (r["party_gstin"] and r["party_gstin"].strip()):
                    cgst = round(r["t_gst"]/2, 2); sgst = round(r["t_gst"]/2, 2)
                    g1_b2c_tv.insert("","end", values=(
                        r["bill_no"], r["bill_date"], r["party"],
                        rupee(r["t_taxable"]), rupee(cgst), rupee(sgst),
                        rupee(r["grand_total"])))
                    b2c_taxable += r["t_taxable"]; b2c_cgst += cgst
                    b2c_sgst += sgst; b2c_total += r["grand_total"]
            g1b2c_sum.config(text=f"B2C Total — Taxable: {rupee(b2c_taxable)} | CGST: {rupee(b2c_cgst)} | SGST: {rupee(b2c_sgst)} | Grand: {rupee(b2c_total)}")

            # HSN Summary
            g1_hsn_tv.delete(*g1_hsn_tv.get_children())
            hsn_data = conn.execute("""
                SELECT si.hsn, si.unit,
                       SUM(si.qty) as tqty, SUM(si.taxable) as ttax,
                       SUM(si.gst_amt) as tgst
                FROM sale_items si JOIN sales s ON s.id=si.sale_id
                WHERE s.bill_date BETWEEN ? AND ? AND s.gst_type='GST'
                GROUP BY si.hsn ORDER BY si.hsn
            """, (fd, td)).fetchall()
            for h in hsn_data:
                cgst = round(h["tgst"]/2, 2); sgst = round(h["tgst"]/2, 2)
                g1_hsn_tv.insert("","end", values=(
                    h["hsn"] or "3208", "Car Paint/Coating", h["unit"] or "Ltr",
                    f"{h['tqty']:.2f}", rupee(h["ttax"]),
                    rupee(cgst), rupee(sgst), rupee(h["tgst"])))

            # ════════════════════════════════════════════
            # ── GSTR-3B ──
            # Output tax (sales)
            out_taxable  = sum(r["t_taxable"] for r in gst_sales)
            out_gst      = sum(r["t_gst"]     for r in gst_sales)
            out_cgst     = round(out_gst/2, 2)
            out_sgst     = round(out_gst/2, 2)

            # Input tax credit (purchases)
            pur_gst_data = conn.execute("""
                SELECT COALESCE(SUM(taxable),0) as t_tax, COALESCE(SUM(gst_amt),0) as t_gst
                FROM purchase_items pi JOIN purchases p ON p.id=pi.purchase_id
                WHERE p.bill_date BETWEEN ? AND ?
            """, (fd, td)).fetchone()
            itc_taxable = pur_gst_data["t_tax"]
            itc_total   = pur_gst_data["t_gst"]
            itc_cgst    = round(itc_total/2, 2)
            itc_sgst    = round(itc_total/2, 2)
            net_cgst    = round(out_cgst - itc_cgst, 2)
            net_sgst    = round(out_sgst - itc_sgst, 2)
            net_gst_payable = round(net_cgst + net_sgst, 2)

            W3 = 46
            gstr3b_text.config(state="normal"); gstr3b_text.delete("1.0","end")
            gstr3b_text.insert("end", f"  GSTR-3B — Monthly Return Summary\n", "heading")
            gstr3b_text.insert("end", f"  Period: {fd}  to  {td}\n\n", "sub")
            gstr3b_text.insert("end", f"{'─'*W3}\n")
            gstr3b_text.insert("end", "  3.1  OUTWARD SUPPLIES (Aapki Bikri)\n", "section")
            gstr3b_text.insert("end", f"{'─'*W3}\n")
            gstr3b_text.insert("end", f"  (a) GST Bills (Pakka Bill):\n")
            gstr3b_text.insert("end", f"      Taxable Value     : {rupee(out_taxable):>14}\n")
            gstr3b_text.insert("end", f"      CGST (9%)         : {rupee(out_cgst):>14}\n")
            gstr3b_text.insert("end", f"      SGST (9%)         : {rupee(out_sgst):>14}\n")
            gstr3b_text.insert("end", f"      Total GST Output  : {rupee(out_gst):>14}\n", "bold")
            gstr3b_text.insert("end", f"\n{'─'*W3}\n")
            gstr3b_text.insert("end", "  4.   INPUT TAX CREDIT (Purchases)\n", "section")
            gstr3b_text.insert("end", f"{'─'*W3}\n")
            gstr3b_text.insert("end", f"      Purchase Taxable  : {rupee(itc_taxable):>14}\n")
            gstr3b_text.insert("end", f"      ITC CGST          : {rupee(itc_cgst):>14}\n")
            gstr3b_text.insert("end", f"      ITC SGST          : {rupee(itc_sgst):>14}\n")
            gstr3b_text.insert("end", f"      Total ITC         : {rupee(itc_total):>14}\n", "bold")
            gstr3b_text.insert("end", f"\n{'═'*W3}\n")
            gstr3b_text.insert("end", "  6.   NET TAX PAYABLE\n", "section")
            gstr3b_text.insert("end", f"{'═'*W3}\n")
            gstr3b_text.insert("end", f"      Net CGST Payable  : {rupee(net_cgst):>14}\n")
            gstr3b_text.insert("end", f"      Net SGST Payable  : {rupee(net_sgst):>14}\n")
            clr_pay = "big_green" if net_gst_payable <= 0 else "big_red"
            gstr3b_text.insert("end", f"  ▶   TOTAL GST PAYABLE : {rupee(net_gst_payable):>14}\n", clr_pay)
            gstr3b_text.insert("end", f"{'═'*W3}\n")
            gstr3b_text.insert("end", "\n  ⚠️  Note: Ye sirf estimate hai. GSTIN portal pe\n")
            gstr3b_text.insert("end", "  actual return file karna aavashyak hai.\n", "note")
            for tag, cfg in [
                ("heading",   {"font":("Courier New",12,"bold"), "foreground":C_ORANGE}),
                ("sub",       {"font":("Courier New",9),         "foreground":C_GRAY}),
                ("section",   {"font":("Courier New",10,"bold"), "foreground":C_DARK}),
                ("bold",      {"font":("Courier New",10,"bold"), "foreground":"#111"}),
                ("big_green", {"font":("Courier New",12,"bold"), "foreground":C_GREEN}),
                ("big_red",   {"font":("Courier New",12,"bold"), "foreground":C_RED}),
                ("note",      {"font":("Courier New",9),         "foreground":C_GRAY}),
            ]:
                gstr3b_text.tag_config(tag, **cfg)
            gstr3b_text.config(state="disabled")

            # ════════════════════════════════════════════
            # ── BALANCE SHEET ──
            # Assets
            # 1. Stock value (current)
            all_prods = conn.execute("SELECT name, opening_stock, sale_rate, purchase_rate FROM products").fetchall()
            stock_value = 0
            for p in all_prods:
                pur_q = conn.execute("SELECT COALESCE(SUM(qty),0) as t FROM purchase_items WHERE product=?", (p["name"],)).fetchone()["t"]
                sal_q = conn.execute("SELECT COALESCE(SUM(qty),0) as t FROM sale_items     WHERE product=?", (p["name"],)).fetchone()["t"]
                curr  = p["opening_stock"] + pur_q - sal_q
                stock_value += curr * p["purchase_rate"]

            # 2. Cash in hand = total received (paid bills) - total paid (cash purchases)
            total_received  = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM sales WHERE pay_mode!='Credit'").fetchone()["t"]
            total_paid_cash = conn.execute("SELECT COALESCE(SUM(grand_total),0) as t FROM purchases WHERE pay_mode IN ('Cash','Online/UPI','Cheque')").fetchone()["t"]
            expenses_paid   = conn.execute("SELECT COALESCE(SUM(amount),0) as t FROM expenses").fetchone()["t"]
            cash_in_hand    = total_received - total_paid_cash - expenses_paid

            # Liabilities
            # 1. Creditors (credit purchases unpaid)
            credit_pur = conn.execute(
                "SELECT COALESCE(SUM(grand_total),0) as t FROM purchases WHERE pay_mode='Credit'").fetchone()["t"]
            credit_pur_paid = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as t FROM bill_payments WHERE bill_type='purchase'").fetchone()["t"]
            creditors = max(0, credit_pur - credit_pur_paid)

            # 2. GST Payable (net)
            gst_payable = max(0, net_gst_payable)

            # 3. Debtors (credit sales unpaid)
            credit_sale = conn.execute(
                "SELECT COALESCE(SUM(grand_total),0) as t FROM sales WHERE pay_mode='Credit'").fetchone()["t"]
            credit_sale_recv = conn.execute(
                "SELECT COALESCE(SUM(amount),0) as t FROM bill_payments WHERE bill_type='sale'").fetchone()["t"]
            debtors = max(0, credit_sale - credit_sale_recv)

            # Capital = Assets - Liabilities
            total_assets      = stock_value + max(0, cash_in_hand) + debtors
            total_liabilities = creditors + gst_payable
            capital           = total_assets - total_liabilities

            conn.close()

            WB = 46
            bs_text.config(state="normal"); bs_text.delete("1.0","end")
            bs_text.insert("end", "  BALANCE SHEET\n", "heading")
            bs_text.insert("end", f"  As on: {td}\n\n", "sub")

            bs_text.insert("end", f"  {'ASSETS':─^{WB}}\n", "section")
            bs_text.insert("end", f"  Stock-in-Hand (at cost) : {rupee(stock_value):>12}\n")
            bs_text.insert("end", f"  Sundry Debtors          : {rupee(debtors):>12}\n")
            bs_text.insert("end", f"  Cash / Bank             : {rupee(max(0,cash_in_hand)):>12}\n")
            bs_text.insert("end", f"  {'':─^{WB}}\n")
            bs_text.insert("end", f"  TOTAL ASSETS            : {rupee(total_assets):>12}\n", "bold")

            bs_text.insert("end", f"\n  {'LIABILITIES':─^{WB}}\n", "section")
            bs_text.insert("end", f"  Sundry Creditors        : {rupee(creditors):>12}\n")
            bs_text.insert("end", f"  GST Payable (Net)       : {rupee(gst_payable):>12}\n")
            bs_text.insert("end", f"  {'':─^{WB}}\n")
            bs_text.insert("end", f"  TOTAL LIABILITIES       : {rupee(total_liabilities):>12}\n", "bold")

            bs_text.insert("end", f"\n  {'':═^{WB}}\n")
            clr_cap = "big_green" if capital >= 0 else "big_red"
            bs_text.insert("end", f"  NET CAPITAL / EQUITY    : {rupee(capital):>12}\n", clr_cap)
            bs_text.insert("end", f"  {'':═^{WB}}\n")
            bs_text.insert("end", "\n  ⚠️  Note: Ye ek simple working capital based\n")
            bs_text.insert("end", "  balance sheet hai. CA se verify karein.\n", "note")
            for tag, cfg in [
                ("heading",   {"font":("Courier New",13,"bold"), "foreground":C_ORANGE}),
                ("sub",       {"font":("Courier New",9),         "foreground":C_GRAY}),
                ("section",   {"font":("Courier New",10,"bold"), "foreground":C_DARK}),
                ("bold",      {"font":("Courier New",11,"bold"), "foreground":"#111"}),
                ("big_green", {"font":("Courier New",13,"bold"), "foreground":C_GREEN}),
                ("big_red",   {"font":("Courier New",13,"bold"), "foreground":C_RED}),
                ("note",      {"font":("Courier New",9),         "foreground":C_GRAY}),
            ]:
                bs_text.tag_config(tag, **cfg)
            bs_text.config(state="disabled")

        load_reports_btn = make_btn(df, "📊 Load Reports", load_reports, bg=C_ORANGE)
        load_reports_btn.pack(side="left", padx=12)
        bind_enter_nav([fde.entry, tde.entry, load_reports_btn])
        load_reports()



    # ══════════════════════════════════════════════════════════════════════════
    # SALE RETURN
    # ══════════════════════════════════════════════════════════════════════════
    def _show_sale_return(self):
        self._clear()
        self._section_header("🔄 Sale Return", "Grahak se wapas aaya maal")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # ── Top: New Return form ──
        form = tk.LabelFrame(main, text="Naya Sale Return", font=("Segoe UI",9,"bold"),
                             bg=C_BG, fg=C_DARK, padx=8, pady=6)
        form.pack(fill="x", pady=(0,8))

        row1 = tk.Frame(form, bg=C_BG); row1.pack(fill="x", pady=2)
        tk.Label(row1, text="Return No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn0 = get_db()
        rno = conn0.execute("SELECT COUNT(*) as c FROM sale_returns").fetchone()["c"] + 1
        conn0.close()
        v_rno   = tk.StringVar(value=f"SR/{rno:04d}")
        sr_rno_e = ttk.Entry(row1, textvariable=v_rno, width=12); sr_rno_e.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Date:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_date  = tk.StringVar()
        sr_date_dp = make_date_entry(row1, v_date, width=12); sr_date_dp.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Orig Bill No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_orig  = tk.StringVar()
        conn0   = get_db()
        sale_bills = [r["bill_no"] for r in conn0.execute(
            "SELECT bill_no FROM sales ORDER BY bill_date DESC, id DESC").fetchall()]
        conn0.close()
        orig_cb = ttk.Combobox(row1, textvariable=v_orig, values=sale_bills, width=16)
        orig_cb.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Party:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_party = tk.StringVar()
        party_cb = ttk.Combobox(row1, textvariable=v_party, width=18, values=[])
        party_cb.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Reason:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_reason = tk.StringVar(value="Defective")
        ttk.Combobox(row1, textvariable=v_reason, width=14,
                     values=["Defective","Wrong Product","Excess Qty","Quality Issue","Other"]
                     ).pack(side="left", padx=(4,0))

        # ── Item entry row ──
        irow = tk.Frame(form, bg=C_BG); irow.pack(fill="x", pady=(6,2))
        tk.Label(irow, text="Product:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_prod  = tk.StringVar()
        conn0   = get_db()
        prods   = [r["name"] for r in conn0.execute("SELECT name FROM products ORDER BY name").fetchall()]
        conn0.close()
        prod_cb = ttk.Combobox(irow, textvariable=v_prod, values=prods, width=22)
        prod_cb.pack(side="left", padx=(4,8))

        tk.Label(irow, text="Qty:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_qty   = tk.StringVar(value="1")
        sr_qty_e = ttk.Entry(irow, textvariable=v_qty, width=7); sr_qty_e.pack(side="left", padx=(4,8))

        tk.Label(irow, text="Unit:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_unit  = tk.StringVar(value="Ltr")
        sr_unit_cb = ttk.Combobox(irow, textvariable=v_unit, width=6, state="readonly",
                     values=["Ltr","Kg","Pcs","Box","Set"]); sr_unit_cb.pack(side="left", padx=(4,8))

        tk.Label(irow, text="Rate:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_rate  = tk.StringVar(value="0")
        sr_rate_e = ttk.Entry(irow, textvariable=v_rate, width=9); sr_rate_e.pack(side="left", padx=(4,8))

        _sr_add_btn = make_btn(irow, "➕ Add", None, bg=C_ORANGE, pady=3); _sr_add_btn.pack(side="left", padx=(8,4))
        _sr_rem_btn = make_btn(irow, "🗑️ Remove", None, bg=C_RED, pady=3); _sr_rem_btn.pack(side="left", padx=4)

        sr_items = []
        total_lbl = tk.Label(form, text="Grand Total: ₹0.00",
                             font=("Segoe UI",11,"bold"), bg=C_BG, fg=C_ORANGE)
        total_lbl.pack(anchor="e", pady=2)

        # ── Items treeview ──
        itv = ttk.Treeview(form, columns=("Product","Qty","Unit","Rate","Total"),
                           show="headings", height=4)
        for c, w in zip(("Product","Qty","Unit","Rate","Total"), [220,70,70,90,100]):
            itv.heading(c, text=c); itv.column(c, width=w, anchor="center")
        itv.pack(fill="x", pady=(4,0))

        def refresh_total():
            t = sum(i["total"] for i in sr_items)
            total_lbl.config(text=f"Grand Total: {rupee(t)}")

        def prod_selected(ev=None):
            name = v_prod.get().strip()
            if not name: return
            conn2 = get_db()
            p = conn2.execute("SELECT sale_rate, unit FROM products WHERE name=?", (name,)).fetchone()
            conn2.close()
            if p:
                v_rate.set(str(p["sale_rate"]))
                v_unit.set(p["unit"])

        prod_cb.bind("<<ComboboxSelected>>", prod_selected)

        def bill_selected(ev=None):
            bill_no = v_orig.get().strip()
            if not bill_no: return
            conn2 = get_db()
            sale = conn2.execute("SELECT party FROM sales WHERE bill_no=?", (bill_no,)).fetchone()
            if sale:
                v_party.set(sale["party"])
                # Items prefill karo
                items = conn2.execute(
                    """SELECT si.product, si.qty, si.unit, si.rate
                       FROM sale_items si
                       JOIN sales s ON s.id = si.sale_id
                       WHERE s.bill_no=?""", (bill_no,)).fetchall()
                # Clear existing items
                sr_items.clear()
                itv.delete(*itv.get_children())
                for it in items:
                    total = round(it["qty"] * it["rate"], 2)
                    sr_items.append({"product": it["product"], "qty": it["qty"],
                                     "unit": it["unit"], "rate": it["rate"], "total": total})
                    itv.insert("","end", values=(it["product"], it["qty"], it["unit"],
                                                 rupee(it["rate"]), rupee(total)))
                refresh_total()
            conn2.close()

        orig_cb.bind("<<ComboboxSelected>>", bill_selected)
        orig_cb.bind("<FocusOut>", bill_selected)

        def add_sr_item():
            prod = v_prod.get().strip()
            if not prod: messagebox.showwarning("","Product chunein!"); return
            try: qty = float(v_qty.get())
            except: messagebox.showwarning("","Qty sahi daalen!"); return
            try: rate = float(v_rate.get())
            except: messagebox.showwarning("","Rate sahi daalen!"); return
            if qty <= 0: messagebox.showwarning("","Qty 0 se zyada honi chahiye!"); return
            total = round(qty * rate, 2)
            item = {"product": prod, "qty": qty, "unit": v_unit.get(), "rate": rate, "total": total}
            sr_items.append(item)
            itv.insert("", "end", values=(prod, qty, v_unit.get(), rupee(rate), rupee(total)))
            v_prod.set(""); v_qty.set("1"); v_rate.set("0")
            refresh_total()

        def remove_sr_item():
            sel = itv.selection()
            if not sel: return
            idx = itv.index(sel[0])
            if idx < len(sr_items):
                sr_items.pop(idx)
                itv.delete(sel[0])
                refresh_total()

        def save_sr():
            if not sr_items: messagebox.showwarning("","Koi item nahi!"); return
            rno_val = v_rno.get().strip()
            if not rno_val: messagebox.showwarning("","Return No daalen!"); return
            grand = sum(i["total"] for i in sr_items)
            conn2 = get_db()
            try:
                cur = conn2.execute("""INSERT INTO sale_returns
                    (return_no, return_date, orig_bill, party, reason, grand_total)
                    VALUES(?,?,?,?,?,?)""",
                    (rno_val, v_date.get(), v_orig.get().strip(),
                     v_party.get().strip(), v_reason.get(), grand))
                ret_id = cur.lastrowid
                for it in sr_items:
                    conn2.execute("""INSERT INTO sale_return_items
                        (return_id, product, qty, unit, rate, total)
                        VALUES(?,?,?,?,?,?)""",
                        (ret_id, it["product"], it["qty"], it["unit"], it["rate"], it["total"]))
                conn2.commit()
                messagebox.showinfo("Saved!", f"Sale Return {rno_val} save ho gaya!\nTotal: {rupee(grand)}")
                self._show_sale_return()
            except Exception as ex:
                conn2.rollback()
                messagebox.showerror("Error", str(ex))
            finally:
                conn2.close()

        btn_row = tk.Frame(form, bg=C_BG); btn_row.pack(fill="x", pady=(4,0))
        # Wire inline buttons
        _sr_add_btn.config(command=add_sr_item)
        _sr_rem_btn.config(command=remove_sr_item)
        make_btn(btn_row, "💾 Save Return", save_sr, bg=C_GREEN, pady=6).pack(side="left", padx=4)

        # Chain 1: Return No → Date → Orig Bill → Party → Product
        bind_enter_nav([sr_rno_e, sr_date_dp.entry, orig_cb, party_cb, prod_cb])
        # Chain 2: item entry row → Enter on rate = add item + focus back to product
        def _sr_item_nav(widgets_list):
            def _make(i):
                def _fn(ev):
                    if i == len(widgets_list) - 1:
                        add_sr_item()
                        prod_cb.focus_force()
                    else:
                        widgets_list[i+1].focus_force()
                        try: widgets_list[i+1].select_range(0, "end")
                        except: pass
                    return "break"
                return _fn
            for idx, w in enumerate(widgets_list):
                w.bind("<Return>",   _make(idx))
                w.bind("<KP_Enter>", _make(idx))
        _sr_item_nav([prod_cb, sr_qty_e, sr_unit_cb, sr_rate_e])

        # ── History ──
        tk.Label(main, text="Sale Return History", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w", pady=(8,2))
        hcols = ("Return No","Date","Orig Bill","Party","Reason","Total")
        htv = ttk.Treeview(main, columns=hcols, show="headings", height=6)
        for c, w in zip(hcols, [100,90,110,150,120,100]):
            htv.heading(c, text=c); htv.column(c, width=w, anchor="center")
        hsb = ttk.Scrollbar(main, orient="vertical", command=htv.yview)
        htv.configure(yscrollcommand=hsb.set)
        hsb.pack(side="right", fill="y"); htv.pack(fill="both", expand=True)

        def load_sr_history():
            htv.delete(*htv.get_children())
            conn2 = get_db()
            rows = conn2.execute("SELECT * FROM sale_returns ORDER BY return_date DESC, id DESC").fetchall()
            conn2.close()
            for r in rows:
                htv.insert("","end", values=(r["return_no"], r["return_date"],
                    r["orig_bill"], r["party"], r["reason"], rupee(r["grand_total"])))

        def delete_sr():
            sel = htv.selection()
            if not sel: return
            if not messagebox.askyesno("Delete","Ye return delete karein?"): return
            idx = htv.index(sel[0])
            conn2 = get_db()
            rows = conn2.execute("SELECT id FROM sale_returns ORDER BY return_date DESC, id DESC").fetchall()
            if idx < len(rows):
                conn2.execute("DELETE FROM sale_returns WHERE id=?", (rows[idx]["id"],))
                conn2.commit()
            conn2.close(); load_sr_history()

        make_btn(main, "🗑️ Delete Selected", delete_sr, bg=C_RED, pady=4).pack(anchor="w", pady=(4,0))
        load_sr_history()

    # ══════════════════════════════════════════════════════════════════════════
    # PURCHASE RETURN
    # ══════════════════════════════════════════════════════════════════════════
    def _show_purchase_return(self):
        self._clear()
        self._section_header("🔃 Purchase Return", "Supplier ko wapas kiya maal")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # ── Top: New Return form ──
        form = tk.LabelFrame(main, text="Naya Purchase Return", font=("Segoe UI",9,"bold"),
                             bg=C_BG, fg=C_DARK, padx=8, pady=6)
        form.pack(fill="x", pady=(0,8))

        row1 = tk.Frame(form, bg=C_BG); row1.pack(fill="x", pady=2)
        tk.Label(row1, text="Return No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        conn0 = get_db()
        prno = conn0.execute("SELECT COUNT(*) as c FROM purchase_returns").fetchone()["c"] + 1
        conn0.close()
        v_rno   = tk.StringVar(value=f"PR/{prno:04d}")
        pr_rno_e = ttk.Entry(row1, textvariable=v_rno, width=12); pr_rno_e.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Date:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_date  = tk.StringVar()
        pr_date_dp = make_date_entry(row1, v_date, width=12); pr_date_dp.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Orig Bill No:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_orig  = tk.StringVar()
        conn0   = get_db()
        pur_bills = [r["bill_no"] for r in conn0.execute(
            "SELECT bill_no FROM purchases ORDER BY bill_date DESC, id DESC").fetchall()]
        conn0.close()
        orig_cb2 = ttk.Combobox(row1, textvariable=v_orig, values=pur_bills, width=16)
        orig_cb2.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Supplier:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_party = tk.StringVar()
        supp_cb = ttk.Combobox(row1, textvariable=v_party, width=18, values=[])
        supp_cb.pack(side="left", padx=(4,16))

        tk.Label(row1, text="Reason:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_reason = tk.StringVar(value="Defective")
        ttk.Combobox(row1, textvariable=v_reason, width=14,
                     values=["Defective","Wrong Product","Excess Qty","Quality Issue","Other"]
                     ).pack(side="left", padx=(4,0))

        # ── Item entry row ──
        irow = tk.Frame(form, bg=C_BG); irow.pack(fill="x", pady=(6,2))
        tk.Label(irow, text="Product:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_prod  = tk.StringVar()
        conn0   = get_db()
        prods   = [r["name"] for r in conn0.execute("SELECT name FROM products ORDER BY name").fetchall()]
        conn0.close()
        prod_cb2 = ttk.Combobox(irow, textvariable=v_prod, values=prods, width=22)
        prod_cb2.pack(side="left", padx=(4,8))

        tk.Label(irow, text="Qty:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_qty   = tk.StringVar(value="1")
        pr_qty_e = ttk.Entry(irow, textvariable=v_qty, width=7); pr_qty_e.pack(side="left", padx=(4,8))

        tk.Label(irow, text="Unit:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_unit  = tk.StringVar(value="Ltr")
        pr_unit_cb = ttk.Combobox(irow, textvariable=v_unit, width=6, state="readonly",
                     values=["Ltr","Kg","Pcs","Box","Set"]); pr_unit_cb.pack(side="left", padx=(4,8))

        tk.Label(irow, text="Rate:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_rate  = tk.StringVar(value="0")
        pr_rate_e = ttk.Entry(irow, textvariable=v_rate, width=9); pr_rate_e.pack(side="left", padx=(4,8))

        _pr_add_btn = make_btn(irow, "➕ Add", None, bg=C_ORANGE, pady=3); _pr_add_btn.pack(side="left", padx=(8,4))
        _pr_rem_btn = make_btn(irow, "🗑️ Remove", None, bg=C_RED, pady=3); _pr_rem_btn.pack(side="left", padx=4)

        pr_items = []
        total_lbl2 = tk.Label(form, text="Grand Total: ₹0.00",
                              font=("Segoe UI",11,"bold"), bg=C_BG, fg=C_ORANGE)
        total_lbl2.pack(anchor="e", pady=2)

        itv2 = ttk.Treeview(form, columns=("Product","Qty","Unit","Rate","Total"),
                            show="headings", height=4)
        for c, w in zip(("Product","Qty","Unit","Rate","Total"), [220,70,70,90,100]):
            itv2.heading(c, text=c); itv2.column(c, width=w, anchor="center")
        itv2.pack(fill="x", pady=(4,0))

        def refresh_total2():
            t = sum(i["total"] for i in pr_items)
            total_lbl2.config(text=f"Grand Total: {rupee(t)}")

        def prod_selected2(ev=None):
            name = v_prod.get().strip()
            if not name: return
            conn2 = get_db()
            p = conn2.execute("SELECT purchase_rate, unit FROM products WHERE name=?", (name,)).fetchone()
            conn2.close()
            if p:
                v_rate.set(str(p["purchase_rate"]))
                v_unit.set(p["unit"])

        prod_cb2.bind("<<ComboboxSelected>>", prod_selected2)

        def bill_selected2(ev=None):
            bill_no = v_orig.get().strip()
            if not bill_no: return
            conn2 = get_db()
            pur = conn2.execute("SELECT party FROM purchases WHERE bill_no=?", (bill_no,)).fetchone()
            if pur:
                v_party.set(pur["party"])
                items = conn2.execute(
                    """SELECT pi.product, pi.qty, pi.unit, pi.rate
                       FROM purchase_items pi
                       JOIN purchases p ON p.id = pi.purchase_id
                       WHERE p.bill_no=?""", (bill_no,)).fetchall()
                pr_items.clear()
                itv2.delete(*itv2.get_children())
                for it in items:
                    total = round(it["qty"] * it["rate"], 2)
                    pr_items.append({"product": it["product"], "qty": it["qty"],
                                     "unit": it["unit"], "rate": it["rate"], "total": total})
                    itv2.insert("","end", values=(it["product"], it["qty"], it["unit"],
                                                  rupee(it["rate"]), rupee(total)))
                refresh_total2()
            conn2.close()

        orig_cb2.bind("<<ComboboxSelected>>", bill_selected2)
        orig_cb2.bind("<FocusOut>", bill_selected2)

        def add_pr_item():
            prod = v_prod.get().strip()
            if not prod: messagebox.showwarning("","Product chunein!"); return
            try: qty = float(v_qty.get())
            except: messagebox.showwarning("","Qty sahi daalen!"); return
            try: rate = float(v_rate.get())
            except: messagebox.showwarning("","Rate sahi daalen!"); return
            if qty <= 0: messagebox.showwarning("","Qty 0 se zyada honi chahiye!"); return
            total = round(qty * rate, 2)
            item = {"product": prod, "qty": qty, "unit": v_unit.get(), "rate": rate, "total": total}
            pr_items.append(item)
            itv2.insert("", "end", values=(prod, qty, v_unit.get(), rupee(rate), rupee(total)))
            v_prod.set(""); v_qty.set("1"); v_rate.set("0")
            refresh_total2()

        def remove_pr_item():
            sel = itv2.selection()
            if not sel: return
            idx = itv2.index(sel[0])
            if idx < len(pr_items):
                pr_items.pop(idx)
                itv2.delete(sel[0])
                refresh_total2()

        def save_pr():
            if not pr_items: messagebox.showwarning("","Koi item nahi!"); return
            rno_val = v_rno.get().strip()
            if not rno_val: messagebox.showwarning("","Return No daalen!"); return
            grand = sum(i["total"] for i in pr_items)
            conn2 = get_db()
            try:
                cur = conn2.execute("""INSERT INTO purchase_returns
                    (return_no, return_date, orig_bill, party, reason, grand_total)
                    VALUES(?,?,?,?,?,?)""",
                    (rno_val, v_date.get(), v_orig.get().strip(),
                     v_party.get().strip(), v_reason.get(), grand))
                ret_id = cur.lastrowid
                for it in pr_items:
                    conn2.execute("""INSERT INTO purchase_return_items
                        (return_id, product, qty, unit, rate, total)
                        VALUES(?,?,?,?,?,?)""",
                        (ret_id, it["product"], it["qty"], it["unit"], it["rate"], it["total"]))
                conn2.commit()
                messagebox.showinfo("Saved!", f"Purchase Return {rno_val} save ho gaya!\nTotal: {rupee(grand)}")
                self._show_purchase_return()
            except Exception as ex:
                conn2.rollback()
                messagebox.showerror("Error", str(ex))
            finally:
                conn2.close()

        btn_row2 = tk.Frame(form, bg=C_BG); btn_row2.pack(fill="x", pady=(4,0))
        # Wire inline buttons
        _pr_add_btn.config(command=add_pr_item)
        _pr_rem_btn.config(command=remove_pr_item)
        make_btn(btn_row2, "💾 Save Return", save_pr, bg=C_GREEN, pady=6).pack(side="left", padx=4)

        # Chain 1: Return No → Date → Orig Bill → Supplier → Product
        bind_enter_nav([pr_rno_e, pr_date_dp.entry, orig_cb2, supp_cb, prod_cb2])
        # Chain 2: item entry row → Enter on rate = add item + focus back to product
        def _pr_item_nav(widgets_list):
            def _make(i):
                def _fn(ev):
                    if i == len(widgets_list) - 1:
                        add_pr_item()
                        prod_cb2.focus_force()
                    else:
                        widgets_list[i+1].focus_force()
                        try: widgets_list[i+1].select_range(0, "end")
                        except: pass
                    return "break"
                return _fn
            for idx, w in enumerate(widgets_list):
                w.bind("<Return>",   _make(idx))
                w.bind("<KP_Enter>", _make(idx))
        _pr_item_nav([prod_cb2, pr_qty_e, pr_unit_cb, pr_rate_e])

        # ── History ──
        tk.Label(main, text="Purchase Return History", font=("Segoe UI",10,"bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w", pady=(8,2))
        hcols2 = ("Return No","Date","Orig Bill","Supplier","Reason","Total")
        htv2 = ttk.Treeview(main, columns=hcols2, show="headings", height=6)
        for c, w in zip(hcols2, [100,90,110,150,120,100]):
            htv2.heading(c, text=c); htv2.column(c, width=w, anchor="center")
        hsb2 = ttk.Scrollbar(main, orient="vertical", command=htv2.yview)
        htv2.configure(yscrollcommand=hsb2.set)
        hsb2.pack(side="right", fill="y"); htv2.pack(fill="both", expand=True)

        def load_pr_history():
            htv2.delete(*htv2.get_children())
            conn2 = get_db()
            rows = conn2.execute("SELECT * FROM purchase_returns ORDER BY return_date DESC, id DESC").fetchall()
            conn2.close()
            for r in rows:
                htv2.insert("","end", values=(r["return_no"], r["return_date"],
                    r["orig_bill"], r["party"], r["reason"], rupee(r["grand_total"])))

        def delete_pr():
            sel = htv2.selection()
            if not sel: return
            if not messagebox.askyesno("Delete","Ye return delete karein?"): return
            idx = htv2.index(sel[0])
            conn2 = get_db()
            rows = conn2.execute("SELECT id FROM purchase_returns ORDER BY return_date DESC, id DESC").fetchall()
            if idx < len(rows):
                conn2.execute("DELETE FROM purchase_returns WHERE id=?", (rows[idx]["id"],))
                conn2.commit()
            conn2.close(); load_pr_history()

        make_btn(main, "🗑️ Delete Selected", delete_pr, bg=C_RED, pady=4).pack(anchor="w", pady=(4,0))
        load_pr_history()


    def _show_expenses(self):
        self._clear()
        self._section_header("💸 Expenses", "Shop ke kharche record karein")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=12, pady=8)

        # Add expense form
        form = tk.LabelFrame(main, text="Naya Kharcha", font=("Segoe UI",9,"bold"),
                             bg=C_BG, fg=C_DARK, padx=8, pady=8)
        form.pack(fill="x", pady=(0,8))

        rf = tk.Frame(form, bg=C_BG); rf.pack(fill="x")
        tk.Label(rf, text="Date:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_date = tk.StringVar()
        de = make_date_entry(rf, v_date, width=12, bg=C_BG); de.pack(side="left", padx=(4,16))

        tk.Label(rf, text="Category:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_cat = tk.StringVar(value="Rent")
        cat_cb = ttk.Combobox(rf, textvariable=v_cat, width=14,
                     values=["Rent","Salary","Electricity","Transport","Repair","Office","Other"])
        cat_cb.pack(side="left", padx=(4,16))

        tk.Label(rf, text="Amount ₹:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_amt = tk.StringVar(value="0")
        amt_e = ttk.Entry(rf, textvariable=v_amt, width=10); amt_e.pack(side="left", padx=(4,16))

        tk.Label(rf, text="Payment:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_pay = tk.StringVar(value="Cash")
        ttk.Combobox(rf, textvariable=v_pay, width=10, state="readonly",
                     values=["Cash","Online/UPI","Cheque"]).pack(side="left", padx=(4,16))

        rf2 = tk.Frame(form, bg=C_BG); rf2.pack(fill="x", pady=(6,0))
        tk.Label(rf2, text="Description:", font=("Segoe UI",9), bg=C_BG).pack(side="left")
        v_desc = tk.StringVar()
        desc_e = ttk.Entry(rf2, textvariable=v_desc, width=40); desc_e.pack(side="left", padx=(4,0))

        def save_exp():
            try:
                amt = float(v_amt.get() or 0)
                if amt <= 0: messagebox.showwarning("","Amount daalen!"); return
            except: messagebox.showwarning("","Amount sahi daalen!"); return
            conn = get_db()
            try:
                conn.execute("INSERT INTO expenses (exp_date,category,description,amount,pay_mode) VALUES(?,?,?,?,?)",
                             (v_date.get(), v_cat.get(), v_desc.get().strip(), amt, v_pay.get()))
                conn.commit(); conn.close()
                v_amt.set("0"); v_desc.set("")
                load_exps()
            except Exception as ex:
                conn.close(); messagebox.showerror("Error", str(ex))

        save_btn = make_btn(rf2, "💾 Save", save_exp, bg=C_GREEN, pady=6)
        save_btn.pack(side="left", padx=8)
        bind_enter_nav([de.entry, cat_cb, amt_e, v_pay, desc_e, save_btn])

        # List
        cols = ("Date","Category","Description","Amount","Payment")
        tv = ttk.Treeview(main, columns=cols, show="headings")
        for c, w in zip(cols, [100,100,260,100,90]):
            tv.heading(c, text=c); tv.column(c, width=w, anchor="center")
        sb = ttk.Scrollbar(main, orient="vertical", command=tv.yview)
        tv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); tv.pack(fill="both", expand=True)

        total_lbl = tk.Label(main, text="", font=("Segoe UI",10,"bold"), bg=C_BG, fg=C_ORANGE)
        total_lbl.pack(anchor="e", pady=4)

        def load_exps():
            tv.delete(*tv.get_children())
            conn = get_db()
            rows = conn.execute("SELECT * FROM expenses ORDER BY exp_date DESC, id DESC").fetchall()
            conn.close()
            total = sum(r["amount"] for r in rows)
            total_lbl.config(text=f"Total Expenses: {rupee(total)}")
            for r in rows:
                tv.insert("","end", values=(r["exp_date"], r["category"], r["description"],
                          rupee(r["amount"]), r["pay_mode"]))

        def delete_exp():
            sel = tv.selection()
            if not sel: return
            if not messagebox.askyesno("Delete","Ye kharcha delete karein?"): return
            idx = tv.index(sel[0])
            conn = get_db()
            rows = conn.execute("SELECT id FROM expenses ORDER BY exp_date DESC, id DESC").fetchall()
            if idx < len(rows):
                conn.execute("DELETE FROM expenses WHERE id=?", (rows[idx]["id"],))
                conn.commit()
            conn.close(); load_exps()

        make_btn(main, "🗑️ Delete Selected", delete_exp, bg=C_RED).pack(anchor="w", pady=(4,0))
        load_exps()

    # ── SETTINGS ──────────────────────────────────────────────────────────────
    def _show_settings(self):
        self._clear()
        self._section_header("🔧  Settings", "Shop details & configuration")
        main = tk.Frame(self.content, bg=C_BG); main.pack(fill="both", expand=True, padx=16, pady=12)

        nb = ttk.Notebook(main); nb.pack(fill="both", expand=True)

        # ── Shop Settings ──
        tab1 = tk.Frame(nb, bg=C_BG); nb.add(tab1, text="🏪 Shop Details")
        form1 = tk.Frame(tab1, bg=C_BG, padx=20, pady=16); form1.pack(fill="x")

        conn = get_db()
        rows = {r[0]: r[1] for r in conn.execute("SELECT key,value FROM settings WHERE key LIKE 'shop_%' OR key IN ('bill_prefix')").fetchall()}
        conn.close()

        shop_fields = {}
        for label, key, width in [
            ("Shop Name *","shop_name",30), ("Address","shop_address",36),
            ("City","shop_city",20), ("State","shop_state",20),
            ("GSTIN","shop_gstin",20), ("Mobile","shop_mobile",16),
            ("Email","shop_email",28), ("Bank Name","shop_bank",24),
            ("Account No","shop_account",20), ("IFSC","shop_ifsc",14),
            ("UPI ID","shop_upi",24), ("Bill Prefix","bill_prefix",8),
        ]:
            f = tk.Frame(form1, bg=C_BG); f.pack(fill="x", pady=4)
            tk.Label(f, text=label+":", font=("Segoe UI",9), bg=C_BG, width=16, anchor="w").pack(side="left")
            var = tk.StringVar(value=rows.get(key,""))
            ttk.Entry(f, textvariable=var, width=width).pack(side="left")
            shop_fields[key] = var

        f_tnc = tk.Frame(form1, bg=C_BG); f_tnc.pack(fill="x", pady=4)
        tk.Label(f_tnc, text="Terms & Cond.:", font=("Segoe UI",9), bg=C_BG, width=16, anchor="w").pack(side="left")
        v_tnc = tk.StringVar(value=rows.get("shop_print_tnc",""))
        ttk.Entry(f_tnc, textvariable=v_tnc, width=46).pack(side="left")
        shop_fields["shop_print_tnc"] = v_tnc

        def save_shop():
            conn2 = get_db()
            for k, var in shop_fields.items():
                conn2.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (k, var.get()))
            conn2.commit(); conn2.close()
            messagebox.showinfo("Saved!","Settings save ho gayi!")
            self._rebuild_header()

        shop_save_btn = make_btn(tab1, "💾 Save Settings", save_shop, bg=C_GREEN, pady=10)
        shop_save_btn.pack(pady=8, padx=20, anchor="w")

        # Enter nav: all shop detail Entry fields → Save Settings button
        shop_entries = [w for f in form1.winfo_children()
                        for w in f.winfo_children()
                        if isinstance(w, ttk.Entry)]
        bind_enter_nav(shop_entries + [shop_save_btn])

        # ── User Settings ──
        tab2 = tk.Frame(nb, bg=C_BG); nb.add(tab2, text="🔐 Change Password")
        f2 = tk.Frame(tab2, bg=C_BG, padx=24, pady=20); f2.pack(fill="x")
        tk.Label(f2, text="Purana Password:", font=("Segoe UI",10), bg=C_BG).pack(anchor="w")
        v_old_pw = tk.StringVar()
        old_pw_entry = ttk.Entry(f2, textvariable=v_old_pw, show="*", width=24)
        old_pw_entry.pack(anchor="w", pady=(4,12))
        tk.Label(f2, text="Naya Password:", font=("Segoe UI",10), bg=C_BG).pack(anchor="w")
        v_pw = tk.StringVar()
        pw_entry = ttk.Entry(f2, textvariable=v_pw, show="*", width=24)
        pw_entry.pack(anchor="w", pady=(4,12))
        tk.Label(f2, text="Naya Password Confirm:", font=("Segoe UI",10), bg=C_BG).pack(anchor="w")
        v_pw2 = tk.StringVar()
        pw2_entry = ttk.Entry(f2, textvariable=v_pw2, show="*", width=24)
        pw2_entry.pack(anchor="w", pady=(4,12))
        def change_pw():
            old_pw = v_old_pw.get().strip()
            pw     = v_pw.get().strip()
            pw2    = v_pw2.get().strip()
            if not old_pw: messagebox.showwarning("","Purana password daalen!"); return
            if not pw:     messagebox.showwarning("","Naya password daalen!"); return
            if pw != pw2:  messagebox.showwarning("","Naya password match nahi kar raha!"); return
            conn3 = get_db()
            user = conn3.execute("SELECT password FROM users WHERE username='admin'").fetchone()
            if not user or user["password"] != old_pw:
                conn3.close()
                messagebox.showerror("Error","Purana password galat hai!"); return
            conn3.execute("UPDATE users SET password=? WHERE username='admin'", (pw,))
            conn3.commit(); conn3.close()
            messagebox.showinfo("Done!","Password change ho gaya!")
            v_old_pw.set(""); v_pw.set(""); v_pw2.set("")
        change_pw_btn = make_btn(f2, "🔐 Change Password", change_pw, bg=C_ORANGE, pady=10)
        change_pw_btn.pack(anchor="w")
        bind_enter_nav([old_pw_entry, pw_entry, pw2_entry, change_pw_btn])

        # ── License ──
        tab3 = tk.Frame(nb, bg=C_BG); nb.add(tab3, text="🔑 License")
        f3 = tk.Frame(tab3, bg=C_BG, padx=24, pady=20); f3.pack(fill="x")
        status, days_left, install_date, customer = _get_license_info()
        conn4 = get_db()
        lic_rows = {r[0]: r[1] for r in conn4.execute(
            "SELECT key,value FROM settings WHERE key IN ('license_expiry','license_customer','license_serial')").fetchall()}
        conn4.close()
        tk.Label(f3, text=f"Status: {'✅ Active' if status=='ok' else '❌ '+status.upper()}",
                 font=("Segoe UI",12,"bold"), bg=C_BG, fg=C_GREEN if status=='ok' else C_RED).pack(anchor="w")
        tk.Label(f3, text=f"Customer: {lic_rows.get('license_customer','')}",
                 font=("Segoe UI",10), bg=C_BG).pack(anchor="w", pady=4)
        tk.Label(f3, text=f"Expiry: {lic_rows.get('license_expiry','')} ({days_left} din baaki)",
                 font=("Segoe UI",10), bg=C_BG).pack(anchor="w")
        tk.Label(f3, text=f"Serial: {lic_rows.get('license_serial','')}",
                 font=("Courier New",10), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=4)

        # ── Backup & Restore ──
        tab4 = tk.Frame(nb, bg=C_BG); nb.add(tab4, text="💾 Backup & Restore")
        # Scrollable canvas so full content visible on small screens
        _canvas4 = tk.Canvas(tab4, bg=C_BG, highlightthickness=0)
        _sb4     = tk.Scrollbar(tab4, orient="vertical", command=_canvas4.yview)
        _canvas4.configure(yscrollcommand=_sb4.set)
        _sb4.pack(side="right", fill="y")
        _canvas4.pack(side="left", fill="both", expand=True)
        f4 = tk.Frame(_canvas4, bg=C_BG, padx=24, pady=20)
        _win4 = _canvas4.create_window((0, 0), window=f4, anchor="nw")
        def _on_f4_configure(e):
            _canvas4.configure(scrollregion=_canvas4.bbox("all"))
        def _on_canvas4_resize(e):
            _canvas4.itemconfig(_win4, width=e.width)
        f4.bind("<Configure>", _on_f4_configure)
        _canvas4.bind("<Configure>", _on_canvas4_resize)
        def _on_mousewheel4(e):
            _canvas4.yview_scroll(int(-1*(e.delta/120)), "units")
        _canvas4.bind("<Enter>", lambda e: _canvas4.bind_all("<MouseWheel>", _on_mousewheel4))
        _canvas4.bind("<Leave>", lambda e: _canvas4.unbind_all("<MouseWheel>"))

        # ── Manual Backup section ──
        tk.Label(f4, text="📦 Manual Backup", font=("Segoe UI",11,"bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w")
        tk.Label(f4, text="Aapka saara data (bills, parties, products) ek .db file me save ho jaata hai.",
                 font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(2,8))

        def manual_backup():
            do_backup(silent=False)
            _refresh_backup_list()

        make_btn(f4, "💾 Abhi Backup Karo", manual_backup, bg=C_GREEN, pady=8).pack(anchor="w")

        def open_backup_folder():
            folder = _backup_dir()
            try:
                import subprocess
                subprocess.Popen(["explorer", folder])   # Windows
            except Exception:
                try:
                    import subprocess
                    subprocess.Popen(["xdg-open", folder])  # Linux fallback
                except Exception:
                    messagebox.showinfo("Backup Folder", folder)
        make_btn(f4, "📂 Backup Folder Kholein", open_backup_folder, bg=C_BLUE, pady=8).pack(anchor="w", pady=(6,0))

        tk.Frame(f4, bg=C_GRAY, height=1).pack(fill="x", pady=16)

        # ── Auto-backup notice ──
        tk.Label(f4, text="⚙️ Auto Backup", font=("Segoe UI",11,"bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w")
        tk.Label(f4, text="✅ Software band karne par automatic backup hota hai (last 30 backups rakhe jaate hain).",
                 font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(2,16))

        tk.Frame(f4, bg=C_GRAY, height=1).pack(fill="x", pady=(0,16))

        # ── Google Drive Backup section ──
        tk.Label(f4, text="☁️ Google Drive Backup", font=("Segoe UI",11,"bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w")
        tk.Label(f4,
                 text="Apne PC ka Google Drive sync folder set karo — backup wahan bhi automatically save hoga.",
                 font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(2,6))

        drive_row = tk.Frame(f4, bg=C_BG); drive_row.pack(fill="x", pady=(0,4))
        v_drive_path = tk.StringVar(value=get_drive_path())
        drive_entry = ttk.Entry(drive_row, textvariable=v_drive_path, width=46)
        drive_entry.pack(side="left", padx=(0,6))

        def browse_drive():
            folder = filedialog.askdirectory(
                title="Google Drive Sync Folder Chunein",
                initialdir=v_drive_path.get() or os.path.expanduser("~")
            )
            if folder:
                v_drive_path.set(folder)
                set_drive_path(folder)
                _update_drive_status()

        make_btn(drive_row, "📂 Browse", browse_drive, bg=C_BLUE, pady=4).pack(side="left", padx=(0,4))

        def save_drive_path():
            p = v_drive_path.get().strip()
            set_drive_path(p)
            _update_drive_status()
            if p and not os.path.isdir(p):
                messagebox.showwarning("⚠️ Folder Nahi Mila", f"Ye folder exist nahi karta:\n{p}")
            elif p:
                messagebox.showinfo("✅ Saved", f"Drive folder set ho gaya:\n{p}")

        make_btn(drive_row, "💾 Save", save_drive_path, bg=C_GREEN, pady=4).pack(side="left")

        drive_status_lbl = tk.Label(f4, text="", font=("Segoe UI",9), bg=C_BG)
        drive_status_lbl.pack(anchor="w", pady=(2,6))

        def _update_drive_status():
            p = v_drive_path.get().strip()
            if not p:
                drive_status_lbl.config(text="⚠️  Drive folder set nahi hai", fg="#E65100")
            elif not os.path.isdir(p):
                drive_status_lbl.config(text=f"❌  Folder nahi mila: {p}", fg=C_RED)
            else:
                # Count existing backups there
                existing = [f for f in os.listdir(p)
                            if f.startswith("bhugtanease_paint_") and f.endswith(".db")]
                drive_status_lbl.config(
                    text=f"✅  Drive folder set hai  |  {len(existing)} backup(s) present",
                    fg=C_GREEN)

        _update_drive_status()

        drive_btn_row = tk.Frame(f4, bg=C_BG); drive_btn_row.pack(anchor="w", pady=(2,0))
        def manual_drive_backup():
            do_drive_backup(silent=False)
            _update_drive_status()
        make_btn(drive_btn_row, "☁️ Abhi Drive Backup Karo", manual_drive_backup, bg=C_TEAL, pady=8).pack(side="left", padx=(0,8))

        tk.Label(f4, text="(Auto: software band karne par automatically Drive me bhi backup hoga — agar folder set ho)",
                 font=("Segoe UI",8), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(4,0))

        tk.Frame(f4, bg=C_GRAY, height=1).pack(fill="x", pady=16)

        # ── Restore section ──
        tk.Label(f4, text="🔁 Restore", font=("Segoe UI",11,"bold"),
                 bg=C_BG, fg=C_DARK).pack(anchor="w")
        tk.Label(f4, text="Kisi purani backup se data wapas laane ke liye niche se file chunein:",
                 font=("Segoe UI",9), bg=C_BG, fg=C_GRAY).pack(anchor="w", pady=(2,8))

        # Backup list frame
        list_frame = tk.Frame(f4, bg=C_BG); list_frame.pack(fill="x", pady=(0,8))
        scrollbar = tk.Scrollbar(list_frame, orient="vertical")
        backup_lb = tk.Listbox(list_frame, height=8, font=("Courier New",9),
                               yscrollcommand=scrollbar.set, selectmode="single",
                               bg="white", activestyle="dotbox")
        scrollbar.config(command=backup_lb.yview)
        backup_lb.pack(side="left", fill="x", expand=True)
        scrollbar.pack(side="right", fill="y")

        def _refresh_backup_list():
            backup_lb.delete(0, "end")
            folder = _backup_dir()
            files = sorted([
                f for f in os.listdir(folder) if f.endswith(".db")
            ], reverse=True)   # latest first
            for f in files:
                fpath = os.path.join(folder, f)
                size_kb = os.path.getsize(fpath) // 1024
                mtime   = datetime.datetime.fromtimestamp(os.path.getmtime(fpath))
                display = f"  {mtime.strftime('%d-%m-%Y  %H:%M:%S')}   ({size_kb} KB)   {f}"
                backup_lb.insert("end", display)
            if not files:
                backup_lb.insert("end", "  (Koi backup nahi mila — pehle backup karo)")

        _refresh_backup_list()

        def do_restore():
            sel = backup_lb.curselection()
            if not sel:
                messagebox.showwarning("Select Karo", "Pehle restore karne wali backup file chunein!")
                return
            # Extract filename from display string
            line = backup_lb.get(sel[0]).strip()
            fname = line.split("   ")[-1].strip()
            if not fname.endswith(".db"):
                return
            src = os.path.join(_backup_dir(), fname)
            if not os.path.exists(src):
                messagebox.showerror("Error", f"File nahi mili:\n{src}"); return
            confirm = messagebox.askyesno(
                "⚠️ Confirm Restore",
                f"Kya aap sure hain?\n\n"
                f"Backup: {fname}\n\n"
                f"⚠️ Ye restore karne se aaj ka saara naya data OVERWRITE ho jaayega!\n\n"
                f"Haan — restore karo\nNa — cancel karo"
            )
            if not confirm: return
            # Take a safety backup of current data first
            safety = do_backup(silent=True)
            try:
                shutil.copy2(src, DB_FILE)
                messagebox.showinfo(
                    "✅ Restore Complete",
                    f"Data restore ho gaya!\n\n"
                    f"Restore se: {fname}\n"
                    f"(Aapka purana data bhi backup hua hai:\n{os.path.basename(safety) if safety else 'N/A'})\n\n"
                    f"Software restart karo taaki sab sahi load ho."
                )
                _refresh_backup_list()
            except Exception as e:
                messagebox.showerror("❌ Restore Failed", f"Restore nahi hua:\n{e}")

        def browse_and_restore():
            path = filedialog.askopenfilename(
                title="Backup File Chunein",
                filetypes=[("SQLite Backup", "*.db"), ("All Files", "*.*")],
                initialdir=_backup_dir()
            )
            if not path: return
            confirm = messagebox.askyesno(
                "⚠️ Confirm Restore",
                f"Kya aap sure hain?\n\n"
                f"File: {os.path.basename(path)}\n\n"
                f"⚠️ Ye restore karne se aaj ka saara naya data OVERWRITE ho jaayega!"
            )
            if not confirm: return
            safety = do_backup(silent=True)
            try:
                shutil.copy2(path, DB_FILE)
                messagebox.showinfo(
                    "✅ Restore Complete",
                    f"Data restore ho gaya!\n\n"
                    f"(Purana data safety backup: {os.path.basename(safety) if safety else 'N/A'})\n\n"
                    f"Software restart karo taaki sab sahi load ho."
                )
                _refresh_backup_list()
            except Exception as e:
                messagebox.showerror("❌ Restore Failed", f"Restore nahi hua:\n{e}")

        def drive_restore():
            dp = get_drive_path()
            if not dp or not os.path.isdir(dp):
                messagebox.showwarning("Drive Folder Nahi Mila",
                    "Pehle Drive folder set karo (upar Drive section me).")
                return
            drive_files = sorted([
                f for f in os.listdir(dp)
                if f.startswith("bhugtanease_paint_") and f.endswith(".db")
            ], reverse=True)
            if not drive_files:
                messagebox.showinfo("Koi File Nahi", f"Drive folder me koi backup nahi mila:\n{dp}")
                return
            # Show a simple picker dialog
            pick = tk.Toplevel(self.root)
            pick.title("☁️ Drive se Restore")
            pick.configure(bg=C_BG)
            pick.geometry("520x320")
            pick.grab_set()
            tk.Label(pick, text="☁️ Drive se Restore",
                     font=("Segoe UI",12,"bold"), bg=C_BG).pack(pady=(12,4))
            tk.Label(pick, text=f"Folder: {dp}",
                     font=("Segoe UI",8), bg=C_BG, fg=C_GRAY).pack()
            tk.Label(pick, text="Kaunsi backup restore karni hai?",
                     font=("Segoe UI",9), bg=C_BG).pack(pady=(8,4))
            lb_frame = tk.Frame(pick, bg=C_BG); lb_frame.pack(fill="x", padx=16)
            sb2 = tk.Scrollbar(lb_frame)
            lb2 = tk.Listbox(lb_frame, height=8, font=("Courier New",9),
                              yscrollcommand=sb2.set, selectmode="single", bg="white")
            sb2.config(command=lb2.yview)
            lb2.pack(side="left", fill="x", expand=True); sb2.pack(side="right", fill="y")
            for f in drive_files:
                fpath = os.path.join(dp, f)
                size_kb = os.path.getsize(fpath) // 1024
                mtime = datetime.datetime.fromtimestamp(os.path.getmtime(fpath))
                lb2.insert("end", f"  {mtime.strftime('%d-%m-%Y  %H:%M:%S')}  ({size_kb} KB)  {f}")
            lb2.selection_set(0)

            def confirm_drive_restore():
                sel = lb2.curselection()
                if not sel: return
                line = lb2.get(sel[0]).strip()
                fname = line.split("  ")[-1].strip()
                src = os.path.join(dp, fname)
                pick.destroy()
                if not os.path.exists(src):
                    messagebox.showerror("Error", f"File nahi mili:\n{src}"); return
                ok = messagebox.askyesno("⚠️ Confirm",
                    f"Drive backup restore karein?\n\nFile: {fname}\n\n"
                    f"⚠️ Aaj ka data OVERWRITE ho jaayega!")
                if not ok: return
                safety = do_backup(silent=True)
                try:
                    shutil.copy2(src, DB_FILE)
                    messagebox.showinfo("✅ Restore Complete",
                        f"Drive se restore ho gaya!\n\nFile: {fname}\n"
                        f"(Safety backup: {os.path.basename(safety) if safety else 'N/A'})\n\n"
                        f"Software restart karo.")
                    _refresh_backup_list()
                except Exception as e:
                    messagebox.showerror("❌ Failed", f"Restore nahi hua:\n{e}")

            btn_f = tk.Frame(pick, bg=C_BG); btn_f.pack(pady=10)
            make_btn(btn_f, "☁️ Is File se Restore Karo", confirm_drive_restore, bg=C_TEAL, pady=8).pack(side="left", padx=8)
            make_btn(btn_f, "Cancel", pick.destroy, bg=C_GRAY, pady=8).pack(side="left")

        btn_row4 = tk.Frame(f4, bg=C_BG); btn_row4.pack(fill="x", pady=4)
        make_btn(btn_row4, "🔁 Local Backup se Restore", do_restore, bg=C_ORANGE, pady=8).pack(side="left", padx=(0,8))
        make_btn(btn_row4, "☁️ Drive se Restore", drive_restore, bg=C_TEAL, pady=8).pack(side="left", padx=(0,8))
        make_btn(btn_row4, "📂 Kisi Bhi File se...", browse_and_restore, bg=C_GRAY, pady=8).pack(side="left")

    def _rebuild_header(self):
        self._hdr_logo_ref = None  # GC allow karo purana logo
        for w in self.root.winfo_children():
            if isinstance(w, tk.Frame) and w.cget("bg") == C_ORANGE:
                w.destroy(); break
        self._build_header()

# ─── MAIN ENTRY POINT ─────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    init_db()

    status, days_left, install_date, customer = _get_license_info()

    if status == 'new':
        _show_activate_window()
        status, _, _, _ = _get_license_info()
        if status != 'ok': sys.exit(0)
    elif status == 'expired':
        LicenseExpiredWin(days_left, customer)
        status, _, _, _ = _get_license_info()
        if status != 'ok': sys.exit(0)
    elif status != 'ok':
        sys.exit(0)

    LoginWin()
